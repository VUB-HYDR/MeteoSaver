import os
import openpyxl
from openpyxl.styles import PatternFill
import shutil
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from statistics import mean

# Setting up the current working directory; for both the input and output folders
cwd = os.getcwd()



def is_string_convertible_to_float(value):
    '''
    Determines if a given value can be converted to a float.

    This function is used to verify if a value in a cell, such as one from an MS Excel sheet, can be converted to a float. 
    It ensures that the value can be used in subsequent numerical calculations without causing errors.

    Parameters
    --------------
    value: Any
        The value from a cell in an MS Excel sheet. This can be of any data type, including string, integer, float, or None.

    Returns
    --------------
    bool
        Returns True if the value can be converted to a float, indicating that it is safe for numerical calculations.
        Returns False if the value cannot be converted to a float or if the value is None (e.g., for empty cells).
    '''

    if value is None: # Check to handle None cases (empty cells)
        return False
    try:
        float(value)
        return True
    except ValueError:
        return False




def highlight_change(color, worksheet_and_cell_coordinate, filename):
    '''
    Highlights a cell in an Excel worksheet to indicate the status of data processing or transcription.

    This function applies a color highlight to a specific cell in an Excel worksheet to visually represent one of the following scenarios:
    (1) A correction has been made during data post processing (QA/QC checks) to fix an error.
    (2) A value has been confirmed as correctly transcribed.
    (3) A value has been identified as wrongly transcribed.

    Parameters
    --------------
    color: str
        The color to be used for highlighting, which corresponds to a specific status as detailed in the "Key_for_post_processed_data_sheets" document in docs (in the repository).
    worksheet_and_cell_coordinate: openpyxl.cell.Cell
        The specific cell in the worksheet that needs to be highlighted.
    filename: str
        The name or path of the Excel file containing the worksheet.

    Returns
    --------------
    None
        The function applies the highlight directly to the Excel file and does not return any value. The specified cell in the Excel file will be highlighted according to the provided color.
    '''
    
    # Highlight cells with strings instead of floats
    highlighting_color = color # Highlighting color of choice
    highlighting_strings = PatternFill(start_color = highlighting_color, end_color = highlighting_color, fill_type = 'solid')
    cell_to_highlight = worksheet_and_cell_coordinate
    cell_to_highlight.fill = highlighting_strings



def is_highlighted(cell, color):
    '''
    Checks if a cell is highlighted with a specific color during earlier post-processing steps.

    This function is used to determine whether a particular cell in an Excel worksheet has been highlighted with a specified color. 
    This check is essential for identifying cells that may need corrections based on the errors identified during earlier post-processing.

    Parameters
    --------------   
    cell: openpyxl.cell.Cell
        The cell in the worksheet to check for highlighting.
    color: str
        The color (in RGB or hex format) to compare against the cell's current highlight color.

    Returns
    -------------- 
    bool
        Returns True if the cell is highlighted with the specified color, indicating that it was marked during post-processing.
        Returns False if the cell is not highlighted with the specified color.
    '''


    fill = cell.fill.start_color
    if isinstance(fill, openpyxl.styles.colors.Color):
        return fill.rgb == color
    return False



def merge_excel_files(file1, file2, output_file, start_row, end_row):
    '''
    Merges two Excel files for verification purposes: one organized by the mid-point coordinates of bounding boxes and the other by the top coordinates.

    This function merges two preprocessed Excel files that contain transcribed data organized differently (one by mid-point and the other by top coordinates of bounding boxes). 
    The merged output file allows for cross-checking to ensure that cells are correctly placed in their respective rows.

    Parameters
    --------------
    file1: str
        The path to the Excel file containing transcribed data organized in rows using the top coordinates of the bounding boxes (contours).
    file2: str
        The path to the Excel file containing transcribed data organized in rows using the mid-point coordinates of the bounding boxes (contours).
    output_file: str
        The path where the merged Excel file will be saved.
    start_row: int
        The starting row number from which to begin the merge.
    end_row: int
        The ending row number up to which the merge should be conducted.

    Returns
    --------------
    None
        The function creates and saves a merged Excel file at the specified `output_file` location. This file combines the data from `file1` and `file2` for further verification.
    '''


    # Load the Excel files into DataFrames without headers
    df1 = pd.read_excel(file1, header=None)
    df2 = pd.read_excel(file2, header=None)


    # If the indices are not simple integers or do not align with Excel rows as expected,
    # you might need to reset them or adjust how the Excel file is being read (e.g., `index_col=None`)
    df1 = df1.reset_index(drop=True)
    df2 = df2.reset_index(drop=True)

    # Convert start_row and end_row to zero-based index for Python
    start_idx = start_row - 1  # Convert 1-based index to 0-based
    end_idx = end_row - 1  # Convert 1-based index to 0-based

    # Slice to only include the range from start_idx to end_idx
    df1 = df1.iloc[start_idx:end_idx + 1]
    df2 = df2.iloc[start_idx:end_idx + 1]

    # Initialize a new DataFrame to hold merged results
    merged_df = pd.DataFrame(index=df1.index, columns=df1.columns)

    # Iterate over rows by index (assuming the indices are aligned)
    for idx in df1.index:
        for col in df1.columns:
            val1 = df1.at[idx, col]
            val2 = df2.at[idx, col]
            # Simple merge logic: prefer non-empty values from df1, then df2
            if pd.notna(val1):
                merged_df.at[idx, col] = val1
            else:
                merged_df.at[idx, col] = val2


    # Create a new workbook and select the active worksheet
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active

    # Append the merged DataFrame to the new worksheet
    for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            new_worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Merge cells for multi-column headers
    new_worksheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1) #No de la pentade
    new_worksheet.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2) #Date
    new_worksheet.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3) #Bellani
    new_worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8) #Températures extrêmes
    new_worksheet.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10) #Evaportation
    new_worksheet.merge_cells(start_row=1, start_column=11, end_row=3, end_column=11) #Pluies
    new_worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=16) #Température et Humidité de l'air à 6 heures
    new_worksheet.merge_cells(start_row=1, start_column=17, end_row=1, end_column=21) #Température et Humidité de l'air à 15 heures
    new_worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=26) #Température et Humidité de l'air à 18 heures
    new_worksheet.merge_cells(start_row=1, start_column=27, end_row=3, end_column=27) #Date
    # subheaders
    new_worksheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7) #Abri
    new_worksheet.merge_cells(start_row=2, start_column=9, end_row=2, end_column=10) #Piche
    new_worksheet.merge_cells(start_row=2, start_column=12, end_row=2, end_column=16) #(Psychromètre a aspiration)
    new_worksheet.merge_cells(start_row=2, start_column=17, end_row=2, end_column=21) #(Psychromètre a aspiration)
    new_worksheet.merge_cells(start_row=2, start_column=22, end_row=2, end_column=26) #(Psychromètre a aspiration)

    # Set up border styles for excel output
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    # Loop through cells to apply borders
    for row in new_worksheet.iter_rows(min_row=1, max_row=new_worksheet.max_row, min_col=1, max_col=new_worksheet.max_column):
        for cell in row:
            cell.border = thin_border
    new_workbook.save(output_file)
    
    # Iterate through all cells and set the alignment
    for row in new_worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Read headers from the first row of one of the files
    workbook = openpyxl.load_workbook(file1)
    copy_file1 = workbook.active
    headers = [cell.value for cell in copy_file1[1]]  
    for row in new_worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=new_worksheet.max_column):
        for col_num, header in enumerate(headers, start=1):
            cell = new_worksheet.cell(row=1, column=col_num, value=header)
            if header == "No de la pentade" or header == "Date" or header == "Bellani (gr. Cal/cm2) 6-6h" or header == "Pluies en mm. 6-6h":
                cell.alignment = Alignment(textRotation=90)

    # Save the workbook
    new_workbook.save(output_file)




def has_two_digits_in_order(value, calculated_value):
    """
    Determines if at least two consecutive digits from the calculated value appear in the correct order within the transcribed value, ignoring any decimal points.

    This function checks whether any pair of consecutive digits from the `calculated_value` are present in the transcribed `value`, in the same sequence. It is useful for validating or comparing numerical strings where the order of digits is important, such as in certain data verification tasks.

    Parameters:
    --------------
    value: str
        The string to be checked, typically a value extracted from a dataset.
    calculated_value: str
        The reference string containing the calculated value, against which the `value` is compared.

    Returns:
    --------------
    bool
        Returns True if at least two consecutive digits from `calculated_value` appear in the same order within `value`.
        Returns False otherwise.
    """

    # Remove decimal points by replacing them with an empty string
    value_str = str(value).replace('.', '')
    calculated_str = str(calculated_value).replace('.', '')

    # Sliding window approach to check for two consecutive matching digits
    for i in range(len(calculated_str) - 1):
        # Take two consecutive digits from the calculated value
        consecutive_digits = calculated_str[i:i+2]
        # Check if these two consecutive digits appear in the same order in the value
        if consecutive_digits in value_str:
            return True
    return False


    
def post_processing(pre_processed_excel_file, postprocessed_data_dir, month_filename):
    '''
    Performs Quality Assessment and Quality Control (QA/QC) checks on transcribed climate data to ensure data quality and correctness.

    This function applies various checks and corrections to the transcribed data, including threshold checks for temperatures, check sums and averages, as well as ad-hoc data corrections. The goal is to produce a cleaned and reliable dataset for further analysis or storage.

    Parameters
    --------------
    pre_processed_excel_file: str 
        The path to the Excel file containing preprocessed transcribed data or an ExcelFile object.
    postprocessed_data_dir: str
        The directory path where the final postprocessed Excel file will be stored.
    month_filename: str
        The base name for the output files, typically incorporating station metadata and derived from the original images of climate data sheets.

    Returns
    --------------
    new_workbook: Excel file
        The postprocessed Excel file containing the cleaned and verified climate data.
    '''

    # Open the original Excel file
    workbook = openpyxl.load_workbook(pre_processed_excel_file)
    worksheet = workbook.active

    # Path to save a new copy of the workbook for post-processing
    new_version_of_file = f'{postprocessed_data_dir}\{month_filename}_post_processed.xlsx'

    # Save the original workbook to ensure it's on disk
    original_path = f'src\output\original_transcribed_data.xlsx'
    workbook.save(original_path)

    # Copy the original file to a new version for post-processing
    shutil.copy2(original_path, new_version_of_file)

    new_workbook = openpyxl.load_workbook(new_version_of_file)
    new_worksheet = new_workbook.active # To select the first worksheet of the workbook without requiring its name

    workbook.close() # Close original transcribed workbook


    # Inorder to already avoud very large values right from the start, here we edit values that have more than 4 digits (thousands) in certain rows where we know it is impossible
    # List of rows to exclude from processing
    excluded_rows = [1, 2, 3, 9, 16, 23, 30, 37, 45, 47] # Headers and 5 day totals/ totals
    excluded_columns = [3, 15, 20, 25]  # Example of middle columns to exclude: with U
    # Iterate over all rows in the worksheet
    for row in new_worksheet.iter_rows(min_col=3, max_col=new_worksheet.max_column-1):
        # Get the row index from the first cell of the row
        if row[0].row not in excluded_rows:
            for cell in row:
                if cell.column not in excluded_columns:
                    if is_string_convertible_to_float(cell.value): # Checking if the value transcribed in the cell is convertible to a float to avoid strings 
                        # Convert to string to check length and manipulate
                        str_value = str(cell.value)
                        if len(str_value) > 4: # Thousands. Here i chose 4 because it seems the strings are made up up numbers and a newline \n character. Check this when you use 3 instead
                            # Remove the first digit and convert back to the original type
                            new_value = str_value[1:]  # Remove the first character
                            cell.value = type(cell.value)(new_value)  # Convert back to int or float
                            highlight_change('FF9933', cell, new_version_of_file) #FF9933 is 2 ## To highlight manipulation of transcribed data
                            # Save Excel file
                            new_workbook.save(new_version_of_file)
    
    # Save Excel file
    new_workbook.save(new_version_of_file)


    # Since values transcribed ignore decimal points for all cells yet all the cell values initially are in decimals, here we Iterate over all rows, starting from row 4 to skip header rows
    for row in new_worksheet.iter_rows(min_row=4, max_row=new_worksheet.max_row, min_col=3, max_col=new_worksheet.max_column-1): # For columns: avoid the first two columns and the last column (2nd with Date, and last also with the Date)
        for cell in row:
            if is_string_convertible_to_float(cell.value): # Checking if the value transcribed in the cell is convertible to a float to avoid strings 
                cell.value = (float(cell.value))/10 # Divide the cell value by 10
                new_workbook.save(new_version_of_file) # Save the modified workbook


    # Creating thresholds for temperature values
    Maximum_Temperature_Threshold = 40  # Max reported temperatures during 1950-1959 were 30-31 degC + increasing temperatures in 1960-1990 approximated at 0.60°C to 1.62°C per 30 yr period (Alsdorf et.al, 2016)
    Minimum_Temperature_Threshold = 5  # Min reported temperatures during 1950-1959 were 18-20 degC  et.al, 2016). However lower values of up to 9 and 10°C are noted in some sheets        

    # Cell coordinates to check the totals
    # Rows
    rows = [9, 16, 23, 30, 37, 45]
    #rows = [9] #just for trial

    for row in rows:

        # Columns
        columns = ['D', 'E', 'F'] # Where these represent [ Max Temperature, Min Temperature, Average Temperature, Precipitation] 
        #columns = ['E'] #just for trial

        for column in columns:
            # Get the value of the '5/6-day Total' that was transcribed. Note Months with 31 days ahve some 6 day todays considering 26th to 31st
            cell_for_5_day_total_retrieved = new_worksheet[column + str(row)]
            value_in_cell_for_5_day_total_retrieved = cell_for_5_day_total_retrieved.value
            if value_in_cell_for_5_day_total_retrieved is None or value_in_cell_for_5_day_total_retrieved == '':  # Highlight empty cells
                # Highlight to show that cell is empty 
                highlight_change('FFC0CB', cell_for_5_day_total_retrieved, new_version_of_file) #FFC0CB is Pink
            
            # Select the 5 cells just above the '5-day/6-day Total' in the transcribed file
            # Note, while most totals are based on 5-day values, some months with 31 days have 6-day totals or the last days. 
            # Thus here, the definition of '5-day total' is used for both the 5-day and 6-day totals for simplicity since the value is only used as Quality control check of the daily values
            
            if row in [9, 16, 23, 30, 37]: # rows with only 5-day totals for all months
                offset_cells = 5
            
            else: # last rows with 6-day totals during months with 31 days or EVEN less than 5-days totals during February
                offset_cells = 6
            
            blank_cells = [] # Blank cells (None Type) for dates on the sheet that dont correspond to calender dates for example cells for Feb (29th), 30th and 31st that don't exist
            
            cell_values_for_the_5_days = cell_for_5_day_total_retrieved.offset(row = -offset_cells, column = 0)
            list_of_cell_values_for_the_5_days = []
            for cells in range(offset_cells): 
                cell_values_for_5_days_retrieved = cell_values_for_the_5_days.offset(row = cells, column = 0).value
                
                cell_coordinate = column + str(row + cells - offset_cells) # Identify the particular cell coordinate
                cell_coordinate_in_worksheet = new_worksheet[cell_coordinate]
                
                if cell_values_for_5_days_retrieved is None:
                    list_of_cell_values_for_the_5_days.append(0.0) # Incase there was no transcribed in the cell, convert it to 0.0 to avoid errors in the summation in the following steps for 5/6 day totals.         

                else:
                    list_of_cell_values_for_the_5_days.append(float(cell_values_for_5_days_retrieved))
                    cell_coordinate_in_worksheet.value = (float(cell_values_for_5_days_retrieved))
                    new_workbook.save(new_version_of_file)


                # Blank cells
                if cell_values_for_5_days_retrieved is None or cell_values_for_5_days_retrieved == '':
                    blank_cells.append(1.0) # Count the number of blank cells representing no data or dates that dont exist on the calender for example cells for Feb (29th), 30th and 31st
                    highlight_change('FFFFFF', cell_coordinate_in_worksheet, new_version_of_file) #No Fill

                
                if column in ['D', 'E', 'F']: # Maximimum, Minimum and Average Temperatures
                    
                    if cell_coordinate_in_worksheet.value is None or cell_coordinate_in_worksheet.value == '':
                        # Highlight to show that cell is empty
                        highlight_change('FFFFFF', cell_coordinate_in_worksheet, new_version_of_file) ##FFFFFF is white.
                        new_workbook.save(new_version_of_file)
                    else: 
                        if is_string_convertible_to_float(cell_coordinate_in_worksheet.value): 
                            if float(cell_coordinate_in_worksheet.value) > Maximum_Temperature_Threshold:
                                
                                if float(cell_coordinate_in_worksheet.value) == "":
                                    highlight_change('FFFFFF', cell_coordinate_in_worksheet, new_version_of_file) #No Fill
                                    new_workbook.save(new_version_of_file)
                                else:
                                    # Highlight to show that value is out of the expected bounds
                                    if cell_coordinate_in_worksheet.value >= 100:
                                        new_value = (cell_coordinate_in_worksheet.value)/10.0  #try dividing by 10, to avoid very large values due to missing decimal point
                                        cell_coordinate_in_worksheet.value = new_value

                                    highlight_change('CC3300', cell_coordinate_in_worksheet, new_version_of_file) #CC3300 is Dark Red 
                                    new_workbook.save(new_version_of_file)
                            
                            if float(cell_coordinate_in_worksheet.value) < Minimum_Temperature_Threshold:

                                if float(cell_coordinate_in_worksheet.value) == "":
                                    highlight_change('FFFFFF', cell_coordinate_in_worksheet, new_version_of_file) #No Fill
                                    new_workbook.save(new_version_of_file)
                                else:
                                    # Highlight to show that value is out of the expected Minimum Temperature bounds
                                    highlight_change('CC3300', cell_coordinate_in_worksheet, new_version_of_file) #CC3300 is Dark Red 
                                    new_workbook.save(new_version_of_file)
                
                
                    ### Doing the check again after manipulating the values
                    list_of_cell_values_for_the_5_days = []
                    for cells in range(offset_cells): 
                        cell_values_for_5_days_retrieved = cell_values_for_the_5_days.offset(row = cells, column = 0).value

                        cell_coordinate = column + str(row + cells - offset_cells) # Identify the particular cell coordinate
                        cell_coordinate_in_worksheet = new_worksheet[cell_coordinate]
                        
                        if cell_coordinate_in_worksheet.value is not None:
                            
                            if is_string_convertible_to_float(cell_values_for_5_days_retrieved): # Checking if the value transcribed in the cell is convertible to a float to avoid strings 
                                list_of_cell_values_for_the_5_days.append(float(cell_values_for_5_days_retrieved))
                                cell_coordinate_in_worksheet.value = (float(cell_values_for_5_days_retrieved))
                                new_workbook.save(new_version_of_file)

                            else:
                                list_of_cell_values_for_the_5_days.append(0.0) # Incase the value transcribed in the cell is a string, convert it to 0.0 to avoid errors in the summation.
                
            
            total_from_cell_values_for_the_5_days = sum(list_of_cell_values_for_the_5_days)
            
            #Compare the Total of the 5-days/6-days transcribed and that calculated from the transcribed data
            if is_string_convertible_to_float(value_in_cell_for_5_day_total_retrieved):
                if format(float(value_in_cell_for_5_day_total_retrieved),'.1f') == format(float(total_from_cell_values_for_the_5_days), '.1f'):
                    highlight_change('FF6DCD57', new_worksheet[column + str(row)], new_version_of_file) #6DCD57 is Green. The total of the transcribed 5 days values is equal to the Total transcribed for the cells
                    new_workbook.save(new_version_of_file)
                    # highlight transcribed cells that lead to correct transcribed total
                    for i in range(row - offset_cells, row):
                        highlight_change('FF6DCD57', new_worksheet[column + str(i)], new_version_of_file) #6DCD57 is Green. The total of the transcribed 5 days values is equal to the Total transcribed for the cells
                        new_workbook.save(new_version_of_file)
                    print('The total of the transcribed 5 days values is equal to the Total transcribed for the cell; ' +str(column)+ str(row) + ' is OK')

                else:
                    highlight_change('75696F', new_worksheet[column + str(row)], new_version_of_file) #75696F is Grey. When transcribed 5-day total is not equal to total of the 5 days
                    new_workbook.save(new_version_of_file)
                    print('Check the Total transcribed at cell ' + str(column)+ str(row) +', or the transcribed 5 days values above, because the total of the transcribed 5 days values is not equal to the Total transcribed')
            

            
            
            # Compare the Mean of the 5-days/6-days transcribed and that calculated from the transcribed data
            # We, however, need to skip the 5/6 day mean for precipitation 
            column_to_skip = 'K'  # Incase precipiation is to be considered. This isn't however done in this current example
            if column != column_to_skip:

                cell_coordinate_in_worksheet_with_the_mean = new_worksheet[column + str(row+1)]
                mean_of_5_days_retrieved = cell_coordinate_in_worksheet_with_the_mean.value
                if mean_of_5_days_retrieved is None or mean_of_5_days_retrieved == '':  # Highlight empty cells
                    # Highlight to show that cell is empty 
                    highlight_change('FFC0CB', cell_coordinate_in_worksheet_with_the_mean, new_version_of_file) #FFC0CB is Pink
                    new_workbook.save(new_version_of_file)
                else:
                    if is_string_convertible_to_float(mean_of_5_days_retrieved): 
                        
                        total_blank_cells = sum(blank_cells) # Count of the blank cells
                        denominator = offset_cells - total_blank_cells 

                        if denominator != 0: # This to avoid float division by zero, which result in an error
                        
                            if format(float(mean_of_5_days_retrieved), '.1f') != format(total_from_cell_values_for_the_5_days/(offset_cells - total_blank_cells), '.1f'): # If the mean transcribed (retrieved from the transciption) is not equal to the calculated mean of the values , then highlight the cell
                                cell_coordinate_in_worksheet_with_the_mean.value = float(mean_of_5_days_retrieved)
                                highlight_change('75696F', cell_coordinate_in_worksheet_with_the_mean, new_version_of_file) #75696F is Grey. When transcribed 5-day average is not equal to average of the 5 days
                                new_workbook.save(new_version_of_file)

                            else:
                                highlight_change('FF6DCD57', cell_coordinate_in_worksheet_with_the_mean, new_version_of_file) #6DCD57 is Green. When transcribed 5-day average is equal to average of the 5 days
                                
                                # highlight transcribed cells that lead to correct transcribed average
                                for i in range(row - offset_cells, row):
                                    highlight_change('FF6DCD57', new_worksheet[column + str(i)], new_version_of_file) #6DCD57 is Green. The average of the transcribed 5 days values is equal to the average transcribed for the cells
                                    new_workbook.save(new_version_of_file)
                                
                                new_workbook.save(new_version_of_file)

                        else: 
                            # Handle the case where the denominator is zero
                            highlight_change('75696F', cell_coordinate_in_worksheet_with_the_mean, new_version_of_file) # 75696F is Grey
                            new_workbook.save(new_version_of_file)



    # Check row by row. Daily Minimum, Average and Maximum Temperatures. 
    # If two of the three temperature values (min, max or average) are correct (i.e. checked by the 5/6 day averages or total, and thus highlighted with green) and therefore correct the third value
    # First lets assign a margin of uncertainty
    uncertainty_margin = 0.2  # This because of the rounding off that's done by the observer usually during the calculation of the average daily temperature i.e incase value has 2 decimal points
    
    # The check the cells by rows, and manipulate where necessary
    for row in new_worksheet.iter_rows(min_row=4, max_row=new_worksheet.max_row, min_col=4, max_col=6):     
            
        D, E, F = row    # D = Max Temp., E = Min Temp., F = Average Temp.
        D_highlighted = is_highlighted(D, 'FF6DCD57') # Daily maximum temperature
        E_highlighted = is_highlighted(E, 'FF6DCD57') # Daily minimum temperature
        F_highlighted = is_highlighted(F, 'FF6DCD57') # Daily average temperature

        highlighted_count = D_highlighted + E_highlighted + F_highlighted  # Check how many of these daily temperature values (max, min or ave) are confirmed 'GREEN' from the prior checks)

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        exists_count = D_value_exists + E_value_exists + F_value_exists

        if highlighted_count == 2:
            if not D_highlighted:
                if is_string_convertible_to_float(E.value) and is_string_convertible_to_float(F.value):
                    calculated_D = round(2 * float(F.value) - float(E.value), 1)
                    if D_value_exists and abs(float(D.value) - calculated_D) <= uncertainty_margin:
                        highlight_change('FF6DCD57', D, new_version_of_file)
                    else:
                        D.value = calculated_D
                        highlight_change('FF6DCD57', D, new_version_of_file)
            elif not E_highlighted:
                if is_string_convertible_to_float(D.value) and is_string_convertible_to_float(F.value):
                    calculated_E = round(2 * float(F.value) - float(D.value), 1)
                    if E_value_exists and abs(float(E.value) - calculated_E) <= uncertainty_margin:
                        highlight_change('FF6DCD57', E, new_version_of_file)
                    else:
                        E.value = calculated_E
                        highlight_change('FF6DCD57', E, new_version_of_file)
                    
            elif not F_highlighted:
                if is_string_convertible_to_float(D.value) and is_string_convertible_to_float(E.value):
                    calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
                    if F_value_exists and abs(float(F.value) - calculated_F) <= uncertainty_margin:    
                        highlight_change('FF6DCD57', F, new_version_of_file)
                    else:
                        F.value = calculated_F
                        highlight_change('FF6DCD57', F, new_version_of_file)
            #new_workbook.save(new_version_of_file)
            
        # If one cell is highlighted but has no value, calculate its value
        elif highlighted_count == 3 and exists_count == 2:
            if D_highlighted and not D_value_exists and E_value_exists and F_value_exists:
                D.value = round(2 * float(F.value) - float(E.value), 1)
                highlight_change('FF6DCD57', D, new_version_of_file)
            elif E_highlighted and not E_value_exists and D_value_exists and F_value_exists:
                E.value = round(2 * float(F.value) - float(D.value), 1)
                highlight_change('FF6DCD57', E, new_version_of_file)
            elif F_highlighted and not F_value_exists and D_value_exists and E_value_exists:
                F.value = round((float(D.value) + float(E.value)) / 2, 1)
                highlight_change('FF6DCD57', F, new_version_of_file)
            #new_workbook.save(new_version_of_file)
        
        # If none of the cells are highlighted green but they all have values, or maybe only one of the cells is highlighted green
        elif highlighted_count < 2 and D_value_exists and E_value_exists and F_value_exists:

            calculated_D = round(2 * float(F.value) - float(E.value), 1)
            if abs(calculated_D - float(D.value)) <= uncertainty_margin:
                highlight_change('FF6DCD57', D, new_version_of_file)
                #new_workbook.save(new_version_of_file)
            
            calculated_E = round(2 * float(F.value) - float(D.value), 1)
            if abs(calculated_E - float(E.value)) <= uncertainty_margin:
                highlight_change('FF6DCD57', E, new_version_of_file)
                #new_workbook.save(new_version_of_file)
            

            calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
            if abs(calculated_F - float(F.value)) <= uncertainty_margin:
                highlight_change('FF6DCD57', F, new_version_of_file)
                #new_workbook.save(new_version_of_file)
            


            # Check if at least two digits of the calculated value are present in the correct order in the given value. If so, replace the transcribed value with the corrected value
            if has_two_digits_in_order(D.value, calculated_D) and Minimum_Temperature_Threshold <= calculated_D <= Maximum_Temperature_Threshold and calculated_D > E.value:
                D.value = calculated_D
                highlight_change('FF6DCD57', D, new_version_of_file)
                # new_workbook.save(new_version_of_file)

            # Check if at least two digits of the calculated value are present in the correct order in the given value. If so, replace the transcribed value with the correcetd value
            if has_two_digits_in_order(E.value, calculated_E) and Minimum_Temperature_Threshold <= calculated_E <= Maximum_Temperature_Threshold and calculated_E < D.value :
                E.value = calculated_E
                highlight_change('FF6DCD57', E, new_version_of_file)
                # new_workbook.save(new_version_of_file)

            # Check if at least two digits of the calculated value are present in the correct order in the given value. If so, replace the transcribed value with the correcetd value
            if has_two_digits_in_order(F.value, calculated_F) and Minimum_Temperature_Threshold <= calculated_F <= Maximum_Temperature_Threshold and E.value < calculated_F < D.value:
                F.value = calculated_F
                highlight_change('FF6DCD57', F, new_version_of_file)
                # new_workbook.save(new_version_of_file)
                
        #new_workbook.save(new_version_of_file)    

        # RE-CHECKS
        D_highlighted = is_highlighted(D, 'FF6DCD57')
        E_highlighted = is_highlighted(E, 'FF6DCD57')
        F_highlighted = is_highlighted(F, 'FF6DCD57')

        highlighted_count = D_highlighted + E_highlighted + F_highlighted

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        if highlighted_count >=1 :
            if E_value_exists and F_value_exists and Minimum_Temperature_Threshold <= E.value <= Maximum_Temperature_Threshold and Minimum_Temperature_Threshold <= F.value <= Maximum_Temperature_Threshold and F.value > E.value:
                calculated_D = round(2 * float(F.value) - float(E.value), 1)
                if D.value is not None and abs(float(D.value) - calculated_D) <= uncertainty_margin:
                    highlight_change('FF6DCD57', D, new_version_of_file)
                else:
                    if not D_highlighted:
                        D.value = calculated_D
                        highlight_change('FF6DCD57', D, new_version_of_file)
            if D_value_exists and F_value_exists and Minimum_Temperature_Threshold <= D.value <= Maximum_Temperature_Threshold and Minimum_Temperature_Threshold <= F.value <= Maximum_Temperature_Threshold and D.value > F.value:
                calculated_E = round(2 * float(F.value) - float(D.value), 1)
                if E.value is not None and abs(float(E.value) - calculated_E) <= uncertainty_margin:
                    highlight_change('FF6DCD57', E, new_version_of_file)
                else:
                    if not E_highlighted:
                        E.value = calculated_E
                        highlight_change('FF6DCD57', E, new_version_of_file)
            if D_value_exists and E_value_exists and Minimum_Temperature_Threshold <= D.value <= Maximum_Temperature_Threshold and Minimum_Temperature_Threshold <= E.value <= Maximum_Temperature_Threshold and D.value > E.value:
                calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
                if F.value is not None and abs(float(F.value) - calculated_F) <= uncertainty_margin:    
                    highlight_change('FF6DCD57', F, new_version_of_file)
                else:
                    if not F_highlighted:
                        F.value = calculated_F
                        highlight_change('FF6DCD57', F, new_version_of_file)
        
        if D_value_exists and E_value_exists and F_value_exists and Minimum_Temperature_Threshold <= D.value <= Maximum_Temperature_Threshold and Minimum_Temperature_Threshold <= E.value <= Maximum_Temperature_Threshold and Minimum_Temperature_Threshold <= F.value <= Maximum_Temperature_Threshold and E.value < F.value < D.value:
            # Calculate the average of D and E and check if it is within the uncertainty margin of F
            calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
            if abs(calculated_F - float(F.value)) <= uncertainty_margin:
                # Highlight all cells green if the condition is true
                highlight_change('FF6DCD57', D, new_version_of_file)
                highlight_change('FF6DCD57', E, new_version_of_file)
                highlight_change('FF6DCD57', F, new_version_of_file)


    new_workbook.save(new_version_of_file)


    # Check of Max, Min and Avg Temperatures using the Amplitude (Ampl.) commonly known as the Diurnal Temperarure Range. This is column G of our worksheets
    # where: Ampl. = Max - Min ..................... (1)
    #        Ampl. = 2Avg - 2Min ................... (2)
    #        Ampl. = 2Max - 2Avg ................... (3)
    # The check the cells by rows, and manipulate where necessary
    for row in new_worksheet.iter_rows(min_row=4, max_row=new_worksheet.max_row, min_col=4, max_col=7):    
            
        D, E, F, G = row   # D = Max Temp., E = Min Temp., F = Average Temp., G = Diurnal Temperature range

        D_highlighted = is_highlighted(D, 'FF6DCD57')
        E_highlighted = is_highlighted(E, 'FF6DCD57')
        F_highlighted = is_highlighted(F, 'FF6DCD57')

        # highlighted_count = D_highlighted + E_highlighted + F_highlighted

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)
        G_value_exists = is_string_convertible_to_float(G.value)

        if G_value_exists:
            G_value = float(G.value) # Here G_value is the Amplitude
            if G_value >+ 100:
                G_value /= 10
                highlight_change('CC3300', G, new_version_of_file) #CC3300 is Dark Red.  
            G.value = round(G_value, 1)
        
            # Calculate the amplitude based on the given formulas
            if D_value_exists and E_value_exists:
                D_value = float(D.value) # Max Temp
                E_value = float(E.value) # Min Temp

                Ampl_1 = D_value - E_value  # Ampl. = Max - Min ..................... (1)
                if abs(G_value - Ampl_1) <= uncertainty_margin:
                    # Highlight Ampl. cell green if the condition is true
                    highlight_change('FF6DCD57', G, new_version_of_file)
                    if not F_highlighted:
                        calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
                        F.value = calculated_F
                        highlight_change('FF6DCD57', D, new_version_of_file)
                        highlight_change('FF6DCD57', E, new_version_of_file)
                        highlight_change('FF6DCD57', F, new_version_of_file)
            
            if E_value_exists and F_value_exists:
                E_value = float(E.value) # Min Temp
                F_value = float(F.value) # Avg Temp

                Ampl_2 = (2*F_value) - (2*E_value) # Ampl. = 2Avg - 2Min ................... (2)
                if abs(G_value -Ampl_2) <= uncertainty_margin:
                    # Highlight Ampl. cell green if the condition is true
                    highlight_change('FF6DCD57', G, new_version_of_file)
                    if not D_highlighted:
                        calculated_D = round(2 * float(F.value) - float(E.value), 1)
                        D.value = calculated_D
                        highlight_change('FF6DCD57', D, new_version_of_file)
                        highlight_change('FF6DCD57', E, new_version_of_file)
                        highlight_change('FF6DCD57', F, new_version_of_file)
            
            if D_value_exists and F_value_exists:
                D_value = float(D.value) # Max Temp
                F_value = float(F.value) # Avg Temp

                Ampl_3 = (2*D_value) - (2*F_value) # Ampl. = 2Max - 2Avg ................... (3)
                if abs(G_value -Ampl_3) <= uncertainty_margin:
                    # Highlight Ampl. cell green if the condition is true
                    highlight_change('FF6DCD57', G, new_version_of_file)
                    if not E_highlighted:
                        calculated_E = round(2 * float(F.value) - float(D.value), 1)
                        E.value = calculated_E
                        highlight_change('FF6DCD57', D, new_version_of_file)
                        highlight_change('FF6DCD57', E, new_version_of_file)
                        highlight_change('FF6DCD57', F, new_version_of_file)

    new_workbook.save(new_version_of_file)


    # Final check the cells by rows to ensure that Min temp < Avg temp < Max Temp, and that re-checks for 
    excluded_rows = [9, 16, 23, 30, 37, 45] # These are the rows with 5/6  day totals since totals are usually larger than thresholds.  
    for row in new_worksheet.iter_rows(min_row=4, max_row=new_worksheet.max_row, min_col=4, max_col=6):
        if row[0].row in excluded_rows:
            continue 
        
        D, E, F = row

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        if D_value_exists and E_value_exists and F_value_exists:
            D_value = float(D.value)
            E_value = float(E.value)
            F_value = float(F.value)

            # In order to avoid values not within the thresholds and negatives perharps gotten through the calculations/manipulations
            if not (Minimum_Temperature_Threshold < D_value < Maximum_Temperature_Threshold) or D_value < 0: 
                highlight_change('CC3300', D, new_version_of_file)  # Dark Red
            if not (Minimum_Temperature_Threshold < E_value < Maximum_Temperature_Threshold) or E_value < 0:
                highlight_change('CC3300', E, new_version_of_file)  # Dark Red
            if not (Minimum_Temperature_Threshold < F_value < Maximum_Temperature_Threshold) or F_value < 0:
                highlight_change('CC3300', F, new_version_of_file)  # Dark Red
            

    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)

    # FINAL RE-CHECKS
    # New logic to recalculate 5/6 day totals if all five/six cells above are highlighted green
    total_rows = [9, 16, 23, 30, 37, 45]
    columns = ['D', 'E', 'F']   # Max, Min and Ave Temperatures

    for row in total_rows:
        for col in columns:
            # Check if the total is already highlighted
            total_highlighted = is_highlighted(new_worksheet[f"{col}{row}"], 'FF6DCD57')
            if row != 45:  # 5 days sum
                cells_above = [new_worksheet[f"{col}{row - i}"] for i in range(1, 6)]
            else:  # the last sum contains 6 days (Day 26 to Day 31 of the month)
                cells_above = [new_worksheet[f"{col}{row - i}"] for i in range(1, 7)]
            if all(is_highlighted(cell, 'FF6DCD57') for cell in cells_above) and not total_highlighted:
                sum_value = sum(float(cell.value) for cell in cells_above)
                sum_cell = new_worksheet[f"{col}{row}"]
                sum_cell.value = round(sum_value, 1)
                highlight_change('FF6DCD57', sum_cell, new_version_of_file)
            elif sum(is_highlighted(cell, 'FF6DCD57') for cell in cells_above) == len(cells_above) - 1:
                non_highlighted_cell = next(cell for cell in cells_above if not is_highlighted(cell, 'FF6DCD57'))
                if total_highlighted:
                    total_value = float(new_worksheet[f"{col}{row}"].value)
                    non_highlighted_cell.value = round(total_value - sum(float(cell.value) for cell in cells_above if cell != non_highlighted_cell), 1)
                    highlight_change('FF6DCD57', non_highlighted_cell, new_version_of_file)


    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)


    # New logic to recalculate 5/6 day averages if all five/six cells above are highlighted green
    average_rows = [10, 17, 24, 31, 38, 46]
    columns = ['D', 'E', 'F']   # Max, Min and Ave Temperatures

    for row in average_rows:
        for col in columns:
            # Check if the total is already highlighted
            average_highlighted = is_highlighted(new_worksheet[f"{col}{row}"], 'FF6DCD57')
            if row is not 46: # 5 days average
                cells_above = [new_worksheet[f"{col}{row - i}"] for i in range(2, 7)]
            if row is 46: # the last average contains 6 days (Day 26 to Day 31 of the month)
                cells_above = [new_worksheet[f"{col}{row - i}"] for i in range(2, 8)]
            if all(is_highlighted(cell, 'FF6DCD57') for cell in cells_above) and not average_highlighted:
                average_value = mean(float(cell.value) for cell in cells_above)
                average_cell = new_worksheet[f"{col}{row}"]
                average_cell.value = round(average_value, 1)
                highlight_change('FF6DCD57', average_cell, new_version_of_file)
            elif sum(is_highlighted(cell, 'FF6DCD57') for cell in cells_above) == len(cells_above) - 1:
                non_highlighted_cell = next(cell for cell in cells_above if not is_highlighted(cell, 'FF6DCD57'))
                if average_highlighted:
                    total_value = float(new_worksheet[f"{col}{row}"].value) * len(cells_above)
                    non_highlighted_cell.value = round(total_value - sum(float(cell.value) for cell in cells_above if cell != non_highlighted_cell), 1)
                    highlight_change('FF6DCD57', non_highlighted_cell, new_version_of_file)
                                
    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)

    new_workbook.close()

    return new_workbook



# Compare the 'confirmed' transcribed and post corrected data with manually entered data
def compare_workbooks(file1, file2, uncertainty_margin=0.2):
    # Load both workbooks
    wb1 = openpyxl.load_workbook(file1)
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook(file2)
    ws2 = wb2.active

    total_highlighted_cells = 0
    accurate_matches = 0

    # Iterate through the cells in the first workbook
    for row in ws1.iter_rows(min_row=4, max_row=ws1.max_row, min_col=4, max_col=6):
        for cell in row:
            if is_highlighted(cell, 'FF6DCD57'):
                total_highlighted_cells += 1

                cell_value_ws1 = cell.value
                cell_value_ws2 = ws2.cell(row=cell.row, column=cell.column).value

                if is_string_convertible_to_float(cell_value_ws1) and is_string_convertible_to_float(cell_value_ws2):
                    if abs(float(cell_value_ws1) - float(cell_value_ws2)) <= uncertainty_margin:
                        accurate_matches += 1

    if total_highlighted_cells == 0:
        accuracy_percentage = 0.0
    else:
        accuracy_percentage = (accurate_matches / total_highlighted_cells) * 100

    print(f"Total Highlighted Cells: {total_highlighted_cells}")
    #print(f"Accurate Matches: {accurate_matches}")
    print(f"Accuracy Percentage: {accuracy_percentage:.2f}%")

    # Close the workbooks
    wb1.close()
    wb2.close()