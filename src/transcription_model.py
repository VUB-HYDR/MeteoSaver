#!/usr/bin/env python
import os, argparse, glob, tempfile, shutil, warnings
import cv2
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
# from paddleocr import PaddleOCR,draw_ocr
from PIL import Image, ImageDraw, ImageFont
import pytesseract
import easyocr
import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
import math
import random
from scipy.stats import trim_mean

def organize_contours_midpoint(contours, max_rows):
    '''
    # Organizes the bounding boxes, here termed as contours, in rows by their center co-ordinates using the KMeans clustering

    Parameters
    --------------
    contours: list of bounding boxes (with their x, y, w, h coordinates)
        List of contours for the detected text in the table cells with coordinates

    max_rows: int
        Maximum rows, adjusted based on your table's expected structure

    Returns
    -------------- 
    rows: Bounding boxes organised in rows using Kmeans clustering

    '''

    midpoints = [(cv2.boundingRect(contour)[1] + cv2.boundingRect(contour)[3] // 2) for contour in contours]
    if len(midpoints) == 0:
        return []
    kmeans = KMeans(n_clusters=min(max_rows, len(midpoints)), random_state=0)
    kmeans.fit(np.array(midpoints).reshape(-1, 1))
    labels = kmeans.labels_
    rows = [[] for _ in range(max_rows)]
    for label, contour in zip(labels, contours):
        rows[label].append(contour)
    for i in range(len(rows)):
        rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])
    return rows


def organize_contours_top(contours, max_rows):
    '''
    # Organizes the bounding boxes, here termed as contours, in rows by their top co-ordinates using the KMeans clustering

    Parameters
    --------------
    contours: list of bounding boxes (with their x, y, w, h coordinates)
        List of contours for the detected text in the table cells with coordinates

    max_rows: int
        Maximum rows, adjusted based on your table's expected structure

    Returns
    -------------- 
    rows: Bounding boxes organised in rows using Kmeans clustering

    '''

    # Top for vertical clustering
    top = [cv2.boundingRect(contour)[1] for contour in contours]
    if len(top) == 0:
        return []

    kmeans = KMeans(n_clusters=min(max_rows, len(top)), random_state=0)
    kmeans.fit(np.array(top).reshape(-1, 1))
    labels = kmeans.labels_

    rows = [[] for _ in range(max_rows)]
    for label, contour in zip(labels, contours):
        rows[label].append(contour)

    for i in range(len(rows)):
        rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])

    return rows



def filter_contours(contours, min_width_threshold, min_height_threshold, max_width_threshold, max_height_threshold):
    '''
    Filters a list of contours based on specified width and height thresholds.

    This function takes a list of contours and filters them by their bounding rectangle dimensions. Contours whose bounding rectangles fall within the specified width and height thresholds are retained, while others are discarded. This is useful for removing noise or irrelevant contours based on size constraints e.g. small contours that don't contain text or very large contours that cover more than one cell.

    Parameters
    --------------
    contours : list
        A list of contours, where each contour is an array of points defining the contour's shape.
    
    min_width_threshold : int
        The minimum width (in pixels) that a contour's bounding rectangle must have to be included in the filtered results.
    
    min_height_threshold : int
        The minimum height (in pixels) that a contour's bounding rectangle must have to be included in the filtered results.
    
    max_width_threshold : int
        The maximum width (in pixels) that a contour's bounding rectangle can have to be included in the filtered results.
    
    max_height_threshold : int
        The maximum height (in pixels) that a contour's bounding rectangle can have to be included in the filtered results.

    Returns
    --------------
    filtered_contours : list of bounding boxes (filtered)
        A list of contours that meet the specified width and height criteria. Contours whose bounding rectangles do not fall within the given thresholds are excluded.
    '''

    filtered_contours = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if min_width_threshold <= w <= max_width_threshold and min_height_threshold <= h <= max_height_threshold:
            filtered_contours.append(contour)
    return filtered_contours



def calculate_cell_reference(x, row_index, max_columns, table_width):
    '''
    Calculates the Excel cell reference (e.g., 'B3') based on the horizontal position of a detected cell's x-coordinate within a table.

    This function determines the appropriate Excel cell reference by converting the horizontal position of a cell's x-coordinate into a column index. The column index is combined with the given row index to generate the full cell reference in the format used by Excel (e.g., 'A1', 'B2'). The column index is adjusted to ensure it remains within valid boundaries (1 to `max_columns`).

    Parameters
    --------------
    x : float or int
        The x-coordinate of the cell (in pixels), representing its horizontal position within the table.
    
    row_index : int
        The row number in the table to which the cell belongs. This is used directly in the final cell reference.
    
    max_columns : int
        The maximum number of columns in the table. This ensures that the calculated column index does not exceed the available columns.
    
    table_width : int or float
        The total width of the table (in pixels). This is used to determine the relative position of `x` within the table and calculate the corresponding column index.

    Returns
    --------------
    cell_reference : str
        The Excel-style cell reference (e.g., 'B3', 'C7') corresponding to the detected cell's position within the table.
    '''

    column = math.floor(x / table_width * max_columns) + 1

    # Ensure the column index is within valid ranges
    if column < 1:
        column = 1
    elif column > max_columns:
        column = max_columns

    return f'{openpyxl.utils.get_column_letter(column)}{row_index}'



def group_contours_into_columns(contours, num_columns, image_width):
    '''
    Groups contours into columns based on their horizontal position within an image.

    This function organizes a list of contours into a specified number of columns by calculating the column index for each contour based on its x-coordinate. The image is divided into equal-width columns, and each contour is assigned to a column based on where its bounding box falls horizontally. The resulting groups of contours are returned as a dictionary where the keys represent column indices.

    Parameters
    --------------
    contours :  list 
        A list of contours, where each contour is an array of points that define the contour's shape.

    num_columns : int
        The number of columns (from the expected table structure) into which the contours should be grouped. This value determines how the image width is divided.

    image_width : int
        The total width of the image/table (in pixels). This is used to calculate the width of each column and to determine in which column each contour belongs.

    Returns
    --------------
    columns : dict
        A dictionary where each key is a column index (ranging from 0 to `num_columns` - 1), and the value is a list of tuples. Each tuple represents a contour's bounding box in the format `(x, y, w, h)`, indicating its position and size within the image.
    '''

    columns = {i: [] for i in range(num_columns)}
    column_width = image_width // num_columns
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        column_index = x // column_width
        columns[column_index].append((x, y, w, h))
    return columns


def add_missing_rois(sorted_contours, space_threshold, width_threshold, max_height_per_box, max_rows, num_columns, image_width):
    '''
    Identifies and adds missing regions of interest (ROIs) in a table by filling gaps between detected contours within each column.

    This function takes a set of sorted contours representing detected table cells and analyzes gaps between them within each column. If significant gaps are found, it adds new bounding boxes (ROIs) in those gaps to ensure that all expected rows are accounted for. This process helps to detect and fill in any missing cells that were not initially identified during contour detection. The newly generated contours are returned for further processing.

    Parameters
    --------------
    sorted_contours : list
        A list of contours sorted in the desired order (typically by their y-coordinate within the table). These contours represent the detected cells in the table.

    space_threshold : int
        The minimum vertical space (in pixels) between two consecutive contours within a column that should be considered a gap. If the space exceeds this threshold, a new ROI is added to fill the gap.

    width_threshold : int
        The width (in pixels) of the new ROI boxes that will be added to fill the gaps. This ensures that the new boxes have a consistent width relative to the other cells in the column.

    max_height_per_box : int
        The estimated maximum height (in pixels) for the new ROI boxes. This value is used to determine the size of the gaps and how to place the new boxes.

    max_rows : int
        The maximum number of rows expected in each column (from the expected table structure). The function will not add more boxes than this number, ensuring that the final column does not exceed the expected row count.

    num_columns : int
        The number of columns in the table. This determines how the contours are grouped and processed.

    image_width : int
        The total width of the image/table (in pixels). This is used to calculate the width of each column and group contours accordingly.

    Returns
    --------------
    new_contours : list
        A list of new contours representing the updated set of detected and added ROIs. These contours include both the original contours and any new ones created to fill gaps.
    '''


    # Group contours into columns
    columns = group_contours_into_columns(sorted_contours, num_columns, image_width)

    new_boxes = []
    for i in sorted(columns.keys()):  # Ensure columns are processed in order
        column_boxes = sorted(columns[i], key=lambda b: b[1])  # Sort by y-coordinate
        column_count = len(column_boxes)
        print(f'Number of current rows in the current column: {column_count}')  # Debug statement
        # Calculate gaps and sort them by size (largest first)
        gaps = []
        for j in range(1, len(column_boxes)):
            prev_box = column_boxes[j - 1]
            curr_box = column_boxes[j]
            space_between = curr_box[1] - (prev_box[1] + prev_box[3])
            if space_between > space_threshold:
                gaps.append((space_between, prev_box, curr_box))
        
        gaps.sort(key=lambda x: x[0], reverse=True)  # Sort gaps in between the cells by size (largest first)

        # Add new boxes for the gaps in priority order
        for gap in gaps:
            if column_count >= max_rows:
                break
            space_between, prev_box, curr_box = gap
            # Calculate the y position for the new contour
            new_y = prev_box[1] + prev_box[3] + (space_between - max_height_per_box) // 2
            new_height = max_height_per_box
            new_box = (prev_box[0], new_y, width_threshold, new_height)

            # Check for neighboring boxes within the same row
            has_left_neighbor = any(abs(prev_box[0] - b[0]) <= width_threshold for b in column_boxes)
            has_right_neighbor = any(abs(curr_box[0] - b[0]) <= width_threshold for b in column_boxes)

            if has_left_neighbor or has_right_neighbor:
                print(f'Added new box at: {new_box}')  # Debug statement
                column_boxes.append(new_box)
                column_count += 1

        column_boxes = sorted(column_boxes, key=lambda b: b[1])  # Sort again after adding new boxes
        new_boxes.extend(column_boxes)
    
    new_contours = [np.array([[box[0], box[1]], [box[0] + box[2], box[1]], [box[0] + box[2], box[1] + box[3]], [box[0], box[1] + box[3]]], dtype=np.int32) for box in new_boxes]
    
    return new_contours



def generate_random_colors(n):
    '''
    Generates a list of `n` random distinct colors in BGR format.

    This function creates a list of distinct colors by generating random values for hue, saturation, and value (HSV) and then converting them to the BGR color space, which is commonly used in OpenCV. The generated colors are vivid and bright, ensuring they are visually distinct when used in applications such as object detection, image segmentation, or visual annotations.

    Parameters
    --------------
    n : int
        The number of distinct colors to generate.

    Returns
    --------------
    colors : list of tuples
        A list containing `n` tuples, each representing a color in BGR format. The colors are designed to be vivid and distinct for clear visualization.
    '''


    colors = []
    for i in range(n):
        # Generate random values for hue, saturation, and value
        hue = random.randint(0, 179)  # Hue range in OpenCV is [0, 179]
        saturation = random.randint(100, 255)  # To ensure the colors are vivid
        value = random.randint(100, 255)  # To ensure the colors are bright

        # Convert the random HSV color to BGR
        color = cv2.cvtColor(np.uint8([[[hue, saturation, value]]]), cv2.COLOR_HSV2BGR)[0][0].tolist()
        colors.append((int(color[0]), int(color[1]), int(color[2])))
    return colors




def draw_row_markers_and_boxes(image, rows, colors):
    '''
    Draws bounding boxes around contours in each row and adds numbered markers to each row with distinct colors.

    This function takes an image and a list of rows, where each row is a list of contours representing table cells or regions of interest. It draws a bounding box around each contour in the row and places a numbered marker to the left of the row. Each row is highlighted with a distinct color, cycling through the provided color list if necessary.

    Parameters
    --------------
    image : 
        The image on which the bounding boxes and row markers will be drawn. This image is modified in place.
    
    rows : list of lists
        A list of rows, where each row is a list of contours (represented as arrays of points) that are part of the same row in a table or grid.
    
    colors : list of tuples
        A list of BGR color tuples used to color the bounding boxes and markers for each row. The function cycles through this list if there are more rows than colors.

    Returns
    --------------
    None
        The function modifies the input image in place and does not return any value. The image will have bounding boxes drawn around the contours and numbered markers for each row.
    '''


    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.5
    thickness = 1

    for idx, row in enumerate(rows):
        color = colors[idx % len(colors)]  # Cycle through colors if more rows than colors
        y_coords = [cv2.boundingRect(c)[1] + cv2.boundingRect(c)[3] // 2 for c in row]
        if y_coords:
            y_position = int(np.median(y_coords))
            x_position = 10  # arbitrary x position to place the marker
            cv2.putText(image, str(idx + 1), (x_position, y_position), font, font_scale, color, thickness)
        
        # Draw bounding boxes for each contour in the row
        for contour in row:
            x, y, w, h = cv2.boundingRect(contour)
            cv2.rectangle(image, (x, y), (x + w, y + h), color, 2)




def calculate_trimmed_mean(values, proportion_to_cut=0.2):
    '''
    Calculates the trimmed mean of a list of values, excluding a specified proportion of the smallest and largest values.

    This function computes the trimmed mean of a given list of numerical values. The trimmed mean is a measure of central tendency that removes a certain proportion of the lowest and highest values before calculating the mean. This helps to reduce the effect of outliers on the mean.

    Parameters
    --------------
    values : list
        A list of numerical values for which the trimmed mean is to be calculated.
    
    proportion_to_cut : float, optional
        The proportion of values to remove from each end of the sorted list before calculating the mean. 
        For example, a proportion of 0.2 means removing the lowest 20% and the highest 20% of the values. 
        The default value is 0.2.

    Returns
    --------------
    float
        The trimmed mean of the input values, after excluding the specified proportion of extreme values.
    '''

    return trim_mean(values, proportion_to_cut)



def transcription(detected_table_cells, ocr_model, no_of_rows, no_of_columns, min_cell_width_threshold, max_cell_width_threshold, min_cell_height_threshold, max_cell_height_threshold):
    '''
    Performs OCR (Optical Character Recognition) on detected table cells from a pre-processed grayscale image to extract text and store the results in an Excel workbook.

    This function processes a detected table within an image, using contours to identify individual cells. It clips each cell and applies the specified OCR/HTR (Handwritten Text Recognition) model to transcribe the text within each cell. The results are then saved in an Excel workbook.

    Parameters
    --------------
    detected_table_cells : list
        detected_table_cells[0]: contours. Contours for the detected text in the table cells.
        detected_table_cells[1]: image_with_all_bounding_boxes. Image with bounding boxes drawn around each detected table cell.
        detected_table_cells[2]: table_copy. A copy of the processed table image used for further operations.
        detected_table_cells[3]: table_original_image. The original image of the table before any processing.
    
    ocr_model : string
        The OCR or HTR model used to recognize and transcribe text from the detected table cells. This can be any model that supports text recognition, such as Tesseract, EasyOCR, or a custom deep learning model.
    
    no_of_rows : int
        The number of rows in the detected table. This parameter helps in structuring the data correctly in the output workbook.
    
    no_of_columns : int
        The number of columns in the table, used to structure the OCR results.
    
    min_cell_width_threshold : int
        The minimum width of a cell (in pixels) that should be considered for OCR. Cells smaller than this threshold are ignored.
    
    max_cell_width_threshold : int
        The maximum width of a cell (in pixels) that should be considered for OCR. Cells wider than this threshold are ignored.
    
    min_cell_height_threshold : int
        The minimum height of a cell (in pixels) that should be considered for OCR. Cells smaller than this threshold are ignored.
    
    max_cell_height_threshold : int
        The maximum height of a cell (in pixels) that should be considered for OCR. Cells taller than this threshold are ignored.

    Returns
    -------------- 
    wb : openpyxl.Workbook
        An Excel workbook object where the OCR results are stored. Each cell's transcribed text is placed in the corresponding cell of the Excel sheet based on its detected position in the image.
    '''
    
    
    if ocr_model == 'Tesseract-OCR':
        ## Lauching Tesseract-OCR
        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' ## Here input the PATH to the Tesseract executable on your computer. See more information here: https://pypi.org/project/pytesseract/
    # if ocr_model == 'PaddleOCR':
    #     ## Lauching PaddleOCR, which would be used by downloading necessary files as shown below
    #     paddle_ocr = PaddleOCR(use_angle_cls=True, lang = 'en', use_gpu=False) ## Run only once to download all required files
    if ocr_model == 'EasyOCR':
        ## Lauching EasyOCR
        easyocr_reader = easyocr.Reader(['en']) # this needs to run only once to load the model into memory

    easyocr_reader = easyocr.Reader(['en'])

    contours = detected_table_cells[0]

    # Filter out smaller or larger bounding boxes from all the detected text contours. This is helpful to avoid overly large cells or small cells with no text. Remember to adjust these values based on the table structure in your specific case 
    filtered_contours = filter_contours(contours, min_cell_width_threshold, min_cell_height_threshold, max_cell_width_threshold, max_cell_height_threshold)

    image_with_all_bounding_boxes = detected_table_cells[1]
    table_copy = detected_table_cells[2]

    # Get the dimensions of the loaded image. Here, particulary the image/table width is very important for the column placement of cells/bounding boxes
    image_height, image_width, image_channels = image_with_all_bounding_boxes.shape

    # Image onto which the regions of interest (ROIs) i.e the cells, will be drawn for illustration purposes
    ROIs_image = table_copy.copy()
    
    results = []

    # Here we use two methods to arrange the boundung boxes (as a double check): (1) Using the middle coordinates of the bounding boxes , and (2) Using the top coordinated of the boudning boxes.
    organize_methods = {  
        'Midpoint': organize_contours_midpoint,
        'Top': organize_contours_top
    }

    for method_name, organize_method in organize_methods.items():

        ## Create an Excel workbook and add a worksheet where the transcribed text will be saved
        wb = Workbook()
        ws = wb.active
        ws.title = 'OCR_Results'

        # Sort contours by y-coordinate
        contours_sorted = sorted(filtered_contours, key=lambda c: cv2.boundingRect(c)[1])

        # Adding missing bounding boxes. Here, we define the minimum height space and minimum width space between the bounding boxes in a column and row respectively, in case of a missing bounding box.
        space_height_threshold = 50 # Minimum height space between the bounding boxes in a column in case of a missing bounding box. Modify as needed for your case
        space_width_threshold = 120 # Minimum width space between the bounding boxes in a row in case of a missing bounding box. Modify as needed for your case
        max_cell_height_per_box = 50 # Set height for a new box (missing box) to add to the bounding boxes list
        # Add missing ROIs to the contours
        new_contours = add_missing_rois(contours_sorted, space_height_threshold, space_width_threshold, max_cell_height_per_box, no_of_rows, no_of_columns, image_width)

        # Organize the contours
        organized_rows = organize_method(new_contours, no_of_rows)
        # Sorting the cell (bounding box) rows from first to last using trimmed mean of coordinates
        sorted_rows = sorted(organized_rows, key=lambda row: calculate_trimmed_mean([cv2.boundingRect(c)[1] + cv2.boundingRect(c)[3] // 2 for c in row]))

    
        ### FOR ILLUSTRATION PURPOSES: Uncomment the following lines if you want to see an example of how the sorting function works. This is for illustration purposes only and not necessary for the main functionality of the script. 
        
        # # Create a copy of the image for visualization
        # image_before_sorting = table_copy.copy()
        # image_after_sorting = table_copy.copy()

        # # Generate colors for 43 rows. For sorted row visualitation purposes
        # colors = generate_random_colors(no_of_rows)  # Random colours for the maximum number of rows, such that each row has its own color for easy identification

        # # Draw markers on the image before sorting
        # draw_row_markers_and_boxes(image_before_sorting, organized_rows, colors)  # Green color for original order

        # # Draw markers on the image after sorting
        # draw_row_markers_and_boxes(image_after_sorting, sorted_rows, colors)  # Red color for sorted order

        # # Save or display the images for inspection
        # cv2.imwrite(f'before_sorting_{method_name}.png', image_before_sorting)  # or use cv2.imshow and cv2.waitKey for immediate display
        # plt.imshow('before_sorting.png')
        # plt.show()

        # cv2.imwrite(f'after_sorting_{method_name}.png', image_after_sorting)  # or use cv2.imshow and cv2.waitKey for immediate display
        # plt.imshow('after_sorting.png')
        # plt.show()
        
    

        # Sort boxes within each column of each row by y-coordinate
        for row in sorted_rows:
            row.sort(key=lambda c: cv2.boundingRect(c)[1])

        for row_index, row in enumerate(sorted_rows, start=1):
            
            for contour in row:

                x, y, w, h = cv2.boundingRect(contour)              

                # Calculate a factor to modify the bounding box area (e.g., in this case 5% of width and 25% of height)
                factor_width = 0.05  # Modify this factor as needed
                increase_factor_height = 0.25 # Modify this factor as needed
            
                x += int(w * factor_width)  # Increase width
                y -= int(h * increase_factor_height)  # Increase height
                w -= int(1 * w * factor_width)  # Decrease width a little to avoid vertical lines that may be transcribed as the number 1 yet they aren't a number
                h += int(2 * h * increase_factor_height)  # Increase height
                
                ### FOR ILLUSTRATION PURPOSES: This line below is about drawing a rectangle on the image with the shape of the bounding box. Its not needed for the OCR. This is only for debugging purposes.
                # image_with_all_bounding_boxes = cv2.rectangle(image_with_all_bounding_boxes, (x, y), (x + w, y + h), (0, 255, 0), 5)

                # Draw the adjusted ROI on the output image
                cv2.rectangle(ROIs_image, (x, y), (x + w, y + h), (0, 255, 0), 5)  # (0, 255, 0) represent a green color for ROI, and 5 is the thicnkess of the ROI bounbdary boxes
                
                
                # OCR
                # Crop each cell using the bounding rectangle coordinates
                ROI = table_copy[y:y+h, x:x+w] # Here, the Region Of Interest (ROI) represent the cells (boundary boxes) clipped out from the table image as a prerequiste for text recogniton by the Optical Character Recognition/Handwritten Text Recognition (OCR/HTR) model
                
                if ROI.size != 0:  # Check if the height and width are greater than zero. This is to prevent invalid ROIs
                    
                    cv2.imwrite('detected.png', ROI)
                    if ocr_model == 'Tesseract-OCR':
                    # Using Tesseract-OCR
                        ocr_result = pytesseract.image_to_string('detected.png', lang='cobedore-V6', config='--psm 7 -c tessedit_char_whitelist=0123456789') # Just added -c tessedit_char_whitelist=0123456789 to really limit the text type/values detected

                        # Here's a brief explanation of some Page Segmentation Modes (PSMs) available in Tesseract:
                        # 0: Orientation and script detection (OSD) only.
                        # 1: Automatic page segmentation with OSD.
                        # 2: Automatic page segmentation, but no OSD, or OCR.
                        # 3: Fully automatic page segmentation, but no OSD. (Default)
                        # 4: Assume a single column of text of variable sizes.
                        # 5: Assume a single uniform block of vertically aligned text.
                        # 6: Assume a single uniform block of text.
                        # 7: Treat the image as a single text line.
                        # 8: Treat the image as a single word.
                        # 9: Treat the image as a single word in a circle.
                        # 10: Treat the image as a single character.
                        # 11: Sparse text. Find as much text as possible in no particular order.
                        # 12: Sparse text with OSD.
                        # 13: Raw line. Treat the image as a single text line, bypassing hacks that are Tesseract-specific.

                    # Uncomment the following lines if you'd like to make use of PaddleOCR
                    # if ocr_model == 'PaddleOCR':
                    #     ## Using PaddleOCR
                    #     ocr_result = paddle_ocr.ocr('detected.png', cls = True)

                    if ocr_model == 'EasyOCR':
                    # Using EasyOCR
                        ocr_result = easyocr_reader.readtext('detected.png', detail = 0, allowlist='0123456789')
                        # In EasyOCR, the detail parameter specifies the level of detail in the output. 
                                # When using the readtext method, the detail parameter can be set to different values to control what kind of output you get. 
                                # Specifically:
                                # detail=1: The output will be a list of tuples, where each tuple contains detailed information about the detected text, 
                                # including the bounding box coordinates, the text string, and the confidence score. Example: [(bbox, text, confidence), ...].

                                # detail=0: The output will be a list of strings, where each string is the detected text without any additional details. 
                                # This is a simpler output format that only provides the recognized text. Example: ["text1", "text2", ...].
                        if isinstance(ocr_result, list): # This is because EasyOCR's results are returned as a list
                                ocr_result = ''.join(ocr_result)  # Convert list to a string
                    
                    
                    # Using OCR for handwritten text recognition
                    if ocr_result is not None:
                        
                        if not ocr_result.strip(): # Check if the result is empty or only whitespace. This could be due to the selected OCR (in this case: Tesseract-OCR) not being able to recognize the text in the ROI.
                            # For this reason, we can try another OCR, say for example Easy OCR, to try to recognize the text in this ROI
                            ocr_result = easyocr_reader.readtext('detected.png', detail = 0, allowlist='0123456789')
                            if isinstance(ocr_result, list): # This is because EasyOCR's results are returned as a list
                                ocr_result = ''.join(ocr_result)  # Convert list to a string


                                               
                        # Attain the Ms Excel Template cell coordinates
                        # Determine the cell reference using the x coodrinate of the bounding box, row index, maximum column number, and image/table width
                        cell_ref = calculate_cell_reference(x, row_index, max_columns=24, table_width=image_width) # e.g., A1, B5, etc.
                        
                        # Additional check: Incase the cell (cell_ref) is already occupied with transcribed text. We then opt for the cell below within the same column
                        column_letter = openpyxl.utils.get_column_letter(math.floor(x / image_width * no_of_columns) + 1)
                        initial_row_index = row_index  # Store the initial row index
                        # Check if the cell is already occupied
                        if ws[cell_ref].value is not None:
                            row_index += 1
                            cell_ref = f'{column_letter}{row_index}'
                        
                        # Place the OCR/HTR recognized text in its respective Ms Excel cell 
                        ws[cell_ref].value = ocr_result   

                        # Restore the row index to the initial value
                        row_index = initial_row_index

                        # Set up border styles for excel output
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

                        # Loop through cells to apply borders
                        for row in ws.iter_rows(min_row=1, max_row=no_of_rows, min_col=1, max_col=no_of_columns):
                            for cell in row:
                                cell.border = thin_border
                    

                    else:
                        print('No values detected in clip')
                else:
                    print('ROI is empty or invalid')


        ### SPECIAL ADDITION TO CODE / CUSTOMIZATION: The lines below are particular to our tables. Here we replace the headers and row labels as in the original table format. This will vary from your setup and this should be customized to your particular sheets.
        # Insert two columns on the left side of the excel sheet
        ws.insert_cols(1, 2)
        
        # Insert a new row at the top for headers
        ws.insert_rows(1, amount = 3)

        # Define your headers (adjust as needed)
        headers = ["No de la pentade", "Date", "Bellani (gr. Cal/cm2) 6-6h", "Températures extrêmes", "", "", "", "", "Evaportation en cm3 6 - 6h", "", "Pluies en mm. 6-6h", "Température et Humidité de l'air à 6 heures", "", "", "", "", "Température et Humidité de l'air à 15 heures",  "", "", "", "", "Température et Humidité de l'air à 18 heures",  "", "", "", "", "Date"]
        sub_headers_1 = ["", "", "", "Abri", "", "", "", "", "Piche", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", ""]
        sub_headers_2 =["", "", "", "Max.", "Min.", "(M+m)/2", "Ampl.", "Min. gazon", "Abri.", "Ext.", "", "T", "T'a", "e.", "U", "∆e", "T", "T'a", "e.", "U", "∆e","T", "T'a", "e.", "U", "∆e", ""]

        # Add the headers to the first row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            if header == "No de la pentade" or header == "Date" or header == "Bellani (gr. Cal/cm2) 6-6h" or header == "Pluies en mm. 6-6h":
                cell.alignment = Alignment(textRotation=90)

        # Add the first row of sub-headers to the second row
        for col_num, sub_header in enumerate(sub_headers_1, start=1):
            ws.cell(row=2, column=col_num, value=sub_header)

        # Add the second row of sub-headers to the third row
        for col_num, sub_header in enumerate(sub_headers_2, start=1):
            ws.cell(row=3, column=col_num, value=sub_header)
        
        # Merge cells for multi-column headers
        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1) #No de la pentade
        ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2) #Date
        ws.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3) #Bellani
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8) #Températures extrêmes
        ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10) #Evaportation
        ws.merge_cells(start_row=1, start_column=11, end_row=3, end_column=11) #Pluies
        ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=16) #Température et Humidité de l'air à 6 heures
        ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=21) #Température et Humidité de l'air à 15 heures
        ws.merge_cells(start_row=1, start_column=22, end_row=1, end_column=26) #Température et Humidité de l'air à 18 heures
        ws.merge_cells(start_row=1, start_column=27, end_row=3, end_column=27) #Date
        # subheaders
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7) #Abri
        ws.merge_cells(start_row=2, start_column=9, end_row=2, end_column=10) #Piche
        ws.merge_cells(start_row=2, start_column=12, end_row=2, end_column=16) #(Psychromètre a aspiration)
        ws.merge_cells(start_row=2, start_column=17, end_row=2, end_column=21) #(Psychromètre a aspiration)
        ws.merge_cells(start_row=2, start_column=22, end_row=2, end_column=26) #(Psychromètre a aspiration)


        # Label Date, Total and Average rows
        row_labels = ["1","2", "3", "4", "5", "Tot.", "Moy.", "6", "7", "8", "9", "10", "Tot.", "Moy.", "11", "12", "13", "14", "15", "Tot.", "Moy.", "16", "17", "18", "19", "20", "Tot.", "Moy.", "21", "22", "23", "24", "25", "Tot.", "Moy.", "26", "27", "28", "29", "30", "31", "Tot.", "Moy.", "Tot.", "Moy."]
        # Update the cells in the second and last column with the date values
        columns = [2, 27]
        for col in columns:
            for i, value in enumerate(row_labels, start=4):
                cell = ws.cell(row=i, column=col)
                cell.value = value
                # wb.save(new_version_of_file) # Save the modified workbook
        
        # Save Excel file
        file_path = f'src\output\{method_name}_Excel_with_OCR_Results.xlsx'
        wb.save(file_path)
        results.append([file_path])

        
        ### FOR ILLUSTRATION PURPOSES: Uncomment the following lines if you want to see the ROIs per table
        # plt.imshow(ROIs_image)
        # plt.show()

    return results