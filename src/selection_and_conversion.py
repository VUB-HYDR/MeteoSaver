import os
import openpyxl
import pandas as pd
import re
import matplotlib.pyplot as plt


# Function to extract year, month from filename
def extract_date_from_filename(filename):
    '''
    # Extract the date from the filename

    Parameters
    --------------
    filename: filename of the postprocessed data. These are in the structure/format "STN_YYYYMM_SF" or "STN_YYYYMM_HD" using the data inventory, where:

            STN is the three digit station number,
            YYYY is the year
            MM is the month,
            SF represents Standard Format
            HD represents a hand drawn form / or photocopied form of the standard format

    Returns
    -------------- 
    year, month : Year (YYYY), Month (MM)
    
    '''

    match = re.search(r'_(\d{4})(\d{2})_', filename)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return year, month
    return None, None


def is_highlighted_green(cell, color):
    ''' Checks if a cell is highlighted 'GREEN' during the earlier post processing steps that symbolizes that this data was confirmed 

    Parameters
    --------------   
    cell: coordinates of cell to check
    color: highlighted color

    Returns
    -------------- 
    boolean: 1 (True) if the cell is highlighted with GREEN (i.e. confirmed in Quality Control)

    '''

    fill = cell.fill.start_color
    if isinstance(fill, openpyxl.styles.colors.Color):
        return fill.rgb == color
    return False

# Create a function to plot with dashed lines for missing data
def plot_with_missing(ax, series, label, color):
    # Plot the main line
    ax.plot(series.index, series, label=label, color=color, linestyle='-')
    
    # Create a mask for missing values
    is_nan = series.isna()
    
    # Find start and end points of missing data segments
    missing_segments = []
    start = None
    for i in range(len(is_nan)):
        if is_nan.iloc[i] and start is None:
            start = i
        elif not is_nan.iloc[i] and start is not None:
            missing_segments.append((start, i))
            start = None
    if start is not None:
        missing_segments.append((start, len(is_nan)))
    
    # Plot dashed lines for missing data
    for start, end in missing_segments:
        if start > 0 and end < len(series):
            ax.plot(series.index[start-1:end+1], series[start-1:end+1], linestyle='--', color=color)


def select_and_convert_postprocessed_data(input_folder_path, output_folder_path, station):

    ''' Selects the confirmed data (from the quality control) and converts the excel file to a format ready to be converted to the Station Exchange Format
    Parameters
    --------------   
    input_folder_path: path of postprocessed files
    output_folder_path: output path for the selected data
    station: station number

    Returns
    -------------- 
    Excel files with selected (confirmed) transcribed data -> in new format 
    Timeseries plot of max, min and avg temperature (confirmed) at particular station
    
    '''

    output_file = f'{output_folder_path}\Daily_all_temperatures.xlsx' # Combined output file with the three variables: Max, Min and Average Temperature
    output_file_max = f'{output_folder_path}\Daily_max_temperatures.xlsx' # Output file with Max Temperature
    output_file_min = f'{output_folder_path}\Daily_min_temperatures.xlsx' # Output file with Min Temperature
    output_file_avg = f'{output_folder_path}\Daily_mean_temperatures.xlsx' # Output file with Average Temperature

    # List to hold all data
    data = [] # All the variables

    # Lists to hold all data for each temperature type
    data_max = []
    data_min = []
    data_avg = []

    # Rows to exclude. Adjust these according to your specific sheet
    excluded_rows = [1, 2, 3, 9, 10, 16, 17, 23, 24, 30, 31, 37, 38, 45, 46, 47, 48] # These include titles or 5/6 day totals and averages.

    # Iterate over all files in the folder
    for filename in os.listdir(input_folder_path):
        if filename.endswith(".xlsx"):
            # Extract year and month from filename
            year, month = extract_date_from_filename(filename)
            if year and month:
                file_path = os.path.join(input_folder_path, filename)
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active

                # Extract data from rows and columns, excluding specific rows.  
                for row_num in range(4, worksheet.max_row + 1): #Here this represents Max, Min and Average Temperatures
                    if row_num not in excluded_rows: 
                        day_cell = worksheet.cell(row=row_num, column=2)  # Assuming the day is in the first column
                        max_temperature_cell = worksheet.cell(row=row_num, column=4)  # Column for Max Temperature
                        min_temperature_cell = worksheet.cell(row=row_num, column=5)  # Column for Min Temperature
                        average_temperature_cell = worksheet.cell(row=row_num, column=6)  # Column for Avg Temperature


                        if day_cell.value :
                            day = int(day_cell.value)
                            max_temperature = max_temperature_cell.value if is_highlighted_green(max_temperature_cell, 'FF6DCD57') else 'NaN'
                            min_temperature = min_temperature_cell.value if is_highlighted_green(min_temperature_cell, 'FF6DCD57') else 'NaN'
                            average_temperature = average_temperature_cell.value if is_highlighted_green(average_temperature_cell, 'FF6DCD57') else 'NaN'
                            
                            data.append([year, month, day, max_temperature, min_temperature, average_temperature])

                            data_max.append([year, month, day, max_temperature])
                            data_min.append([year, month, day, min_temperature])
                            data_avg.append([year, month, day, average_temperature])


    # Create a DataFrame from the data
    df = pd.DataFrame(data, columns=["Year", "Month", "Day", "Max_Temperature", "Min_Temperature", "Avg_Temperature"])
    df_max = pd.DataFrame(data_max, columns=["Year", "Month", "Day", "Max_Temperature"])
    df_min = pd.DataFrame(data_min, columns=["Year", "Month", "Day", "Min_Temperature"])
    df_avg = pd.DataFrame(data_avg, columns=["Year", "Month", "Day", "Avg_Temperature"])



    # Generate a complete date range for each year and month combination
    years_months = df[['Year', 'Month']].drop_duplicates()
    complete_data = []

    for _, row in years_months.iterrows():
        year = row['Year']
        month = row['Month']
        num_days = pd.Period(f'{year}-{month}').days_in_month
        for day in range(1, num_days + 1):
            complete_data.append([year, month, day])

    complete_df = pd.DataFrame(complete_data, columns=["Year", "Month", "Day"])

    # Merge the complete date range with the extracted data
    merged_df = pd.merge(complete_df, df, on=["Year", "Month", "Day"])
    merged_df_max = pd.merge(complete_df, df_max, on=["Year", "Month", "Day"])
    merged_df_min = pd.merge(complete_df, df_min, on=["Year", "Month", "Day"])
    merged_df_avg = pd.merge(complete_df, df_avg, on=["Year", "Month", "Day"])


    # Fill missing temperatures with a placeholder value (e.g., NaN or a specific value)
    # For combined sheet
    merged_df['Max_Temperature'] = merged_df['Max_Temperature'].fillna('NaN')  # Since this is temperature, missing vales cannot be zero (0)
    merged_df['Min_Temperature'] = merged_df['Min_Temperature'].fillna('NaN')
    merged_df['Avg_Temperature'] = merged_df['Avg_Temperature'].fillna('NaN')
    
    # For the individual excel sheets
    merged_df_max['Max_Temperature'] = merged_df_max['Max_Temperature'].fillna('NaN')
    merged_df_min['Min_Temperature'] = merged_df_min['Min_Temperature'].fillna('NaN')
    merged_df_avg['Avg_Temperature'] = merged_df_avg['Avg_Temperature'].fillna('NaN')
    
    # Sort DataFrame by Year, Month, Day
    merged_df = merged_df.sort_values(by=['Year', 'Month', 'Day'])
    merged_df_max = merged_df_max.sort_values(by=['Year', 'Month', 'Day'])
    merged_df_min = merged_df_min.sort_values(by=['Year', 'Month', 'Day'])
    merged_df_avg = merged_df_avg.sort_values(by=['Year', 'Month', 'Day'])

    # Save the DataFrame to a new Excel file
    merged_df.to_excel(output_file, index=False)
    merged_df_max.to_excel(output_file_max, index=False)
    merged_df_min.to_excel(output_file_min, index=False)
    merged_df_avg.to_excel(output_file_avg, index=False)

    
    # Convert temperature columns to numeric, coerce errors to NaN.
    merged_df['Max_Temperature'] = pd.to_numeric(merged_df['Max_Temperature'], errors='coerce')
    merged_df['Min_Temperature'] = pd.to_numeric(merged_df['Min_Temperature'], errors='coerce')
    merged_df['Avg_Temperature'] = pd.to_numeric(merged_df['Avg_Temperature'], errors='coerce')


    merged_df_max['Max_Temperature'] = pd.to_numeric(merged_df_max['Max_Temperature'], errors='coerce')
    merged_df_min['Min_Temperature'] = pd.to_numeric(merged_df_min['Min_Temperature'], errors='coerce')
    merged_df_avg['Avg_Temperature'] = pd.to_numeric(merged_df_avg['Avg_Temperature'], errors='coerce')


    # Plot and save the graph
    merged_df['Date'] = pd.to_datetime(merged_df[['Year', 'Month', 'Day']])
    
    # Assuming the standard error is 0.5 for the sake of demonstration
    standard_error = 0.2   # here i used the uncertainity margin allowed in the post processing of the transcribed data
    z = 1.96  # Z-score for 95% confidence interval

    
    #****TWO PLOT OPTIONS***
    # (OPTION 1)
    # ## Plot continous (timeseries) lines with breaks in cases with missing data
    # Do not drop rows with NaN values
    plot_df = merged_df[['Date', 'Max_Temperature', 'Min_Temperature', 'Avg_Temperature']]

    # Set the Date column as the index
    plot_df.set_index('Date', inplace=True)

    # Create a figure and axis for the plot
    fig, ax = plt.subplots(figsize=(10, 6))

    # Plot Max Temperature with confidence interval band using pandas plot function
    plot_df['Max_Temperature'].plot(ax=ax, label='Maximum', color='red')
    ax.fill_between(plot_df.index, plot_df['Max_Temperature'] - z * standard_error, plot_df['Max_Temperature'] + z * standard_error, color='red', alpha=0.2)

    # Plot Min Temperature with confidence interval band using pandas plot function
    plot_df['Min_Temperature'].plot(ax=ax, label='Minimum', color='blue')
    ax.fill_between(plot_df.index, plot_df['Min_Temperature'] - z * standard_error, plot_df['Min_Temperature'] + z * standard_error, color='blue', alpha=0.2)

    # Plot Avg Temperature with confidence interval band using pandas plot function
    plot_df['Avg_Temperature'].plot(ax=ax, label='Average', color='orange')
    ax.fill_between(plot_df.index, plot_df['Avg_Temperature'] - z * standard_error, plot_df['Avg_Temperature'] + z * standard_error, color='orange', alpha=0.2)

    # Set plot labels and title
    ax.set_xlabel('Date')
    ax.set_ylabel('Temperature (°C)')
    ax.set_title('Daily Max, Min, and Avg Temperatures at station ' + str(station))
    ax.legend()
    ax.grid(True)

    # Save the plot
    plt.savefig(f'{output_folder_path}/temperature_plot.jpg', format='jpg')
    plt.show()

    # (OPTION 2)
    # Plot with dashed lines for missing data
    # Create a figure and axis for the plot
    fig, ax = plt.subplots(figsize=(10, 6))

    # Plot Max Temperature with confidence interval band
    plot_with_missing(ax, plot_df['Max_Temperature'], 'Maximum', 'red')

    # Plot Min Temperature with confidence interval band
    plot_with_missing(ax, plot_df['Min_Temperature'], 'Minimum', 'blue')

    # Plot Avg Temperature with confidence interval band
    plot_with_missing(ax, plot_df['Avg_Temperature'], 'Average', 'orange')

    # Set plot labels and title
    ax.set_xlabel('Date')
    ax.set_ylabel('Temperature (°C)')
    ax.set_title('Daily Max, Min, and Avg Temperatures at station ' + str(station))
    ax.legend()
    ax.grid(True)

    # Save the plot
    plt.savefig(f'{output_folder_path}/continous_with_missing_data_as_dashed_temperature_plot.jpg', format='jpg')
    plt.show()


    # (OPTION 3)
    ## Plot continous (timeseries) lines without breaks in cases with missing data
    # Drop rows with NaN values in temperature columns only for plotting
    plot_df = merged_df.dropna(subset=['Max_Temperature', 'Min_Temperature', 'Avg_Temperature'])
    plt.figure(figsize=(10, 6))
    # plt.plot(plot_df['Date'], plot_df['Max_Temperature'], label='Maximum', color = 'red')
    # plt.plot(plot_df['Date'], plot_df['Min_Temperature'], label='Minimum', color = 'blue')
    # plt.plot(plot_df['Date'], plot_df['Avg_Temperature'], label='Average', color = 'orange')

    # Plot Max Temperature with confidence interval band
    plt.plot(plot_df['Date'], plot_df['Max_Temperature'], label='Maximum', color='red')
    plt.fill_between(plot_df['Date'], plot_df['Max_Temperature'] - z * standard_error, plot_df['Max_Temperature'] + z * standard_error, color='red', alpha=0.2)

    # Plot Min Temperature with confidence interval band
    plt.plot(plot_df['Date'], plot_df['Min_Temperature'], label='Minimum', color='blue')
    plt.fill_between(plot_df['Date'], plot_df['Min_Temperature'] - z * standard_error, plot_df['Min_Temperature'] + z * standard_error, color='blue', alpha=0.2)

    # Plot Avg Temperature with confidence interval band
    plt.plot(plot_df['Date'], plot_df['Avg_Temperature'], label='Average', color='orange')
    plt.fill_between(plot_df['Date'], plot_df['Avg_Temperature'] - z * standard_error, plot_df['Avg_Temperature'] + z * standard_error, color='orange', alpha=0.2)

    plt.xlabel('Date')
    plt.ylabel('Temperature(°C)')
    plt.title('Daily Max, Min, and Avg Temperatures at station '+str(station))
    plt.legend()
    plt.grid(True)
    plt.savefig(f'{output_folder_path}/continous_temperature_plot.jpg', format='jpg')
    plt.show()






   
    


