import os
import openpyxl
import pandas as pd
import numpy as np
import re
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
import matplotlib.dates as mdates

def load_data(filepath):
    """Load data from the given Excel file."""
    return pd.read_excel(filepath)


# Function to extract year, month from filename
def extract_date_from_filename(filename):
    '''
    # Extract the date from the filename

    Parameters
    --------------
    filename: filename of the postprocessed data. These are in the structure/format "STN_YYYYMM_SF" or "STN_YYYYMM_HD" using the data inventory, where:

            STN is the three digit station number,
            YYYY is the year
            MM is the month,
            SF represents Standard Format
            HD represents a hand drawn form / or photocopied form of the standard format

    Returns
    -------------- 
    year, month : Year (YYYY), Month (MM)
    
    '''

    match = re.search(r'_(\d{4})(\d{2})_', filename)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return year, month
    return None, None


def is_highlighted_green(cell, color):
    ''' Checks if a cell is highlighted 'GREEN' during the earlier post processing steps that symbolizes that this data was confirmed 

    Parameters
    --------------   
    cell: coordinates of cell to check
    color: highlighted color

    Returns
    -------------- 
    boolean: 1 (True) if the cell is highlighted with GREEN (i.e. confirmed in Quality Control)

    '''

    fill = cell.fill.start_color
    if isinstance(fill, openpyxl.styles.colors.Color):
        return fill.rgb == color
    return False



def select_and_convert_postprocessed_data(filepath, date_column, columns_to_check, header_rows, multi_day_totals, multi_day_averages, excluded_rows, additional_excluded_rows, final_totals_rows):

    ''' 
    Selects confirmed transcribed data after QA/QC checks and converts it into an .xlsx format ready for conversion to the Station Exchange Format (SEF).

    This function processes a QA/QC-verified Excel file, extracting validated temperature data for max, min, and average values. 
    It filters out invalid rows, handles sharp transitions, and identifies outliers. The resulting DataFrame is structured for further analysis or export.

    Parameters
    --------------
    filepath : str
        Path to the QA/QC-verified Excel file.
    date_column : str
        Column letter containing the date values (e.g., 'B').
    columns_to_check : list of str
        List of column letters to extract temperature values (e.g., ['D', 'E', 'F'] for max, min, and average temperatures).
    header_rows : int
        Number of header rows to exclude from processing.
    multi_day_totals : bool
        Whether the dataset includes multi-day total rows.
    multi_day_averages : bool
        Whether the dataset includes multi-day average rows.
    excluded_rows : list of int
        List of row indices to exclude from processing.
    additional_excluded_rows : list of int
        Additional row indices to exclude if multi-day averages are included.
    final_totals_rows : list of int
        Row indices of final totals to exclude if included

    Returns
    --------------
    merged_df : pandas.DataFrame
        A DataFrame containing validated and cleaned daily temperature data with the following columns:
        - "Year": Year of the record.
        - "Month": Month of the record.
        - "Day": Day of the record.
        - "Max_Temperature": Maximum temperature (validated or NaN if invalid/missing).
        - "Min_Temperature": Minimum temperature (validated or NaN if invalid/missing).
        - "Avg_Temperature": Average temperature (validated or NaN if invalid/missing).

    
    '''

    # Extract year and month from the filename
    filename = os.path.basename(filepath)  # Get the filename from the full file path
    year, month = extract_date_from_filename(filename)  # Assuming `extract_date_from_filename` is a utility that parses the year/month from the filename

    if year and month:
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active

    if not (year and month):
        raise ValueError(f"Could not extract year and month from filename: {filename}")


    # List to hold all data
    data = [] # All the variables

    # Lists to hold all data for each temperature type
    data_max = []
    data_min = []
    data_avg = []

    # Rows to exclude. Adjust these according to your specific sheet
    if multi_day_totals and not multi_day_averages:
        excluded_rows =  excluded_rows # These include titles and multi-day (e.g. 5/6) day totals
    if multi_day_totals and multi_day_averages:
        excluded_rows = excluded_rows + additional_excluded_rows # including both multi day totals and averages
    if not multi_day_totals:
        excluded_rows = final_totals_rows # Exlude only the final totals

    # Convert the day column letter and temperature columns to numeric indices
    date_column_idx = ord(date_column) - ord('A') + 1  # Convert 'B' -> 2  # Date
    column_indices = [ord(col.strip()) - ord('A') + 1 for col in columns_to_check] # Max, min and average temperatures 

    # Now `column_indices` will contain [4, 5, 6] for 'D', 'E', 'F'
    max_temp_column_idx = column_indices[0]  # 'D' column index -> Maximum temperature
    min_temp_column_idx = column_indices[1]  # 'E' column index -> Minimum temperature
    avg_temp_column_idx = column_indices[2]  # 'F' column index -> Avergae temperature

    total_transcribed_cells = 0 


    # Extract data from rows and columns, excluding specific rows.  
    for row_num in range(header_rows+1, worksheet.max_row+1): #Here this represents Max, Min and Average Temperatures
        if row_num not in excluded_rows: 
            day_cell = worksheet.cell(row=row_num, column=date_column_idx)  # Assuming the day is in the first column
            max_temperature_cell = worksheet.cell(row=row_num, column=max_temp_column_idx)  # Column for Max Temperature
            min_temperature_cell = worksheet.cell(row=row_num, column=min_temp_column_idx)  # Column for Min Temperature
            average_temperature_cell = worksheet.cell(row=row_num, column=avg_temp_column_idx)  # Column for Avg Temperature

            # Count the total transcribed values, including even the non-confirmed (GREEN) ones
            total_transcribed_cells += sum(cell.value is not None for cell in [max_temperature_cell, min_temperature_cell, average_temperature_cell])

            if day_cell.value is not None and day_cell.value.isdigit():
                day = int(day_cell.value)
                max_temperature = max_temperature_cell.value if is_highlighted_green(max_temperature_cell, 'FF6DCD57') else 'NaN'
                min_temperature = min_temperature_cell.value if is_highlighted_green(min_temperature_cell, 'FF6DCD57') else 'NaN'
                average_temperature = average_temperature_cell.value if is_highlighted_green(average_temperature_cell, 'FF6DCD57') else 'NaN'
                
                data.append([year, month, day, max_temperature, min_temperature, average_temperature])

                data_max.append([year, month, day, max_temperature])
                data_min.append([year, month, day, min_temperature])
                data_avg.append([year, month, day, average_temperature])


    # Create a DataFrame from the data
    df = pd.DataFrame(data, columns=["Year", "Month", "Day", "Max_Temperature", "Min_Temperature", "Avg_Temperature"])


    # Generate a complete date range for each year and month combination
    years_months = df[['Year', 'Month']].drop_duplicates()
    complete_data = []

    for _, row in years_months.iterrows():
        year = row['Year']
        month = row['Month']
        num_days = pd.Period(f'{year}-{month}').days_in_month
        for day in range(1, num_days + 1):
            complete_data.append([year, month, day])

    complete_df = pd.DataFrame(complete_data, columns=["Year", "Month", "Day"])

    # Merge the complete date range with the extracted data
    merged_df = pd.merge(complete_df, df, on=["Year", "Month", "Day"], how='left')
   
    # Fill missing temperatures with a placeholder value (e.g., NaN or a specific value)
    for column in ["Max_Temperature", "Min_Temperature", "Avg_Temperature"]: # Since this is temperature, missing vales cannot be zero (0)
        merged_df[column] = merged_df[column].fillna(np.nan)
    
    # Convert temperature columns to numeric, coerce errors to NaN.
    merged_df['Max_Temperature'] = pd.to_numeric(merged_df['Max_Temperature'], errors='coerce')
    merged_df['Min_Temperature'] = pd.to_numeric(merged_df['Min_Temperature'], errors='coerce')
    merged_df['Avg_Temperature'] = pd.to_numeric(merged_df['Avg_Temperature'], errors='coerce')

    # Sort DataFrame by Year, Month, Day
    merged_df = merged_df.sort_values(by=['Year', 'Month', 'Day'])

    # Standard deviation and outlier detection
    for column in ["Max_Temperature", "Min_Temperature", "Avg_Temperature"]:
        # Calculate the standard deviation and mean for the column
        std = merged_df[column].std()
        mean = merged_df[column].mean()

        # # Condition 1: Remove values more than 3 standard deviations from the mean
        # merged_df[column] = merged_df[column].apply(lambda x: np.nan if abs(x - mean) > 3 * std else x)

        # Condition 2: Detect sharp transitions between days (e.g., +4sd to -4sd)
        # Calculate the standard deviation differences for each day relative to the mean
        merged_df['std_diff'] = (merged_df[column] - mean) / std
        for i in range(1, len(merged_df) - 1):  # Avoid the first and last rows to prevent boundary issues
            prev_std_diff = merged_df.loc[i - 1, 'std_diff'] if not pd.isna(merged_df.loc[i - 1, 'std_diff']) else 0   # std deviation difference of previous day
            curr_std_diff = merged_df.loc[i, 'std_diff'] # std deviation difference of current day
            next_std_diff = merged_df.loc[i + 1, 'std_diff'] if not pd.isna(merged_df.loc[i + 1, 'std_diff']) else 0 # std deviation difference of following day

            if not pd.isna(curr_std_diff):
                # Detect sharp opposite changes (e.g., large negative difference to large positive difference or vice versa)
                if (prev_std_diff < -4 and curr_std_diff > 4 and next_std_diff < -4) or \
                    (prev_std_diff > 4 and curr_std_diff < -4 and next_std_diff > 4):
                    merged_df.loc[i, column] = np.nan  # Mark the current value as an outlier

        # Drop the temporary 'std_diff' column
        merged_df.drop(columns=['std_diff'], inplace=True)


    print(f"Total automatically transcribed Cells: {total_transcribed_cells}")

    return merged_df


def select_and_convert_manually_transcribed_data(filepath, date_column, columns_to_check, header_rows, multi_day_totals, multi_day_averages, excluded_rows, additional_excluded_rows, final_totals_rows):

    ''' 
    Processes and selects manually transcribed meteorological data for analysis and conversion to a standardized format.

    This function reads an Excel file containing manually transcribed meteorological observations, selects relevant 
    data columns, excludes rows as specified, and generates a complete time series. Missing values are handled, 
    and the data is returned as a structured DataFrame, ready for conversion to the Station Exchange Format (SEF).

    Parameters
    ----------
    filepath : str
        Path to the manually transcribed data sheet (Excel file).
    date_column : str
        Column letter representing the day/date of observation (e.g., 'B').
    columns_to_check : list of str
        List of column letters to extract (e.g., ['D', 'E', 'F'] for Max, Min, and Average temperatures).
    header_rows : int
        Number of rows at the top of the sheet used as headers (not part of the data).
    multi_day_totals : bool
        Indicates whether the sheet contains multi-day totals that should be excluded.
    multi_day_averages : bool
        Indicates whether the sheet contains multi-day averages that should be excluded.
    excluded_rows : list of int
        Row indices to exclude from the processing (e.g., rows with multi-day totals).
    additional_excluded_rows : list of int
        Additional rows to exclude if multi-day averages are present.
    final_totals_rows : list of int
        Rows containing final totals to exclude if no multi-day totals are present.

    Returns
    -------
    pandas.DataFrame
        A DataFrame containing the processed data with columns for year, month, day, 
        maximum temperature, minimum temperature, and average temperature.
    
    '''

    # Extract year and month from the filename
    filename = os.path.basename(filepath)  # Get the filename from the full file path
    year, month = extract_date_from_filename(filename)  # Assuming `extract_date_from_filename` is a utility that parses the year/month from the filename

    if year and month:
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active

    if not (year and month):
        raise ValueError(f"Could not extract year and month from filename: {filename}")


    # List to hold all data
    data = [] # All the variables

    # Lists to hold all data for each temperature type
    data_max = []
    data_min = []
    data_avg = []

    # Rows to exclude. Adjust these according to your specific sheet
    if multi_day_totals and not multi_day_averages:
        excluded_rows =  excluded_rows # These include titles and multi-day (e.g. 5/6) day totals
    if multi_day_totals and multi_day_averages:
        excluded_rows =  excluded_rows + additional_excluded_rows # including both multi day totals and averages
    if not multi_day_totals:
        excluded_rows = final_totals_rows # Exlude only the final totals

    # Convert the day column letter and temperature columns to numeric indices
    date_column_idx = ord(date_column) - ord('A') + 1  # Convert 'B' -> 2  # Date
    column_indices = [ord(col.strip()) - ord('A') + 1 for col in columns_to_check] # Max, min and average temperatures 

    # Now `column_indices` will contain [4, 5, 6] for 'D', 'E', 'F'
    max_temp_column_idx = column_indices[0]  # 'D' column index -> Maximum temperature
    min_temp_column_idx = column_indices[1]  # 'E' column index -> Minimum temperature
    avg_temp_column_idx = column_indices[2]  # 'F' column index -> Avergae temperature

    total_transcribed_cells = 0 

    # Extract data from rows and columns, excluding specific rows.  
    for row_num in range(header_rows+1, worksheet.max_row+1): #Here this represents Max, Min and Average Temperatures
        if row_num not in excluded_rows: 
            day_cell = worksheet.cell(row=row_num, column=date_column_idx)  # Assuming the day is in the first column
            max_temperature_cell = worksheet.cell(row=row_num, column=max_temp_column_idx)  # Column for Max Temperature
            min_temperature_cell = worksheet.cell(row=row_num, column=min_temp_column_idx)  # Column for Min Temperature
            average_temperature_cell = worksheet.cell(row=row_num, column=avg_temp_column_idx)  # Column for Avg Temperature

            # Count the total manually transcribed values
            total_transcribed_cells += sum(cell.value is not None for cell in [max_temperature_cell, min_temperature_cell, average_temperature_cell])

            if day_cell.value is not None and day_cell.value.isdigit():
                day = int(day_cell.value)
                max_temperature = float(max_temperature_cell.value) if max_temperature_cell.value is not None else 'NaN'
                min_temperature = float(min_temperature_cell.value) if min_temperature_cell.value is not None else 'NaN'
                average_temperature = float(average_temperature_cell.value) if average_temperature_cell.value is not None else 'NaN'
                
                data.append([year, month, day, max_temperature, min_temperature, average_temperature])

                data_max.append([year, month, day, max_temperature])
                data_min.append([year, month, day, min_temperature])
                data_avg.append([year, month, day, average_temperature])


    # Create a DataFrame from the data
    df = pd.DataFrame(data, columns=["Year", "Month", "Day", "Max_Temperature", "Min_Temperature", "Avg_Temperature"])


    # Generate a complete date range for each year and month combination
    years_months = df[['Year', 'Month']].drop_duplicates()
    complete_data = []

    for _, row in years_months.iterrows():
        year = row['Year']
        month = row['Month']
        num_days = pd.Period(f'{year}-{month}').days_in_month
        for day in range(1, num_days + 1):
            complete_data.append([year, month, day])

    complete_df = pd.DataFrame(complete_data, columns=["Year", "Month", "Day"])

    # Merge the complete date range with the extracted data
    merged_df = pd.merge(complete_df, df, on=["Year", "Month", "Day"], how='left')
   
    # Fill missing temperatures with a placeholder value (e.g., NaN or a specific value)
    for column in ["Max_Temperature", "Min_Temperature", "Avg_Temperature"]: # Since this is temperature, missing vales cannot be zero (0)
        merged_df[column] = merged_df[column].fillna(np.nan)
    

    # Sort DataFrame by Year, Month, Day
    merged_df = merged_df.sort_values(by=['Year', 'Month', 'Day'])

    print(f"Total manually transcribed Cells: {total_transcribed_cells}")

    return merged_df


def compare_dataframes(df1, df2, uncertainty_margin): 
    """
    Compare temperature data between two dataframes and calculate accuracy percentage and mean absolute error (MAE).

    This function evaluates the similarity between temperature values in two dataframes, typically one containing 
    manually transcribed data and the other containing automatically transcribed post-processed data. It calculates the percentage of values 
    that fall within a specified uncertainty margin and computes the mean absolute error (MAE) for the compared data.

    Parameters
    ----------
    df1 : pandas.DataFrame
        The first dataframe, generally containing manually transcribed data. It should have columns for 'Year', 
        'Month', 'Day', and temperature variables ('Max_Temperature', 'Min_Temperature', 'Avg_Temperature').
    df2 : pandas.DataFrame
        The second dataframe, typically containing automatically transcribed data. It must match 
        the structure and columns of df1.
    uncertainty_margin : float, optional
        The allowable difference between corresponding values in the two dataframes for them to be considered a match. 

    Returns
    -------
    tuple
        A tuple containing:
        - accuracy_percentage (float): The percentage of values that fall within the uncertainty margin.
        - mae (float): The mean absolute error of the compared values.
    """
    
    total_highlighted_cells = 0
    accurate_matches = 0
    absolute_errors = []  # List to store absolute errors for MAE calculation

    # Ensure the dataframes are aligned on the same date
    merged_df = pd.merge(df1, df2, on=['Year', 'Month', 'Day'], suffixes=('_manual', '_post'))

    # Convert temperature columns to numeric, coercing errors to NaN
    for col in ['Max_Temperature', 'Min_Temperature', 'Avg_Temperature']:
        merged_df[col + '_manual'] = pd.to_numeric(merged_df[col + '_manual'], errors='coerce')
        merged_df[col + '_post'] = pd.to_numeric(merged_df[col + '_post'], errors='coerce')

    # Iterate through the relevant columns to compare values
    for col in ['Max_Temperature', 'Min_Temperature', 'Avg_Temperature']:
        col_manual = col + '_manual'
        col_post = col + '_post'

        for i in range(len(merged_df)):
            if not pd.isna(merged_df[col_manual].iloc[i]) and not pd.isna(merged_df[col_post].iloc[i]):
                total_highlighted_cells += 1
                difference = abs(merged_df[col_manual].iloc[i] - merged_df[col_post].iloc[i])
                
                # Add the absolute difference to the list for MAE calculation
                absolute_errors.append(difference)

                # Check if the difference is within the uncertainty margin
                if difference <= uncertainty_margin:
                    accurate_matches += 1

    # Calculate accuracy percentage
    accuracy_percentage = (accurate_matches / total_highlighted_cells) * 100 if total_highlighted_cells > 0 else 0.0
    
    # Calculate Mean Absolute Error (MAE)
    mae = sum(absolute_errors) / len(absolute_errors) if absolute_errors else 0.0

    print(f"Total Highlighted Cells: {total_highlighted_cells}")
    print(f"Accuracy Percentage: {accuracy_percentage:.2f}%")
    print(f"Mean Absolute Error (MAE): {mae:.2f}")

    return accuracy_percentage, mae


def plot_comparison(manual_df, post_processed_df, output_folder_path, station, post_file, accuracy_percentage_and_mean_absolute_error, uncertainty_margin):
    """
    Generate a comparison plot for daily temperature data from manually transcribed and post-processed datasets.

    This function creates a time series plot comparing daily maximum, minimum, and average temperatures from two datasets:
    manually transcribed data and automatically transcribed post-processed data. The post-processed data is displayed as solid markers, while the
    manually transcribed data is visualized as lighter bands representing the uncertainty margin. The plot highlights 
    the accuracy percentage and mean absolute error (MAE) between the datasets.

    Parameters
    ----------
    manual_df : pandas.DataFrame
        Dataframe containing manually transcribed temperature data, with columns for 'Year', 'Month', 'Day', 
        and temperature variables ('Max_Temperature', 'Min_Temperature', 'Avg_Temperature').
    post_processed_df : pandas.DataFrame
        Dataframe containing post-processed temperature data, with the same structure as `manual_df`.
    output_folder_path : str
        Path to the directory where the generated plot will be saved.
    station : str
        Name or identifier of the station (station number), used for labeling the plot and the output file.
    post_file : str
        Filename of the post-processed data file, used to generate a unique name for the saved plot.
    accuracy_percentage_and_mean_absolute_error : tuple
        A tuple containing the accuracy percentage (float) and mean absolute error (float) calculated from the 
        comparison of the datasets.
    uncertainty_margin : float
        The uncertainty margin applied to the manually transcribed data to create the lighter bands in the plot.

    Returns
    -------
    None
        This function saves the generated plot as an image file in the specified output directory and displays it.


    """


    # Ensure Date column exists in both dataframes
    manual_df['Date'] = pd.to_datetime(manual_df[['Year', 'Month', 'Day']])
    post_processed_df['Date'] = pd.to_datetime(post_processed_df[['Year', 'Month', 'Day']])
    
    # Merge data on Date
    merged_df = pd.merge(post_processed_df, manual_df, on='Date', suffixes=('_post', '_manual'))
    
    # Create a figure and axis for the plot
    fig, ax = plt.subplots(figsize=(12, 8))

    # Plot post-QA/QC data using MeteoSaver (Max, Min, Avg) as dotted lines ('--')
    ax.plot(merged_df['Date'], merged_df['Max_Temperature_post'], 'o', label='Max (Automatically transcribed)', color='red')
    ax.plot(merged_df['Date'], merged_df['Min_Temperature_post'], 'o', label='Min (Automatically transcribed)', color='blue')
    ax.plot(merged_df['Date'], merged_df['Avg_Temperature_post'], 'o', label='Avg (Automatically transcribed)', color='orange')


    # Fill uncertainty margins around post-QA/QC data
    ax.fill_between(merged_df['Date'], merged_df['Max_Temperature_manual'] - uncertainty_margin, 
                    merged_df['Max_Temperature_manual'] + uncertainty_margin, color='red', alpha=0.2)
    ax.fill_between(merged_df['Date'], merged_df['Min_Temperature_manual'] - uncertainty_margin, 
                    merged_df['Min_Temperature_manual'] + uncertainty_margin, color='blue', alpha=0.2)
    ax.fill_between(merged_df['Date'], merged_df['Avg_Temperature_manual'] - uncertainty_margin, 
                    merged_df['Avg_Temperature_manual'] + uncertainty_margin, color='orange', alpha=0.2)
    
    # Set plot labels and title
    ax.set_xlabel('Date', fontsize=17, labelpad=15)
    ax.set_ylabel('Temperature (Â°C)', fontsize=17, labelpad=15)
    ax.set_title(f'Daily Max, Min, and Avg Temperatures at station {station}')

    # Adjust tick label font size
    ax.tick_params(axis='x', labelsize=17)  # Set font size for x-axis tick labels
    ax.tick_params(axis='y', labelsize=17)  # Set font size for y-axis tick labels

    # Adjust the y-axis limit to ensure the legend is above the plotted lines
    ylim = ax.get_ylim()
    y_max = max(ylim)
    legend_y_position = y_max + (y_max * 0.1)
    ax.set_ylim(ylim[0], legend_y_position)

    # Format the date labels to show at intervals of 5 days
    # Locator for every 5 days
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=5))
    plt.xticks(rotation=45)  # Rotate x-axis tick labels by 45 degrees

    # Add accuracy percentage below the x-axis
    plt.figtext(0.70, 0.92, f'Accuracy Percentage: {accuracy_percentage_and_mean_absolute_error[0]:.1f}%', ha='left', fontsize=15, bbox={"facecolor":"orange", "alpha":0.5, "pad":5})
    plt.figtext(0.70, 0.87, f'Mean Absolute Error: {accuracy_percentage_and_mean_absolute_error[1]:.1f}', ha='left', fontsize=15, bbox={"facecolor":"orange", "alpha":0.5, "pad":5})

    # # LEGEND
    # # Plot manually transcribed data as points with 'o' markers
    # manual_handle = [
    # Line2D([0], [0], marker='o', color='red', label='Max (Automatically transcribed)', linestyle='None'),
    # Line2D([0], [0], marker='o', color='blue', label='Min (Automatically transcribed)', linestyle='None'),
    # Line2D([0], [0], marker='o', color='orange', label='Avg (Automatically transcribed)', linestyle='None')]

    # # Create custom legend entries for automatically transcribed data + uncertainty
    # custom_handle = [
    # ( Patch(facecolor='red', alpha=0.2)),
    # (Patch(facecolor='blue', alpha=0.2)),
    # (Patch(facecolor='orange', alpha=0.2))]

    # # Combine the manual transcribed and automatically transcribed legend handles
    # all_handles =  manual_handle + custom_handle

    # # Create custom labels
    # labels = [
    #     'Max (Automatically transcribed)', 'Min (Automatically transcribed)', 'Avg (Automatically transcribed)', 
    #     'Max (Manually transcribed)', 'Min (Manually transcribed)', 'Avg (Manually transcribed)'
    # ]

    # # Set legend handles with custom lines for each temperature type
    # ax.legend(handles=all_handles, labels=labels, fontsize='small', loc='upper right')

    # Add the legend below the plot
    # ax.legend(handles=all_handles, labels=labels, loc='upper center', bbox_to_anchor=(0.5, -0.40), ncol=2, fontsize=14)
        
    ax.grid(True)

    # # Add accuracy percentage below the x-axis
    # plt.figtext(0.40, 0.01, f'Accuracy Percentage: {accuracy_percentage_and_mean_absolute_error[0]:.1f}%', ha='center', fontsize=10, bbox={"facecolor":"orange", "alpha":0.5, "pad":5})
    # plt.figtext(0.60, 0.01, f'Mean Absolute Error: {accuracy_percentage_and_mean_absolute_error[1]:.1f}', ha='center', fontsize=10, bbox={"facecolor":"orange", "alpha":0.5, "pad":5})
    
    # Remove the '.xlsx' extension from the post_file to clean the filename
    cleaned_post_file_name = os.path.splitext(post_file)[0]

    plt.tight_layout()
    # Save the plot
    plot_filename = os.path.join(output_folder_path, f'temperature_comparison_plot_{cleaned_post_file_name}.jpg')
    plt.savefig(plot_filename, format='jpg')
    plt.show()



def validate(manually_transcribed_data_dir_station, postprocessed_data_dir_station, output_folder_path, station, date_column, columns_to_check, header_rows, multi_day_totals, multi_day_averages, excluded_rows, additional_excluded_rows, final_totals_rows, uncertainty_margin):
    """
    Validate and compare manually transcribed data with post-processed data, generating accuracy metrics and visualizations.

    This function performs validation by comparing manually transcribed meteorological data with automatically transcribed post-processed data. 
    It matches corresponding files, processes the data, calculates accuracy metrics such as the accuracy percentage 
    and mean absolute error (MAE), and generates time series plots for visual comparison.

    Parameters
    ----------
    manually_transcribed_data_dir_station : str
        Directory containing manually transcribed data files. These files should include the suffix '_manually_entered_temperatures'.
    postprocessed_data_dir_station : str
        Directory containing post-processed data files, named with the suffix '_post_QA_QC'.
    output_folder_path : str
        Path where the generated plots will be saved.
    station : str
        Identifier or name of the station (station no.), used for labeling plots and output files.
    date_column : str
        The column name or letter indicating the date in the input files.
    columns_to_check : list of str
        List of column names or letters corresponding to maximum, minimum, and average temperatures.
    header_rows : int
        Number of header rows in the input files.
    multi_day_totals : bool
        Whether multi-day totals are present in the input files.
    multi_day_averages : bool
        Whether multi-day averages are present in the input files.
    excluded_rows : list of int
        Rows to exclude during processing, such as title rows or non-temperature rows.
    additional_excluded_rows : list of int
        Additional rows to exclude for multi-day totals and averages.
    final_totals_rows : list of int
        Rows containing final totals to exclude if multi-day totals are not included.
    uncertainty_margin : float
        Allowable difference between manually transcribed and post-processed values for validation.

    Returns
    -------
    None
        This function does not return a value but processes the data, calculates metrics, and generates plots.


    """

    manually_transcribed_files = os.listdir(manually_transcribed_data_dir_station)
    postprocessed_files = os.listdir(postprocessed_data_dir_station)
    manual_files = [f for f in manually_transcribed_files if 'manually_entered' in f]
    post_processed_files = [f for f in postprocessed_files if 'post_QA_QC' in f]

    # Create a dictionary to map manual files to post-processed files
    file_pairs = {}
    for manual_file in manual_files:
        base_name = manual_file.replace('_manually_entered_temperatures', '').replace('.xlsx', '')
        corresponding_post_file = f'{base_name}_post_QA_QC.xlsx'
        if corresponding_post_file in post_processed_files:
            file_pairs[manual_file] = corresponding_post_file
        else:
            print(f"Post-processed file not found for {manual_file}")

    if not file_pairs:
        print("No matching files found.")
        return

    # Process each file pair
    for manual_file, post_file in file_pairs.items():
        print(f"Processing pair: {manual_file} and {post_file}")
        manual_filepath = os.path.join(manually_transcribed_data_dir_station, manual_file)
        print(manual_filepath)
        post_processed_filepath = os.path.join(postprocessed_data_dir_station, post_file)

        # Load data
        # Manually transcribed data
        manual_df = select_and_convert_manually_transcribed_data(manual_filepath, date_column, columns_to_check, header_rows, multi_day_totals, multi_day_averages, excluded_rows, additional_excluded_rows, final_totals_rows)
        # Automatically transcribed data
        post_processed_df = select_and_convert_postprocessed_data(post_processed_filepath, date_column, columns_to_check, header_rows, multi_day_totals, multi_day_averages, excluded_rows, additional_excluded_rows, final_totals_rows)

        # # Convert temperature columns to numeric, coerce errors to NaN
        # manual_df['Max_Temperature'] = pd.to_numeric(manual_df['Max_Temperature'], errors='coerce')
        # manual_df['Min_Temperature'] = pd.to_numeric(manual_df['Max_Temperature'], errors='coerce')
        # manual_df['Avg_Temperature'] = pd.to_numeric(manual_df['Max_Temperature'], errors='coerce')

        # Convert temperature columns to numeric, coerce errors to NaN for post-processed data
        post_processed_df['Max_Temperature'] = pd.to_numeric(post_processed_df['Max_Temperature'], errors='coerce')
        post_processed_df['Min_Temperature'] = pd.to_numeric(post_processed_df['Min_Temperature'], errors='coerce')
        post_processed_df['Avg_Temperature'] = pd.to_numeric(post_processed_df['Avg_Temperature'], errors='coerce')

        # Check 
        print(manual_df[['Year', 'Month', 'Day']].head())
        print(post_processed_df[['Year', 'Month', 'Day']].head())

        # Compare the dataframes and get accuracy percentage
        accuracy_percentage_and_mean_absolute_error = compare_dataframes(manual_df, post_processed_df, uncertainty_margin)

        # Plot the comparison
        plot_comparison(manual_df, post_processed_df, output_folder_path, station, post_file, accuracy_percentage_and_mean_absolute_error, uncertainty_margin)