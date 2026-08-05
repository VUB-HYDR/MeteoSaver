import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import re
import matplotlib.pyplot as plt
from matplotlib.colors import Normalize
from matplotlib.ticker import MaxNLocator
import matplotlib.cm as cm
import numpy as np
from sklearn.linear_model import LinearRegression
from scipy.stats import theilslopes
from scipy.stats import gaussian_kde
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from scipy.interpolate import griddata
from cartopy.io.shapereader import Reader
from cartopy.feature import ShapelyFeature


# Function to extract year, month from filename
def extract_date_from_filename(filename):
    '''
    # Extract the date from the filename

    Parameters
    --------------
    filename: filename of the postprocessed data. These are in the structure/format "STN_YYYYMM_SF" or "STN_YYYYMM_HD" using the data inventory, where:

            STN is the three digit station number,
            YYYY is the year
            MM is the month,
            SF represents Standard Format
            HD represents a hand drawn form / or photocopied form of the standard format

    Returns
    -------------- 
    year, month : Year (YYYY), Month (MM)
    
    '''

    match = re.search(r'_(\d{4})(\d{2})_', filename)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return year, month
    return None, None


def is_highlighted_green(cell, color):
    ''' Checks if a cell is highlighted 'GREEN' during the earlier post processing steps that symbolizes that this data was confirmed 

    Parameters
    --------------   
    cell: coordinates of cell to check
    color: highlighted color

    Returns
    -------------- 
    boolean: 1 (True) if the cell is highlighted with GREEN (i.e. confirmed in Quality Control)

    '''

    fill = cell.fill.start_color
    if isinstance(fill, openpyxl.styles.colors.Color):
        return fill.rgb == color
    return False


def dms_to_decimal(dms_str):
    """
    Convert a DMS (Degrees, Minutes, Seconds) string to decimal degrees.

    This function takes a string representing geographic coordinates in the 
    Degrees-Minutes-Seconds (DMS) format and converts it to a decimal degrees 
    representation. It accounts for directional indicators (N, S, E, W) 
    to determine the sign of the decimal value.

    Parameters
    ----------
    dms_str : str
        The DMS string to be converted. The expected format includes an optional 
        direction (N/S/E/W), followed by degrees (°) and minutes ("), e.g., "N 45°30".

    Returns
    -------
    float
        The converted decimal degrees value, rounded to four decimal places. Returns
        NaN if the input is not a valid DMS string.

    Raises
    ------
    ValueError
        If the input string does not match the expected DMS format.

    
    For example:
    --------
    dms_to_decimal("N 45°30") to 45.5000
    """

    if not isinstance(dms_str, str):
        # Return NaN if the input is not a valid string
        return np.nan
    
    # Check if there is a direction (N/S/E/W)
    match = re.match(r'([NSWE])?\s*(\d+)°(\d+)', dms_str)
    if not match:
        raise ValueError(f"Invalid DMS format: {dms_str}")
    
    direction, degrees, minutes = match.groups()
    decimal = int(degrees) + int(minutes) / 60
    
    # Make the decimal negative for S and W directions if direction is specified
    if direction in ['S', 'W']:
        decimal = -decimal
    
    return round(decimal, 4)

def load_station_metadata(file_path, sheet_name='Stations'):
    """
    Load station metadata from an Excel file and convert geographic coordinates to decimal degrees.

    This function reads station metadata from a specified Excel file and processes it to standardize
    column names, trim whitespace, and convert latitude and longitude values from Degrees-Minutes-Seconds 
    (DMS) format to decimal degrees. It returns a cleaned DataFrame with essential station details.

    Parameters
    ----------
    file_path : str
        The file path to the Excel file containing the station metadata.
    sheet_name : str, optional
        The name of the Excel sheet to read. Defaults to 'Stations'.

    Returns
    -------
    pandas.DataFrame
        A DataFrame containing cleaned and processed station metadata with the following columns:
        - 'ID': Unique identifier for each station (station no.).
        - 'name': Station name.
        - 'latitude': Latitude in decimal degrees.
        - 'longitude': Longitude in decimal degrees.
        - 'altitude': Altitude in meters.
    """
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    df = df.rename(columns={
        'Station': 'name',
        'ID': 'ID',
        'Latitude': 'latitude',
        'Longitude': 'longitude',
        'Altitude': 'altitude'
    })
    
    # Trim whitespace from IDs and ensure they're all strings
    df['ID'] = df['ID'].astype(str).str.strip().str.zfill(3)

    # Convert latitude and longitude to decimal degrees using dms_to_decimal
    df['latitude'] = df['latitude'].apply(dms_to_decimal)
    df['longitude'] = df['longitude'].apply(dms_to_decimal)
    
    return df[['ID', 'name', 'latitude', 'longitude', 'altitude']]




def analyze_temperature_trends_with_linear_regression(df, output_folder_path, station, station_name):
    """
    Analyzes and plots temperature trends over time using linear regression.

    Parameters
    ----------
    output_folder_path : str
        Path where the processed temperature dataset is stored.
    station : str
        Unique station identifier.

    Returns
    -------
    trend_slopes : dict
        Slopes of the temperature trends for Max, Min, and Avg temperatures.
    """

    # Add combined date and numeric year
    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")
    df["Year_Num"] = df["Year"].astype(int)

    plt.figure(figsize=(12, 6))
    colors = {"Max_Temperature": "red", "Min_Temperature": "blue", "Avg_Temperature": "orange"}
    trend_colors = {"Max_Temperature": "darkred", "Min_Temperature": "darkblue", "Avg_Temperature": "darkorange"}

    models = {}
    trend_slopes = {}

    for temp_type in ["Max_Temperature", "Min_Temperature", "Avg_Temperature"]:
        df_temp = df.dropna(subset=[temp_type]).copy()
        if df_temp.empty:
            # print(f"[Warning] No valid {temp_type} data at station {station}. Skipping.")
            continue

        if df_temp["Year_Num"].nunique() < 2:
            # print(f"[Warning] Not enough variation in year data for {temp_type} at station {station}. Skipping.")
            continue

        X = df_temp["Year_Num"].values.reshape(-1, 1)
        y = df_temp[temp_type].values

        model = LinearRegression().fit(X, y)
        pred = model.predict(X)
        models[temp_type] = model
        trend_slopes[temp_type] = model.coef_[0]

        # Plot original data
        plt.scatter(df_temp["Date"], y, color=colors[temp_type], alpha=0.5, s=10, label=f"{temp_type.replace('_', ' ')}")
        plt.plot(df_temp["Date"], pred, color=trend_colors[temp_type], linewidth=2, label=f"{temp_type.replace('_', ' ')} Linear")

        # Annotate slope
        last_date = df_temp["Date"].iloc[-1]
        last_temp = pred[-1]
        if np.isfinite(last_temp) and pd.notnull(last_date):
            plt.text(last_date, last_temp, f"{model.coef_[0]:.2f}°C/yr", 
                     fontsize=10, color=trend_colors[temp_type], fontweight='bold')

    if not models:
        # print(f"[Error] No valid data for regression at station {station}. Skipping plot.")
        return None

    plt.xlabel("Year")
    plt.ylabel("Temperature (°C)")
    plt.title(f"Temperature Trends at Station: {station_name} - using Linear Regression")
    plt.legend()
    plt.grid(True)

    os.makedirs(output_folder_path, exist_ok=True)
    trend_plot_path = os.path.join(output_folder_path, f"temperature_trend_plot_{station}_using_linear_regression.jpg")
    plt.savefig(trend_plot_path, format="jpg", dpi=300)
    # print(f"Figure saved at: {trend_plot_path}")
    plt.close()

    return trend_slopes



def analyze_temperature_trends_with_theilsen(df, output_folder_path, station, station_name):
    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")
    df["Year_Num"] = df["Year"].astype(int)

    plt.figure(figsize=(12, 6))
    colors = {"Max_Temperature": "red", "Min_Temperature": "blue", "Avg_Temperature": "orange"}
    trend_colors = {"Max_Temperature": "darkred", "Min_Temperature": "darkblue", "Avg_Temperature": "darkorange"}

    trend_slopes = {}

    for temp_type in ["Max_Temperature", "Min_Temperature", "Avg_Temperature"]:
        df_temp = df.dropna(subset=[temp_type]).copy()

        if df_temp.empty:
            # print(f"[Warning] No valid {temp_type} data at station {station}. Skipping.")
            continue

        if df_temp["Year_Num"].nunique() < 2:
            # print(f"[Warning] Not enough variation in year data for {temp_type} at station {station}. Skipping.")
            continue

        x = df_temp["Year_Num"].values
        y = df_temp[temp_type].values

        # Linear Regression
        model = LinearRegression().fit(x.reshape(-1, 1), y)
        linear_pred = model.predict(x.reshape(-1, 1))
        linear_slope = model.coef_[0]

        # Theil–Sen Estimator
        theil_slope, intercept, _, _ = theilslopes(y, x, 0.95)
        theil_pred = intercept + theil_slope * x

        # Use consistent keys matching your final trend_df
        key_map = {
            "Max_Temperature": "trend_max_temperature",
            "Min_Temperature": "trend_min_temperature",
            "Avg_Temperature": "trend_avg_temperature"
        }

        trend_slopes[key_map[temp_type]] = {
            "linear_slope": linear_slope,
            "theil_slope": theil_slope
        }

        plt.scatter(df_temp["Date"], y, color=colors[temp_type], alpha=0.4, s=10, label=f"{temp_type.replace('_', ' ')}")
        plt.plot(df_temp["Date"], theil_pred, color=trend_colors[temp_type], linewidth=2, label=f"{temp_type.replace('_', ' ')} Theil–Sen")

        plt.text(df_temp["Date"].iloc[-1] - pd.Timedelta(days=500),
                 theil_pred[-1],
                 f"{theil_slope:.2f}°C/yr",
                 fontsize=9,
                 color=trend_colors[temp_type],
                 fontweight='bold',
                 bbox=dict(boxstyle='round,pad=0.3', facecolor='white', edgecolor=trend_colors[temp_type], alpha=0.6))

    plt.xlabel("Year")
    plt.ylabel("Temperature (°C)")
    plt.title(f"Temperature Trends at Station: {station_name} - using Theil–Sen Estimator")
    plt.legend()
    plt.grid(True)

    os.makedirs(output_folder_path, exist_ok=True)
    plot_path = os.path.join(output_folder_path, f"temperature_trend_{station}_using_theilsen.jpg")
    plt.savefig(plot_path, dpi=300)
    # print(f"Theil–Sen trend plot saved at: {plot_path}")
    plt.close()

    # === Compute TXx (annual max of daily max temp) and TNn (annual min of daily min temp) ===
    extreme_trends = {}

    # TXx
    if "Max_Temperature" in df.columns:
        txx_df = df.dropna(subset=["Max_Temperature"])
        txx_annual = txx_df.groupby("Year_Num")["Max_Temperature"].max().dropna()
        if txx_annual.index.nunique() >= 2:
            x = txx_annual.index.values
            y = txx_annual.values
            theil_slope, _, _, _ = theilslopes(y, x, 0.95)
            extreme_trends["trend_TXx"] = theil_slope
        else:
            # print(f"[Warning] Not enough years for TXx trend at {station}")
            extreme_trends["trend_TXx"] = np.nan

    # TNn
    if "Min_Temperature" in df.columns:
        tnn_df = df.dropna(subset=["Min_Temperature"])
        tnn_annual = tnn_df.groupby("Year_Num")["Min_Temperature"].min().dropna()
        if tnn_annual.index.nunique() >= 2:
            x = tnn_annual.index.values
            y = tnn_annual.values
            theil_slope, _, _, _ = theilslopes(y, x, 0.95)
            extreme_trends["trend_TNn"] = theil_slope
        else:
            # print(f"[Warning] Not enough years for TNn trend at {station}")
            extreme_trends["trend_TNn"] = np.nan

    # Merge into the main trend_slopes dict
    trend_slopes.update(extreme_trends)

    return trend_slopes


def analyze_precipitation_trend_theilsen(df, output_folder_path, station, station_name):
    if "Precipitation" not in df.columns:
        # print(f"'Precipitation' column not found in the file.")
        return None

    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")
    df["Year_Num"] = df["Year"].astype(int)
    df_clean = df.dropna(subset=["Precipitation"])

    if df_clean.empty or df_clean["Year_Num"].nunique() < 2:
        # print(f"[Warning] Not enough valid or varied data for station {station}.")
        return None

    x = df_clean["Year_Num"].values
    y = df_clean["Precipitation"].values

    theil_slope, intercept, _, _ = theilslopes(y, x, 0.95)
    theil_pred = intercept + theil_slope * x

    plt.figure(figsize=(12, 6))
    plt.scatter(df_clean["Date"], y, alpha=0.3, s=10, color="dodgerblue", label="Daily Precipitation")
    plt.plot(df_clean["Date"], theil_pred, color="navy", linewidth=2, label="Theil–Sen Trend")

    plt.title(f"Precipitation Trends at Station: {station_name} - using Theil–Sen Estimator")
    plt.xlabel("Year")
    plt.ylabel("Precipitation (mm)")
    plt.legend()
    plt.grid(True)

    last_date = df_clean["Date"].iloc[-1]
    last_predicted = theil_pred[-1]
    plt.text(last_date, last_predicted, f"{theil_slope:.2f} mm/yr", fontsize=10, color="navy", fontweight="bold")

    os.makedirs(output_folder_path, exist_ok=True)
    plot_path = os.path.join(output_folder_path, f"precipitation_trend_theilsen_{station}.jpg")
    plt.savefig(plot_path, dpi=300)
    # print(f"Precipitation trend plot saved at: {plot_path}")
    plt.close()

    return theil_slope




def analyze_monthly_precipitation_trend_theilsen(df, output_folder_path, station, station_name):
    if "Precipitation" not in df.columns:
        # print(f"'Precipitation' column not found in the file.")
        return None

    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")
    df_clean = df.dropna(subset=["Precipitation"])

    if df_clean.empty:
        # print(f"No valid precipitation data at station {station}.")
        return None

    df_clean = df_clean.copy()
    df_clean["YearMonth"] = df_clean["Date"].dt.to_period("M")
    monthly_avg = df_clean.groupby("YearMonth")["Precipitation"].mean().reset_index()
    monthly_avg["Date"] = monthly_avg["YearMonth"].dt.to_timestamp()
    monthly_avg["Year_Fraction"] = monthly_avg["Date"].dt.year + (monthly_avg["Date"].dt.month - 1) / 12.0

    if monthly_avg["Year_Fraction"].nunique() < 2:
        # print(f"[Warning] Not enough monthly data variation for station {station}.")
        return None

    x = monthly_avg["Year_Fraction"].values
    y = monthly_avg["Precipitation"].values

    theil_slope, intercept, _, _ = theilslopes(y, x, 0.95)
    theil_pred = intercept + theil_slope * x

    plt.figure(figsize=(12, 6))
    plt.scatter(monthly_avg["Date"], y, alpha=0.4, s=15, color="skyblue", label="Monthly Avg Precipitation")
    plt.plot(monthly_avg["Date"], theil_pred, color="navy", linewidth=2, label="Theil–Sen Trend")

    plt.title(f"Precipitation Trends at Station: {station_name} - using Theil–Sen Estimator")
    plt.xlabel("Year")
    plt.ylabel("Monthly Avg Precipitation (mm)")
    plt.legend()
    plt.grid(True)

    last_date = monthly_avg["Date"].iloc[-1]
    last_predicted = theil_pred[-1]
    plt.text(last_date, last_predicted, f"{theil_slope:.2f} mm/yr", fontsize=10, color="navy", fontweight="bold")

    os.makedirs(output_folder_path, exist_ok=True)
    plot_path = os.path.join(output_folder_path, f"monthly_precipitation_trend_theilsen_{station}.jpg")
    plt.savefig(plot_path, dpi=300)
    # print(f" Monthly precipitation trend plot saved at: {plot_path}")
    plt.close()

    return theil_slope


def plot_monthly_mean_temperatures(df, output_folder_path, station, station_name):

    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")

    # Resample monthly means
    df = df.set_index("Date")
    monthly_max = df["Max_Temperature"].resample("M").mean()
    monthly_min = df["Min_Temperature"].resample("M").mean()
    monthly_avg = df["Avg_Temperature"].resample("M").mean()

    plt.figure(figsize=(12, 6))
    plt.plot(monthly_max.index, monthly_max.values, color="red", linewidth=1.5, label="Monthly Max Temp")
    plt.plot(monthly_min.index, monthly_min.values, color="blue", linewidth=1.5, label="Monthly Min Temp")
    plt.plot(monthly_avg.index, monthly_avg.values, color="orange", linewidth=1.5, label="Monthly Mean Temp")

    plt.title(f"Monthly Mean Temperatures at {station_name}")
    plt.xlabel("Year")
    plt.ylabel("Temperature (°C)")
    plt.grid(True)
    plt.legend()
    
    os.makedirs(output_folder_path, exist_ok=True)
    plot_path = os.path.join(output_folder_path, f"monthly_mean_temperatures_{station}.jpg")
    plt.savefig(plot_path, dpi=300)
    # print(f"Monthly mean temperature plot saved at: {plot_path}")
    plt.close()


def plot_monthly_precipitation(df, output_folder_path, station, station_name):
    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")
    df = df.dropna(subset=["Precipitation"])

    # Resample to monthly sum
    monthly_precip = df.set_index("Date")["Precipitation"].resample("M").sum()

    plt.figure(figsize=(12, 6))
    plt.plot(monthly_precip.index, monthly_precip.values, color="dodgerblue", linewidth=1.5, label="Monthly Precipitation")

    plt.title(f"Monthly Precipitation at {station_name}")
    plt.xlabel("Year")
    plt.ylabel("Precipitation (mm)")
    plt.grid(True)
    plt.legend()
    
    os.makedirs(output_folder_path, exist_ok=True)
    plot_path = os.path.join(output_folder_path, f"monthly_precipitation_{station}.jpg")
    plt.savefig(plot_path, dpi=300)
    # print(f"Monthly precipitation plot saved at: {plot_path}")
    plt.close()



def plot_temperature_extremes_three_panel(df, output_folder_path, station, station_name):


    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")
    df["Year_Num"] = df["Date"].dt.year
    df = df[(df["Year_Num"] >= 1950) & (df["Year_Num"] <= 2022)]

    variables = {
        "Tx": "Max_Temperature",
        "Tn": "Min_Temperature",
        "Tavg": "Avg_Temperature"
    }

    periods = {
        "1950–1960": {"color": "gold", "range": (1950, 1960)},
        "1960–1970": {"color": "darkorange", "range": (1960, 1970)},
        "1970–1980": {"color": "red", "range": (1970, 1980)},
        "1980–1990": {"color": "firebrick", "range": (1980, 1990)},
        "1990–2000": {"color": "purple", "range": (1990, 2000)},
        "2000–2010": {"color": "mediumorchid", "range": (2000, 2010)},
        "2010–onward": {"color": "darkviolet", "range": (2010, 2022)}
    }

    bins = [0, 1, 2.5, 5, 95, 97.5, 99, 100]
    bin_labels = ['<1%', '1–2.5%', '2.5–5%', '95–97.5%', '97.5–99%', '>99%']
    bin_midpoints = [0.5, 1.75, 3.75, 96.25, 98.25, 99.5]
    x_shifted = list(range(6))

    fig, axes = plt.subplots(1, 3, figsize=(18, 5), sharey=True)

    for ax_idx, (ax, (var_name, temp_column)) in enumerate(zip(axes, variables.items())):
        df_valid = df.dropna(subset=[temp_column])
        if df_valid.empty:
            ax.set_visible(False)
            # print(f"[Warning] No valid {temp_column} data at station {station_name}. Skipping.")
            continue

        thresholds = np.percentile(df_valid[temp_column], bins)
        results = {}

        for label, period in periods.items():
            df_period = df_valid[
                (df_valid["Year_Num"] >= period["range"][0]) &
                (df_valid["Year_Num"] < period["range"][1])
            ]

            if df_period.empty:
                continue

            # Count how many *years* in the decade have at least one valid day
            valid_years_with_data = df_period.groupby("Year_Num")[temp_column].apply(lambda x: x.notna().sum() > 0).sum()
            if valid_years_with_data < 5: # ensures only decades with at least 5 years of valid data are included in the results
                # print(f"[Info] Skipping {label} for {var_name} due to insufficient valid years ({valid_years_with_data}/10).")
                continue

            total_years = period["range"][1] - period["range"][0]
            counts = []

            for i in [0, 1, 2, 4, 5, 6]:
                mask = (df_period[temp_column] > thresholds[i]) & (df_period[temp_column] <= thresholds[i + 1])
                avg_days = mask.sum() / total_years
                counts.append(avg_days)

            results[label] = counts


        # Plot lines
        # for label, values in results.items():
        #     color = periods[label]["color"]
        #     ax.plot(x_shifted, values, label=label, color=color, marker='o', linewidth=1)
        #     ax.fill_between(x_shifted, values, alpha=0.4, color=color)


        for label, values in results.items():
            color = periods[label]["color"]

            # Cold side (bins 0, 1, 2 → x = 0, 1, 2)
            ax.plot(x_shifted[:3], values[:3], label=label, color=color, marker='o', linewidth=1)
            ax.fill_between(x_shifted[:3], values[:3], alpha=0.4, color=color)

            # Hot side (bins 3, 4, 5 → x = 3, 4, 5)
            ax.plot(x_shifted[3:], values[3:], color=color, marker='o', linewidth=1)
            ax.fill_between(x_shifted[3:], values[3:], alpha=0.4, color=color)

        # Plot filled "mask" for the skipped central bin (3: 2.5–5% and 95–97.5%)
        # ax.axvspan(2, 3, color='lightgray', alpha=1.0, zorder=0)
        # ax.axvspan(2, 3, color='white', alpha=1.0, zorder=10) # white opaque version

        ax.set_title(var_name)
        ax.set_xticks(x_shifted)
        ax.set_xticklabels(bin_labels, rotation=45)
        ax.axvline(2.5, color='k', linestyle='--', linewidth=1)
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.grid(False) # Remove internal grid lines

        # Add cold/hot day text inside each panel
        if ax_idx == 0 or ax_idx == 2:
            ax.text(0.1, 0.5, "cold days", fontsize=9, transform=ax.transAxes)
            ax.text(0.8, 0.5, "hot days", fontsize=9, transform=ax.transAxes)
        
        if ax_idx == 1:
            ax.text(0.1, 0.5, "cold nights", fontsize=9, transform=ax.transAxes)
            ax.text(0.8, 0.5, "warm nights", fontsize=9, transform=ax.transAxes)
        

        # Add 99th percentile value top-right
        val_99th = thresholds[6]  # from bins
        ax.text(0.98, 0.95, f"99th: {val_99th:.1f} °C", fontsize=9, ha='right', transform=ax.transAxes)

        if ax_idx == 0:
            ax.set_ylabel("Number of days per year")
            ax.legend(loc='upper left', fontsize=8, frameon=False)

        if ax_idx == 1:
            ax.set_xlabel("Percentiles")

    fig.suptitle(f"Extreme Temperature Percentile Days – {station_name}", fontsize=14, y=1.05)
    fig.tight_layout()

    os.makedirs(output_folder_path, exist_ok=True)
    plot_path = os.path.join(output_folder_path, f"extreme_percentiles_{station}.jpg")
    plt.savefig(plot_path, dpi=300, bbox_inches='tight')
    plt.close()
    # print(f"3-panel percentile extremes plot saved at: {plot_path}")



def plot_full_period_trend_distribution_three_panel(df, output_folder_path, station, station_name, n_bootstrap=100):


    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")
    df["Year_Num"] = df["Date"].dt.year

    variables = {
        "Tx": "Max_Temperature",
        "Tn": "Min_Temperature",
        "Tavg": "Avg_Temperature"
    }

    fig, axes = plt.subplots(1, 3, figsize=(18, 5), sharey=True)

    for ax, (label, column) in zip(axes, variables.items()):
        df_var = df.dropna(subset=[column])
        if df_var["Year_Num"].nunique() < 2:
            ax.set_visible(False)
            # print(f"[Warning] Not enough data for {column} at {station_name}")
            continue

        x = df_var["Year_Num"].values
        y = df_var[column].values

        # Observed trend
        obs_slope, _, _, _ = theilslopes(y, x)

        # # Only use data from 2000 to 2010 for the observed trend
        # df_recent = df_var[(df_var["Year_Num"] >= 2000)]
        # x_recent = df_recent["Year_Num"].values
        # y_recent = df_recent[column].values

        # if len(x_recent) < 2 or len(np.unique(x_recent)) < 2:
        #     obs_slope = np.nan
        #     print(f"[WARNING] Not enough recent data (2000-to-date) for {column}.")
        # else:
        #     obs_slope, _, _, _ = theilslopes(y_recent, x_recent)

        # Bootstrapping
        bootstrap_slopes = []
        for _ in range(n_bootstrap):
            sample = df_var.sample(frac=1, replace=True)
            x_boot = sample["Year_Num"].values
            y_boot = sample[column].values
            if len(np.unique(x_boot)) < 2:
                continue
            slope, _, _, _ = theilslopes(y_boot, x_boot)
            bootstrap_slopes.append(slope)

        bootstrap_slopes = np.array(bootstrap_slopes)
        perc5 = np.percentile(bootstrap_slopes, 5)
        perc95 = np.percentile(bootstrap_slopes, 95)

        # Plot
        # bins = np.linspace(min(bootstrap_slopes.min(), obs_slope) - 0.2,
        #                    max(bootstrap_slopes.max(), obs_slope) + 0.2, 40)

        margin = (bootstrap_slopes.max() - bootstrap_slopes.min()) * 0.2  # small buffer
        bin_min = bootstrap_slopes.min() - margin
        bin_max = bootstrap_slopes.max() + margin
        bins = np.linspace(bin_min, bin_max, 60)

        ax.hist(bootstrap_slopes, bins=bins, color='lightgray', edgecolor='k', alpha=1.0)

        # ax.axvline(obs_slope, color='crimson', linestyle='--', linewidth=1.5,
        #            label=f"Recent trend = {obs_slope:.2f} °C/year")
        if not np.isnan(obs_slope):
            ax.axvline(obs_slope, color='crimson', linestyle='--', linewidth=1.5,
                    label=f"Obs. trend = {obs_slope:.2f} °C/year")

        ax.hlines(y=ax.get_ylim()[1] * 0.8, xmin=perc5, xmax=perc95, color='k', linewidth=2)
        ax.text((perc5 + perc95) / 2, ax.get_ylim()[1] * 0.85,
                f"5th–95th%: [{perc5:.2f}, {perc95:.2f}]", fontsize=9, ha='center')

        ax.set_title(label, fontsize=12)
        ax.set_xlabel("Trend (°C per year)")
        ax.grid(False)
        ax.legend(loc='upper right', fontsize=8, frameon=False)

    axes[0].set_ylabel("Frequency")
    fig.suptitle(f"Full-period Trend Distributions – {station_name}", fontsize=14, y=1.05)
    fig.tight_layout(rect=[0, 0, 1, 0.95])

    os.makedirs(output_folder_path, exist_ok=True)
    save_path = os.path.join(output_folder_path, f"full_period_trend_distribution_{station}.jpg")
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    # print(f"Full-period 3-panel trend distribution plot saved at: {save_path}")


def plot_temperature_distribution_shift_by_decade(df, output_folder_path, station, station_name):

    df["Date"] = pd.to_datetime(df[["Year", "Month", "Day"]], errors="coerce")
    df["Year_Num"] = df["Date"].dt.year

    variables = {
        "Tx": "Max_Temperature",
        "Tn": "Min_Temperature",
        "Tavg": "Avg_Temperature"
    }

    periods = {
        "1950–1960": {"color": "gold", "range": (1950, 1960)},
        "1960–1970": {"color": "darkorange", "range": (1960, 1970)},
        "1970–1980": {"color": "red", "range": (1970, 1980)},
        "1980–1990": {"color": "firebrick", "range": (1980, 1990)},
        "1990–2000": {"color": "purple", "range": (1990, 2000)},
        "2000–2010": {"color": "mediumorchid", "range": (2000, 2010)},
        "2010–onward": {"color": "darkviolet", "range": (2010, 2022)}
    }

    fig, axes = plt.subplots(1, 3, figsize=(18, 5), sharey=True)

    for ax, (label, column) in zip(axes, variables.items()):
        df_var = df.dropna(subset=[column])
        all_data = []

        sorted_periods = dict(sorted(periods.items(), key=lambda x: x[1]["range"][0]))
        for i, (decade, props) in enumerate(sorted_periods.items()):
        # for decade, props in periods.items():
        
            df_decade = df_var[
                (df_var["Year_Num"] >= props["range"][0]) & 
                (df_var["Year_Num"] < props["range"][1])
            ]
            if df_decade.empty:
                continue

            # values = df_decade[column].values
            values = df_decade[column].dropna().values
            if len(values) < 2:
                # print(f"[WARNING] Skipping KDE for {label} in {decade} due to insufficient data.")
                continue  # Skip this decade

            
            # Plotting KDE
            # Estimate and plot KDE (smooth curve)
            density = gaussian_kde(values)
            xs = np.linspace(values.min(), values.max(), 300)
            ys = density(xs)
            
            ax.plot(xs, ys, color=props["color"], label=decade, linewidth=2, alpha=0.9, zorder=i)
            ax.fill_between(xs, ys, color=props["color"], alpha=0.2, zorder=i)

            # # Plotting histograms
            # ax.hist(values, bins=30, density=True, alpha=0.4,
            #         color=props["color"], label=decade, zorder=i)

            # Add 95th percentile vertical line
            perc95 = np.percentile(values, 95)
            ax.axvline(perc95, linestyle='--', color=props["color"], linewidth=1)

        ax.set_title(label, fontsize=12)
        ax.set_xlabel("Temperature (°C)")
        if ax == axes[0]:
            ax.set_ylabel("Density")
        ax.grid(False)
        ax.legend(loc='upper left', fontsize=8, frameon=False)

    fig.suptitle(f"Shift in Temperature Distributions by Decade – {station_name}", fontsize=14, y=1.05)
    fig.tight_layout(rect=[0, 0, 1, 0.95])

    os.makedirs(output_folder_path, exist_ok=True)
    save_path = os.path.join(output_folder_path, f"temperature_distribution_shift_by_decade_{station}.jpg")
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    # print(f"Temperature distribution shift plot saved at: {save_path}")


def convert_to_sef_with_metadata(
    df, station_info, temp_column, temp_type, source="Institut National pour l'Etude et la Recherche Agronomiques",
    link="", stat="point", units="C", observer="", hour=0
):
    """
    Convert a DataFrame containing temperature data into the Station Exchange Format (SEF, .tsv) using station metadata.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame with 'Year', 'Month', 'Day', and the temperature column (`temp_column`).
    station_info : dict or Series
        Station metadata with keys: ID, name, latitude, longitude, altitude.
    temp_column : str
        Column name in `df` containing the temperature data.
    temp_type : str
        Type of temperature variable for SEF 'Vbl' field (e.g., "Tx", "Td").
    hour : int
        Hour of observation (e.g., 6 for 6 AM). Defaults to 0.
    source : str, optional
        Source of the data.
    link : str, optional
        URL link to data source.
    stat : str, optional
        Statistical type (e.g., 'point').
    units : str, optional
        Units of measurement, e.g., 'C' or 'mm'.
    observer : str, optional
        Observer or software metadata.

    Returns
    -------
    tuple
        - sef_headers: dict
        - sef_df: pandas.DataFrame (SEF-compliant data rows)
    """

    sef_headers = {
        "SEF": "1.0.0",
        "ID": station_info['ID'],
        "Name": station_info['name'],
        "Lat": station_info['latitude'],
        "Lon": station_info['longitude'],
        "Alt": station_info['altitude'],
        "Source": source,
        "Link": link,
        "Vbl": temp_type,
        "Stat": stat,
        "Units": units,
        "Meta": f"Observer={observer}  | QC software = MeteoSaver v1.0 | Note = Transcription done using MeteoSaver v1.0 (https://github.com/VUB-HYDR/MeteoSaver) developed by Derrick Muheki at the Department of Water and Climate, Vrije Universiteit Brussel"
    }

    sef_df = pd.DataFrame({
        "Year": df["Year"],
        "Month": df["Month"],
        "Day": df["Day"],
        "Hour": [hour] * len(df),
        "Minute": [0] * len(df),
        "Period": [0] * len(df),
        "Value": df[temp_column].fillna("NaN"),
        "Meta": [""] * len(df)
    })

    sef_column_order = ["Year", "Month", "Day", "Hour", "Minute", "Period", "Value", "Meta"]
    sef_df = sef_df[sef_column_order]

    return sef_headers, sef_df



<<<<<<< HEAD
def integrate_extra_data_if_available(merged_df, station, extra_data_directory):
    station_folder = os.path.join(extra_data_directory, f"{int(station):03d}")
    if not os.path.isdir(station_folder):
        # print(f"[INFO] No extra data folder for station {station}")
        return merged_df
=======
def data_formatting(input_folder_path, output_folder_path, metadata_file_path, station, date_column, header_rows, multi_day_totals, multi_day_averages, excluded_rows, additional_excluded_rows, final_totals_rows, uncertainty_margin):
>>>>>>> origin/main

    # Look for the INERA Excel file
    excel_files = [f for f in os.listdir(station_folder) if f.endswith('.xlsx')]
    if not excel_files:
        # print(f"[INFO] No Excel file found in folder: {station_folder}")
        return merged_df

    file_path = os.path.join(station_folder, excel_files[0])
    extra_df = pd.read_excel(file_path, engine='openpyxl')

    if extra_df.empty:
        return merged_df

    # Clean column names and extract dates
    extra_df = extra_df.rename(columns=lambda x: str(x).strip())
    for col in ['Year', 'Month', 'Day']:
        extra_df[col] = pd.to_datetime(extra_df[[col]].astype(str).agg('-'.join, axis=1), errors='coerce').dt.strftime('%Y-%m-%d')
    extra_df[['Year', 'Month', 'Day']] = extra_df['Year'].str.split('-', expand=True).astype(int)
    extra_df['Date'] = pd.to_datetime(extra_df[['Year', 'Month', 'Day']], errors='coerce')

    # Ensure merged_df has a Date column
    if 'Date' not in merged_df.columns:
        merged_df['Date'] = pd.to_datetime(merged_df[['Year', 'Month', 'Day']], errors='coerce')

    # Define relevant columns to update or add
    variables_to_update = ['Max_Temperature', 'Min_Temperature', 'Avg_Temperature', 'Precipitation']
    if 'Source' not in merged_df.columns:
        merged_df['Source'] = 'Original'

    for _, row in extra_df.iterrows():
        date = row['Date']
        if pd.isna(date):
            continue

        matched_indices = merged_df.index[merged_df['Date'] == date].tolist()

        if matched_indices:
            for idx in matched_indices:
                for var in variables_to_update:
                    if var in merged_df.columns:
                        current_val = merged_df.at[idx, var]
                        is_missing = pd.isna(current_val) or str(current_val).strip().lower() == 'nan'
                        if is_missing and not pd.isna(row.get(var)):
                            merged_df.at[idx, var] = row[var]
                            merged_df.at[idx, 'Source'] = 'Extra'
        else:
            # Add this row as a new record
            new_row = {
                'Year': row['Year'],
                'Month': row['Month'],
                'Day': row['Day'],
                'Date': row['Date'],
                'Source': 'Extra'
            }
            for var in variables_to_update:
                new_row[var] = row.get(var, np.nan)
            merged_df = pd.concat([merged_df, pd.DataFrame([new_row])], ignore_index=True)

    merged_df = merged_df.sort_values(by='Date').reset_index(drop=True)
    # print(f"[INFO] Integrated extra data for station {station}: {merged_df['Source'].value_counts().to_dict()}")
    return merged_df



def data_formatting(input_folder_path, output_folder_path, metadata_file_path, formatted_already_digitized_data_dir, station, date_column,
                    header_rows, multi_day_totals, multi_day_averages, excluded_rows,
                    additional_excluded_rows, final_totals_rows, uncertainty_margin, max_temperature_threshold, min_temperature_threshold):

    # Define your manually assigned column indices (based on Excel structure)
    max_temp_column_idx = 4   # 'D'
    min_temp_column_idx = 5   # 'E'
    avg_temp_column_idx = 6   # 'F'
    precip_column_idx   = 11  # 'K'
    dry06_column_idx    = 12  # 'L'
    wet06_column_idx    = 13  # 'M'
    dry15_column_idx    = 17  # 'Q'
    wet15_column_idx    = 18  # 'R'
    dry18_column_idx    = 22  # 'V'
    wet18_column_idx    = 23  # 'W'

    column_metadata = {
    "Max_Temperature": {"col": max_temp_column_idx, "vbl": "Tx", "hour": 0, "units": "C", "stat": "maximum"},
    "Min_Temperature": {"col": min_temp_column_idx, "vbl": "Tn", "hour": 0, "units": "C", "stat": "minimum"},
    "Avg_Temperature": {"col": avg_temp_column_idx, "vbl": "Ta", "hour": 0, "units": "C", "stat": "mean"},
    "Precipitation":   {"col": precip_column_idx,    "vbl": "rr", "hour": 6, "units": "mm", "stat": "sum"},
    "Dry_bulb_temp_06h00":  {"col": dry06_column_idx,     "vbl": "ta", "hour": 6, "units": "C", "stat": "point"},
    "Dry_bulb_temp_15h00":  {"col": dry15_column_idx,     "vbl": "ta", "hour": 15, "units": "C", "stat": "point"},
    "Dry_bulb_temp_18h00":  {"col": dry18_column_idx,     "vbl": "ta", "hour": 18, "units": "C", "stat": "point"},
    "Wet_bulb_temp_06h00":  {"col": wet06_column_idx,     "vbl": "tb", "hour": 6, "units": "C", "stat": "point"},
    "Wet_bulb_temp_15h00":  {"col": wet15_column_idx,     "vbl": "tb", "hour": 15, "units": "C", "stat": "point"},
    "Wet_bulb_temp_18h00":  {"col": wet18_column_idx,     "vbl": "tb", "hour": 18, "units": "C", "stat": "point"},
        }   ## Check https://datarescue.climate.copernicus.eu/variablenames for other variable names

<<<<<<< HEAD
=======
    # Define your manually assigned column indices (based on Excel structure)
    max_temp_column_idx = 4   # 'D'
    min_temp_column_idx = 5   # 'E'
    avg_temp_column_idx = 6   # 'F'
    precip_column_idx   = 11  # 'K'

    output_file = os.path.join(output_folder_path, 'Daily_all_temperatures_and_precipitation.xlsx')  # Combined output file with the three variables: Max, Min, and Average Temperature
    output_files = {  # Output files for individual temperature columns
        'Max_Temperature': os.path.join(output_folder_path, 'Daily_max_temperatures.xlsx'),
        'Min_Temperature': os.path.join(output_folder_path, 'Daily_min_temperatures.xlsx'),
        'Avg_Temperature': os.path.join(output_folder_path, 'Daily_mean_temperatures.xlsx'),
        'Precipitation': os.path.join(output_folder_path, 'Daily_precipiation.xlsx')
    }
>>>>>>> origin/main

    # Load station metadata
    station_metadata = load_station_metadata(metadata_file_path)
    station = str(station)
    station_metadata['ID'] = station_metadata['ID'].astype(str).str.strip().str.zfill(3)

    # Filter and extract station info
    station_info_df = station_metadata[station_metadata['ID'] == station]
    if station_info_df.empty:
        raise ValueError(f"No metadata found for station ID {station}")

    station_info = station_info_df.iloc[0]  # This is a Series, used in SEF headers
    station_name = station_info['name']  # Used for plot titles

<<<<<<< HEAD
    # Adjust excluded rows
=======
    # Lists to hold all data for each temperature type and precipitation
    data_max = []
    data_min = []
    data_avg = []
    data_precip = []

    # Rows to exclude. Adjust these according to your specific sheet
>>>>>>> origin/main
    if multi_day_totals and not multi_day_averages:
        pass
    elif multi_day_totals and multi_day_averages:
        excluded_rows += additional_excluded_rows
    elif not multi_day_totals:
        excluded_rows = final_totals_rows

<<<<<<< HEAD
    # Convert column letter (e.g. 'B') to column index (1-based)
    date_column_idx = ord(date_column.upper()) - ord('A') + 1

    # Dictionary to store data for each variable
    data_by_variable = {var: [] for var in column_metadata}

    # Loop through Excel files
    green_cell_counter = {var: 0 for var in column_metadata} # Count all the confirmed transcibed values by MeteoSaver
=======
    # Convert the day column letter and temperature columns to numeric indices
    date_column_idx = ord(date_column) - ord('A') + 1  # Convert 'B' -> 2  # Date
    # column_indices = [ord(col.strip()) - ord('A') + 1 for col in columns_to_check] # Max, min and average temperatures 

    # # Now `column_indices` will contain [4, 5, 6] for 'D', 'E', 'F'
    # max_temp_column_idx = column_indices[0]  # 'D' column index -> Maximum temperature
    # min_temp_column_idx = column_indices[1]  # 'E' column index -> Minimum temperature
    # avg_temp_column_idx = column_indices[2]  # 'F' column index -> Avergae temperature
>>>>>>> origin/main

    for filename in os.listdir(input_folder_path):
        if filename.endswith(".xlsx"):
            year, month = extract_date_from_filename(filename)
            if year and month:
                file_path = os.path.join(input_folder_path, filename)
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

<<<<<<< HEAD
                for row_num in range(header_rows + 1, ws.max_row + 1):
                    if row_num in excluded_rows:
                        continue

                    day_cell = ws.cell(row=row_num, column=date_column_idx)
                    if not day_cell.value:
                        continue
                    day = int(day_cell.value)
                    
                    for var_name, meta in column_metadata.items():
                        col = meta["col"]
                        cell = ws.cell(row=row_num, column=col)
                        if is_highlighted_green(cell, 'FF6DCD57'):
                            green_cell_counter[var_name] += 1
                            val = cell.value
                        else:
                            val = 'NaN'
                        data_by_variable[var_name].append([year, month, day, val])

    # Create merged DataFrame from collected data
    merged_df = pd.DataFrame()
    for key in data_by_variable:
        df_key = pd.DataFrame(data_by_variable[key], columns=["Year", "Month", "Day", key])
        if merged_df.empty:
            merged_df = df_key
        else:
            merged_df = pd.merge(merged_df, df_key, on=["Year", "Month", "Day"], how="outer")
=======
                # Extract data from rows and columns, excluding specific rows.  
                for row_num in range(header_rows+1, worksheet.max_row + 1): #Here this represents Max, Min and Average Temperatures
                    if row_num not in excluded_rows: 
                        day_cell = worksheet.cell(row=row_num, column=date_column_idx)  # Assuming the day is in the first column
                        max_temperature_cell = worksheet.cell(row=row_num, column=max_temp_column_idx)  # Column for Max Temperature
                        min_temperature_cell = worksheet.cell(row=row_num, column=min_temp_column_idx)  # Column for Min Temperature
                        average_temperature_cell = worksheet.cell(row=row_num, column=avg_temp_column_idx)  # Column for Avg Temperature
                        precipitation_cell = worksheet.cell(row=row_num, column=precip_column_idx)  # Column for precipitation

                        if day_cell.value :
                            day = int(day_cell.value)
                            max_temperature = max_temperature_cell.value if is_highlighted_green(max_temperature_cell, 'FF6DCD57') else 'NaN'
                            min_temperature = min_temperature_cell.value if is_highlighted_green(min_temperature_cell, 'FF6DCD57') else 'NaN'
                            average_temperature = average_temperature_cell.value if is_highlighted_green(average_temperature_cell, 'FF6DCD57') else 'NaN'
                            precipitation = precipitation_cell.value if is_highlighted_green(precipitation_cell, 'FF6DCD57') else 'NaN'

                            data.append([year, month, day, max_temperature, min_temperature, average_temperature, precipitation])

                            data_max.append([year, month, day, max_temperature])
                            data_min.append([year, month, day, min_temperature])
                            data_avg.append([year, month, day, average_temperature])
                            data_precip.append([year, month, day, precipitation])

    # Create a DataFrame from the data
    df = pd.DataFrame(data, columns=["Year", "Month", "Day", "Max_Temperature", "Min_Temperature", "Avg_Temperature", "Precipitation"])


    # Generate a complete date range for each year and month combination
    years_months = df[['Year', 'Month']].drop_duplicates()
    complete_data = []
>>>>>>> origin/main

    # Create full date range to ensure all days are present
    years_months = merged_df[["Year", "Month"]].drop_duplicates()
    full_dates = []
    for _, row in years_months.iterrows():
        y, m = row["Year"], row["Month"]
        for d in range(1, pd.Period(f'{y}-{m}').days_in_month + 1):
            full_dates.append([y, m, d])
    complete_df = pd.DataFrame(full_dates, columns=["Year", "Month", "Day"])
    merged_df = pd.merge(complete_df, merged_df, on=["Year", "Month", "Day"], how="left")
    merged_df = merged_df.sort_values(["Year", "Month", "Day"])

    # === Optional: Integrate Extra INERA Data ===
    merged_df = integrate_extra_data_if_available(merged_df, station, formatted_already_digitized_data_dir)

<<<<<<< HEAD
    # Convert values to numeric
    for col in merged_df.columns:
        if col not in ["Year", "Month", "Day"]:
            merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce')

    # Detect outliers for temperature variables only (not precipitation or wet bulb)
    temp_vars = ["Max_Temperature", "Min_Temperature", "Avg_Temperature"]
    for column in temp_vars:
        if column not in merged_df.columns:
            continue
        
        # First, remove extreme physically implausible values using the set maximum and minimum thresholds for the region
        merged_df.loc[(merged_df[column] > max_temperature_threshold) | (merged_df[column] < min_temperature_threshold), column] = np.nan
=======
    # Merge the complete date range with the extracted data
    merged_df = pd.merge(complete_df, df, on=["Year", "Month", "Day"])
   
    # Fill missing temperatures with a placeholder value (e.g., NaN or a specific value)
    for column in ["Max_Temperature", "Min_Temperature", "Avg_Temperature", "Precipitation"]: # Since this is temperature, missing vales cannot be zero (0)
        merged_df[column] = merged_df[column].fillna(np.nan)
    
    # Convert temperature columns to numeric, coerce errors to NaN.
    merged_df['Max_Temperature'] = pd.to_numeric(merged_df['Max_Temperature'], errors='coerce')
    merged_df['Min_Temperature'] = pd.to_numeric(merged_df['Min_Temperature'], errors='coerce')
    merged_df['Avg_Temperature'] = pd.to_numeric(merged_df['Avg_Temperature'], errors='coerce')
    merged_df['Precipitation'] = pd.to_numeric(merged_df['Precipitation'], errors='coerce')


    # Sort DataFrame by Year, Month, Day
    merged_df = merged_df.sort_values(by=['Year', 'Month', 'Day'])
>>>>>>> origin/main

        std = merged_df[column].std()
        mean = merged_df[column].mean()
        merged_df[f"{column}_Flag"] = ""

        # Condition 1: value > 3 standard deviations
        for i in range(len(merged_df)):
            if abs(merged_df.loc[i, column] - mean) > 3 * std:
                merged_df.loc[i, column] = np.nan
                merged_df.loc[i, f"{column}_Flag"] = "Condition 1" # Flag as Condition 1 (current value as an outlier)

        # Condition 2: sharp spike flip-flop
        merged_df['std_diff'] = (merged_df[column] - mean) / std
        for i in range(1, len(merged_df) - 1):
            prev_sd = merged_df.loc[i - 1, 'std_diff'] if not pd.isna(merged_df.loc[i - 1, 'std_diff']) else 0
            curr_sd = merged_df.loc[i, 'std_diff']
            next_sd = merged_df.loc[i + 1, 'std_diff'] if not pd.isna(merged_df.loc[i + 1, 'std_diff']) else 0
            if not pd.isna(curr_sd) and (
                (prev_sd < -4 and curr_sd > 4 and next_sd < -4) or
                (prev_sd > 4 and curr_sd < -4 and next_sd > 4)
            ):
                merged_df.loc[i, column] = np.nan
                merged_df.loc[i, f"{column}_Flag"] = "Condition 2" # Flag as Condition 2 (current value as an outlier)
        merged_df.drop(columns=['std_diff'], inplace=True)

<<<<<<< HEAD
=======
    # Detect outliers for precipitation
    if 'Precipitation' in merged_df.columns:
        merged_df['Precipitation_Flag'] = ""

        # Filter out implausibly high precipitation values (e.g., above 150 mm/day)
        precip_threshold = 150  # You can adjust this based on regional context
        merged_df.loc[merged_df['Precipitation'] > precip_threshold, 'Precipitation_Flag'] = "Condition 1"
        merged_df.loc[merged_df['Precipitation'] > precip_threshold, 'Precipitation'] = np.nan

    # Save flagged data to Excel and apply conditional formatting
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        merged_df.to_excel(writer, index=False, sheet_name="Data")
        workbook = writer.book
        worksheet = writer.sheets["Data"]
>>>>>>> origin/main

    # Detect outliers for precipitation
    if 'Precipitation' in merged_df.columns:
        merged_df['Precipitation_Flag'] = ""

<<<<<<< HEAD
        # Filter out implausibly high precipitation values (e.g., above 150 mm/day)
        precip_threshold = 150  # You can adjust this based on regional context
        merged_df.loc[merged_df['Precipitation'] > precip_threshold, 'Precipitation_Flag'] = "Condition 1"
        merged_df.loc[merged_df['Precipitation'] > precip_threshold, 'Precipitation'] = np.nan

=======
        # Apply conditional formatting to only flagged cells
        for column in ["Max_Temperature", "Min_Temperature", "Avg_Temperature", "Precipitation"]:
            flag_column = f"{column}_Flag"
            for row in range(2, len(merged_df) + 2):  # Adjusting for header in Excel
                if merged_df.loc[row - 2, flag_column] in ["Condition 1", "Condition 2"]:
                    cell = worksheet[f"{openpyxl.utils.get_column_letter(merged_df.columns.get_loc(column) + 1)}{row}"]
                    cell.fill = dark_red_fill

    # Clean up flag columns in the DataFrame for further processing, if needed
    merged_df.drop(columns=[f"{col}_Flag" for col in ["Max_Temperature", "Min_Temperature", "Avg_Temperature", "Precipitation"]], inplace=True)

    # # Save the DataFrame to a new Excel file 
    # #merged_df.to_excel(output_file, index=False)
    # timeseries = merged_df.fillna('NaN')
    # timeseries.to_excel(output_file, index=False)

    # After processing, generate the SEF file
    # Loop over each temperature type and create a SEF file for each
    temperature_columns = {
        "Max_Temperature": "Tx",
        "Min_Temperature": "Tn",
        "Avg_Temperature": "Ta"
    }
    
    
    for temp_column, temp_type in temperature_columns.items():
        # Filter data for the specific temperature type
        timeseries_df = merged_df[['Year', 'Month', 'Day', temp_column]].fillna('NaN')
        timeseries_df = timeseries_df.rename(columns={temp_column: "Value"})

        # Convert to SEF format with headers using the function
        sef_headers, sef_df = convert_to_sef_with_metadata(
            df=timeseries_df,
            station_info=station_info,
            temp_column="Value",    # Pass the renamed column "Value"
            temp_type=temp_type      # Pass the type (e.g., Tx, Tn, Ta) for SEF header
        )

        # Define the output file path
        sef_output_file = os.path.join(output_folder_path, f"SEF_station_{station}_{temp_type}_temperature.tsv")
>>>>>>> origin/main
        
    # Update data_by_variable with cleaned values
    for var in data_by_variable:
        if var in merged_df.columns:
            records = merged_df[["Year", "Month", "Day", var]].values.tolist()
            cleaned = [[int(y), int(m), int(d), v if not pd.isna(v) else "NaN"]
                    for y, m, d, v in records]
            data_by_variable[var] = cleaned


    # Export to SEF per variable
    for var_name, records in data_by_variable.items():
        if not records:
            continue

        df = pd.DataFrame(records, columns=["Year", "Month", "Day", "Value"])
        meta = column_metadata[var_name]

        sef_headers, sef_df = convert_to_sef_with_metadata(
        df=df,
        station_info=station_info,
        temp_column="Value",
        temp_type=meta["vbl"],
        hour=meta["hour"],
        units=meta["units"],
        stat=meta["stat"])

        sef_file = os.path.join(output_folder_path, f"SEF_station_{station}_{var_name}.tsv")
        with open(sef_file, 'w') as f:
            for hkey, hval in sef_headers.items():
                f.write(f"{hkey}\t{hval}\n")
            sef_df.to_csv(f, sep='\t', index=False) # Write the main SEF data with tab separation and include the header row for data columns


    # Save all cleaned variables as Excel files
    for var_name in data_by_variable:
        df_var = pd.DataFrame(data_by_variable[var_name], columns=["Year", "Month", "Day", var_name])
        df_var.to_excel(os.path.join(output_folder_path, f"{var_name}_timeseries.xlsx"), index=False)

    # Add Date column
    merged_df['Date'] = pd.to_datetime(merged_df[['Year', 'Month', 'Day']])

    #***PLOTTING***
    # === Plot 1: Combined Temp + Precip ===
    plot_df = merged_df.set_index('Date')[['Max_Temperature', 'Min_Temperature', 'Avg_Temperature', 'Precipitation']]
    fig, ax = plt.subplots(figsize=(10, 6))
    plot_df['Max_Temperature'].plot(ax=ax, label='Max Temp', color='red')
    ax.fill_between(plot_df.index, plot_df['Max_Temperature'] - uncertainty_margin, plot_df['Max_Temperature'] + uncertainty_margin, color='red', alpha=0.2)
    plot_df['Min_Temperature'].plot(ax=ax, label='Min Temp', color='blue')
    ax.fill_between(plot_df.index, plot_df['Min_Temperature'] - uncertainty_margin, plot_df['Min_Temperature'] + uncertainty_margin, color='blue', alpha=0.2)
    plot_df['Avg_Temperature'].plot(ax=ax, label='Avg Temp', color='orange')
    ax.fill_between(plot_df.index, plot_df['Avg_Temperature'] - uncertainty_margin, plot_df['Avg_Temperature'] + uncertainty_margin, color='orange', alpha=0.2)
    ax_precip = ax.twinx()
    ax_precip.scatter(plot_df.index, plot_df['Precipitation'], marker='x', color='cornflowerblue', label='Precipitation', alpha=0.7)
    ax.set_title(f"Temperature & Precipitation at Station: {station_name}")
    ax.set_ylabel("Temperature (°C)")
    ax_precip.set_ylabel("Precipitation (mm)")
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax_precip.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, loc='upper right')
    ax.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(output_folder_path, f"combined_temperature_and_precipitation_plot.jpg"))
    # plt.show()
    plt.close()

    # === Plot 2: Only Temperature ===
    fig, ax = plt.subplots(figsize=(10, 6))
    for col, color in zip(['Max_Temperature', 'Min_Temperature', 'Avg_Temperature'], ['red', 'blue', 'orange']):
        ax.plot(merged_df['Date'], merged_df[col], label=col.replace("_", " "), color=color)
        ax.fill_between(merged_df['Date'], merged_df[col] - uncertainty_margin, merged_df[col] + uncertainty_margin, alpha=0.2, color=color)
    ax.set_title(f"Temperature Time Series at Station: {station_name}")
    ax.set_xlabel("Date")
    ax.set_ylabel("Temperature (°C)")
    ax.legend()
    ax.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(output_folder_path, f"temperature_plot_station_{station}.jpg"))
    # plt.show()
    plt.close()

    # === Plot 3.1: Precipitation ===
    precip = merged_df[merged_df['Precipitation'] >= 0]
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.scatter(precip['Date'], precip['Precipitation'], color='cornflowerblue', marker='x', alpha=0.7)
    ax.set_title(f"Precipitation at Station: {station_name}")
    ax.set_xlabel("Date")
    ax.set_ylabel("Precipitation (mm)")
    ax.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(output_folder_path, f"precipitation_plot_station_{station}.jpg"))
    # plt.show()
    plt.close()


    # === Plot 3.2: Precipitation Trends (Raw Daily Values with Mean & 95% Confidence Band) ===
    if 'Precipitation' in merged_df.columns:
        merged_df['DayOfYear'] = merged_df['Date'].dt.dayofyear
        merged_df['Year_Num'] = merged_df['Year'].astype(int)
        cmap = plt.cm.get_cmap("Blues", merged_df['Year_Num'].nunique())
        year_to_color = dict(zip(sorted(merged_df['Year_Num'].unique()), [cmap(i) for i in range(cmap.N)]))

        fig, ax = plt.subplots(figsize=(12, 6))

        # Plot daily precipitation per year
        for year in sorted(merged_df['Year_Num'].unique()):
            df_year = merged_df[merged_df['Year_Num'] == year]
            ax.scatter(df_year['DayOfYear'], df_year['Precipitation'], label=str(year),
                       color=year_to_color[year], s=5, alpha=0.4)

        # Compute mean and 95% CI per day of year
        mean_by_day = merged_df.groupby('DayOfYear')['Precipitation'].mean()
        lower = merged_df.groupby('DayOfYear')['Precipitation'].quantile(0.025)
        upper = merged_df.groupby('DayOfYear')['Precipitation'].quantile(0.975)

        # Plot mean line and 95% confidence interval band
        ax.plot(mean_by_day.index, mean_by_day.values, color='black', linewidth=2, label='Daily Mean')
        ax.fill_between(mean_by_day.index, lower.values, upper.values, color='gray', alpha=0.3, label='95% CI')

        # Add formatting
        ax.set_xticks([1, 32, 60, 91, 121, 152, 182, 213, 244, 274, 305, 335])
        ax.set_xticklabels(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])
        ax.set_title(f"Daily Precipitation Trends at Station: {station_name}")
        ax.set_xlabel("Month")
        ax.set_ylabel("Precipitation (mm)")
        ax.grid(True)
        # ax.legend(loc='upper left', fontsize=8, ncol=2)
        plt.tight_layout()
        plt.savefig(os.path.join(output_folder_path, f"daily_precipitation_trends_station_{station}.jpg"))
        plt.close()
    

    # === Plot 3.3: Precipitation Anomalies (vs. Long-Term Mean) ===
    if 'Precipitation' in merged_df.columns:
        df_anomaly = merged_df.copy()
        df_anomaly['Date'] = pd.to_datetime(df_anomaly[['Year', 'Month', 'Day']], errors='coerce')
        df_anomaly['Precipitation'] = pd.to_numeric(df_anomaly['Precipitation'], errors='coerce')
        df_anomaly = df_anomaly.dropna(subset=['Date', 'Precipitation'])

        df_anomaly['DayOfYear'] = df_anomaly['Date'].dt.dayofyear
        long_term_mean = df_anomaly.groupby('DayOfYear')['Precipitation'].mean()
        df_anomaly['Anomaly'] = df_anomaly['Precipitation'] - df_anomaly['DayOfYear'].map(long_term_mean)

        valid_years = df_anomaly.groupby('Year')['Precipitation'].count()
        valid_years = valid_years[valid_years > 30].index
        years = sorted(valid_years)

        if len(years) == 0:
            print(f"[WARNING] No valid years with sufficient precipitation data at station {station}. Skipping anomaly plot.")
        else:
            # Define decades and colormap
            decades = {
                "1950–1960": (1950, 1960),
                "1960–1970": (1960, 1970),
                "1970–1980": (1970, 1980),
                "1980–1990": (1980, 1990),
                "1990–2000": (1990, 2000),
                "2000–2010": (2000, 2010),
                "2010–2022": (2010, 2022)
            }

            cmap = plt.cm.get_cmap('Blues', len(decades))  # Use 'Blues' and split into len(decades) shades
            plotted_decades = {}

            fig, ax = plt.subplots(figsize=(14, 6))

            for idx, (label, (start, end)) in enumerate(decades.items()):
                decade_years = [y for y in years if start <= y < end]
                if not decade_years:
                    continue

                color = cmap(idx)
                plotted_decades[label] = color

                for year in decade_years:
                    yearly_data = df_anomaly[df_anomaly['Year'] == year]
                    ax.plot(yearly_data['DayOfYear'], yearly_data['Anomaly'], color=color, alpha=0.9)

            ax.axhline(0, color='black', linewidth=1)
            ax.set_xticks([1, 32, 60, 91, 121, 152, 182, 213, 244, 274, 305, 335])
            ax.set_xticklabels(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])
            ax.set_title(f"Daily Precipitation Anomalies Compared to Long-term Average – {station_name}")
            ax.set_xlabel("Month")
            ax.set_ylabel("Precipitation Anomaly (mm)")

            # Add decade legend
            legend_handles = [plt.Line2D([0], [0], color=c, label=d, linewidth=3)
                            for d, c in plotted_decades.items()]
            ax.legend(handles=legend_handles, title="Decades", loc="upper left", fontsize=8)

            ax.grid(True)
            plt.tight_layout()
            plt.savefig(os.path.join(output_folder_path, f"precipitation_anomalies_station_{station}.jpg"))
            plt.close()


    # === Plot 4: Yearly Temperature Trends with 14-day Rolling Average (One color per decade) ===
    decade_colormaps = {
        "Max_Temperature": "hot_r",   # reversed autumn
        "Avg_Temperature": "autumn_r",   # reversed summer
        "Min_Temperature": "cool"      # reversed cool
    }

    decades = {
        "1950–1960": (1950, 1960),
        "1960–1970": (1960, 1970),
        "1970–1980": (1970, 1980),
        "1980–1990": (1980, 1990),
        "1990–2000": (1990, 2000),
        "2000–2010": (2000, 2010),
        "2010–2022": (2010, 2022)
    }

    for column in ["Max_Temperature", "Min_Temperature", "Avg_Temperature"]:
        cmap_name = decade_colormaps[column]
        merged_df[column] = pd.to_numeric(merged_df[column], errors='coerce')
        valid_years = sorted(merged_df.dropna(subset=[column])['Year'].unique())

        if len(valid_years) == 0:
            print(f"[WARNING] No valid years with {column} data at station {station}. Skipping plot.")
            continue

        fig, ax = plt.subplots(figsize=(12, 6))

        # Use decade-specific colormap and normalize across N decades
        cmap = plt.cm.get_cmap(cmap_name, len(decades))
        plotted_decades = {}

        for idx, (label, (start, end)) in enumerate(decades.items()):
            df_decade = merged_df[(merged_df["Year"] >= start) & (merged_df["Year"] < end)]
            decade_years = sorted(df_decade.dropna(subset=[column])['Year'].unique())

            if not decade_years:
                continue

            decade_color = cmap(idx)  # Same color for all years in this decade
            plotted_decades[label] = decade_color

            for year in decade_years:
                df_year = df_decade[df_decade['Year'] == year].copy()
                df_year[f'{column}_14day'] = df_year[column].rolling(window=14, min_periods=1).mean()
                ax.plot(df_year['Date'].dt.dayofyear, df_year[f'{column}_14day'], color=decade_color, alpha=0.9, linewidth=1)

        # Final formatting
        y_min, y_max = ax.get_ylim()
        padding = (y_max - y_min) * 0.2
        ax.set_ylim(y_min - padding, y_max + padding)

        ax.set_xticks([1, 32, 60, 91, 121, 152, 182, 213, 244, 274, 305, 335])
        ax.set_xticklabels(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])

        ax.set_title(f"{column.replace('_', ' ')} (14-day rolling) Trends at Station: {station_name}")
        ax.set_xlabel("Month")
        ax.set_ylabel(f"{column.replace('_', ' ')} (°C)")

        # Add legend (one entry per decade)
        legend_handles = [plt.Line2D([0], [0], color=color, label=decade, linewidth=3)
                        for decade, color in plotted_decades.items()]
        ax.legend(handles=legend_handles, title="Decades", loc="upper left", fontsize=8)

        ax.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(output_folder_path, f"{column}_trends_14day_station_{station}.jpg"))
        plt.close()



    # === Plot 5: Dry and Wet Bulb Plots ===
    plot_sets = {
        "Dry bulb temperatures": ["Dry_bulb_temp_06h00", "Dry_bulb_temp_15h00", "Dry_bulb_temp_18h00"],
        "Wet bulb temperatures": ["Wet_bulb_temp_06h00", "Wet_bulb_temp_15h00", "Wet_bulb_temp_18h00"],
    }

    for title, columns in plot_sets.items():
        fig, ax = plt.subplots(figsize=(10, 6))
        for col in columns:
            if col in merged_df.columns:
                merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce')
                ax.plot(merged_df['Date'], merged_df[col], label=col.replace('_', ' '))
        ax.set_title(f"{title} at Station: {station_name}")
        ax.set_xlabel("Date")
        ax.set_ylabel("Temperature (°C)")
        ax.legend()
        ax.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(output_folder_path, f"{title.replace(' ', '_').lower()}_plot_station_{station}.jpg"))
        # plt.show()
        plt.close()
    

    

    # Save the cleaned full dataset (needed for trend analysis)
    trend_excel_path = os.path.join(output_folder_path, f"Daily_temperature_and_precipitation_station_{station}.xlsx")
    merged_df.to_excel(trend_excel_path, index=False, sheet_name="Data")

    # Only apply yellow highlight if 'Source' column exists
    if 'Source' in merged_df.columns:
        # Reload the file with openpyxl to apply formatting
        wb = load_workbook(trend_excel_path)
        ws = wb["Data"]

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        header_row = [cell.value for cell in ws[1]]
        if "Source" in header_row:
            source_col_idx = header_row.index("Source") + 1  # 1-based

            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=source_col_idx).value == "Extra":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill

            wb.save(trend_excel_path)
            wb.close()
    


    # === Plot 6: Timeseries and Trend lines with slope annotation ===
    df_for_trends = merged_df.copy()

    # 6.1 Mean monthly timeseries
    timeseries_monthly_mean_temperatures =plot_monthly_mean_temperatures(df_for_trends, output_folder_path, station, station_name)
    timeseries_monthly_mean_precipitation = plot_monthly_precipitation(df_for_trends, output_folder_path, station, station_name)

    # Analyze and plot temperature trends
    # 6.2: Linear Regression
    temperature_trend_slopes_with_linear_regression = analyze_temperature_trends_with_linear_regression(df_for_trends, output_folder_path, station, station_name)
    # 6.3: Theil-Sen Estimator
    temperature_trend_slopes_with_theilsen = analyze_temperature_trends_with_theilsen(df_for_trends, output_folder_path, station, station_name)
    # 6.4 Precipitation_trends
    precipitation_trend_slopes_with_theilsen = analyze_monthly_precipitation_trend_theilsen(df_for_trends, output_folder_path, station, station_name)


    # Scaling the distributions
    # 6.5 Frequency of hot and cold extremes over time
    frequency_of_extremes_over_time = plot_temperature_extremes_three_panel(df_for_trends, output_folder_path, station, station_name)
    # 6.6 Shift in trend distribution
    # shift_in_trend_distribution = plot_full_period_trend_distribution_three_panel(df_for_trends, output_folder_path, station, station_name)
    
    shift_in_temperature_distribution = plot_temperature_distribution_shift_by_decade(df_for_trends, output_folder_path, station, station_name)

    # To return dictionary with station information and theilsen slopes per temperature type for the particular station per year
    return {
    "station_id": station,
    "station_name": station_info["name"],
    "lat": station_info["latitude"],
    "lon": station_info["longitude"],
    "trend_max_temperature": temperature_trend_slopes_with_theilsen.get("trend_max_temperature", {}).get("theil_slope"),
    "trend_min_temperature": temperature_trend_slopes_with_theilsen.get("trend_min_temperature", {}).get("theil_slope"),
    "trend_avg_temperature": temperature_trend_slopes_with_theilsen.get("trend_avg_temperature", {}).get("theil_slope"),
    "trend_TXx": temperature_trend_slopes_with_theilsen.get("trend_TXx"),
    "trend_TNn": temperature_trend_slopes_with_theilsen.get("trend_TNn"),
    "data": df_for_trends,
    "green_cell_count": green_cell_counter}



def plot_trend_interpolation_map(trend_df, region_shapefile_path, output_folder_path):
    """
    Plot interpolated warming trends (Max, Min, Avg Temperature) using Cartopy with DRC shapefile.

    Parameters
    ----------
    trend_df : pd.DataFrame
        DataFrame with columns: ['station_id', 'station_name', 'lat', 'lon',
                                 'trend_max_temperature', 'trend_min_temperature', 'trend_avg_temperature']
    region_shapefile_path : str
        Path to the shapefile of the DRC border.
    output_folder_path : str
        Directory to save the output figure.
    """

    os.makedirs(output_folder_path, exist_ok=True)

    variables = {
        'trend_max_temperature': 'Max Temperature Trend (°C/yr)',
        'trend_min_temperature': 'Min Temperature Trend (°C/yr)',
        'trend_avg_temperature': 'Avg Temperature Trend (°C/yr)'
    }

    fig, axes = plt.subplots(1, 3, figsize=(18, 6), subplot_kw={'projection': ccrs.PlateCarree()})

    for ax, (var_key, title) in zip(axes, variables.items()):
        ax.set_title(title)
        ax.set_extent([12, 32, -14, 6], crs=ccrs.PlateCarree())
        # ax.add_feature(cfeature.BORDERS, linestyle=':')
        ax.add_feature(cfeature.COASTLINE)
        ax.add_feature(cfeature.LAND, edgecolor='black', facecolor='lightgray')
        ax.add_feature(cfeature.LAKES, edgecolor='blue', facecolor='blue')
        ax.add_feature(cfeature.RIVERS)

        # Add DRC shapefile
        region_shape = ShapelyFeature(
            Reader(region_shapefile_path).geometries(),
            ccrs.PlateCarree(), edgecolor='black', facecolor='none', linewidth=1.2
        )
        ax.add_feature(region_shape)

        # Ensure required columns are present
        required_columns = ['lat', 'lon', var_key]
        for col in required_columns:
            if col not in trend_df.columns:
                # print(f"[WARNING] Missing column '{col}' in trend_df. Skipping variable: {var_key}")
                continue

        # Drop rows with missing trend or missing lat/lon
        subset = trend_df.dropna(subset=required_columns)
        if subset.empty:
            # print(f"[WARNING] No valid data for {var_key} after filtering missing lat/lon. Skipping.")
            continue

        lons = subset["lon"].values
        lats = subset["lat"].values
        values = subset[var_key].values

        # Grid interpolation
        xi = np.linspace(lons.min(), lons.max(), 100)
        yi = np.linspace(lats.min(), lats.max(), 100)
        xi, yi = np.meshgrid(xi, yi)
        zi = griddata((lons, lats), values, (xi, yi), method='cubic')

        # Contour plot
        contour = ax.contourf(xi, yi, zi, cmap='plasma', transform=ccrs.PlateCarree(), levels=20)
        plt.colorbar(contour, ax=ax, orientation='vertical', shrink=0.8)

        # Add station locations
        ax.scatter(lons, lats, facecolor='black', edgecolors='white', s=40, linewidth=1.2, label='Stations', zorder=3)

    plt.tight_layout()
    save_path = os.path.join(output_folder_path, "spatial_trend_interpolation.jpg")
    plt.savefig(save_path, dpi=300)
    plt.close()


    ## Second plot with TXx and TNn
    second_variables = {
        'trend_TXx': 'TXx Trend (°C/yr)',
        'trend_TNn': 'TNn Trend (°C/yr)',
    }

    fig, axes = plt.subplots(1, 2, figsize=(18, 6), subplot_kw={'projection': ccrs.PlateCarree()})

    for ax, (var_key, title) in zip(axes, second_variables.items()):
        ax.set_title(title)
        ax.set_extent([12, 32, -14, 6], crs=ccrs.PlateCarree())
        # ax.add_feature(cfeature.BORDERS, linestyle=':')
        ax.add_feature(cfeature.COASTLINE)
        ax.add_feature(cfeature.LAND, edgecolor='black', facecolor='lightgray')
        ax.add_feature(cfeature.LAKES, edgecolor='blue', facecolor='blue')
        ax.add_feature(cfeature.RIVERS)

        # Add DRC shapefile
        region_shape = ShapelyFeature(
            Reader(region_shapefile_path).geometries(),
            ccrs.PlateCarree(), edgecolor='black', facecolor='none', linewidth=1.2
        )
        ax.add_feature(region_shape)

        # Ensure required columns are present
        required_columns = ['lat', 'lon', var_key]
        for col in required_columns:
            if col not in trend_df.columns:
                # print(f"[WARNING] Missing column '{col}' in trend_df. Skipping variable: {var_key}")
                continue

        # Drop rows with missing trend or missing lat/lon
        subset = trend_df.dropna(subset=required_columns)
        if subset.empty:
            # print(f"[WARNING] No valid data for {var_key} after filtering missing lat/lon. Skipping.")
            continue

        lons = subset["lon"].values
        lats = subset["lat"].values
        values = subset[var_key].values

        # Grid interpolation
        xi = np.linspace(lons.min(), lons.max(), 100)
        yi = np.linspace(lats.min(), lats.max(), 100)
        xi, yi = np.meshgrid(xi, yi)
        zi = griddata((lons, lats), values, (xi, yi), method='cubic')

        # Contour plot
        contour = ax.contourf(xi, yi, zi, cmap='plasma', transform=ccrs.PlateCarree(), levels=20)
        plt.colorbar(contour, ax=ax, orientation='vertical', shrink=0.8)

        # Add station locations
        ax.scatter(lons, lats, facecolor='black', edgecolors='white', s=40, linewidth=1.2, label='Stations', zorder=3)

    plt.tight_layout()
    save_path = os.path.join(output_folder_path, "spatial_TXx_and_TNn_trend_interpolation.jpg")
    plt.savefig(save_path, dpi=300)
    plt.close()

    # print(f"[INFO] Interpolated trend map saved at: {save_path}")


def plot_trend_boxplot_by_station(trend_df, output_folder_path):

    trend_df_cleaned = trend_df.dropna(subset=[
        "trend_max_temperature", "trend_min_temperature", "trend_avg_temperature"
    ], how="all")

    data_to_plot = {
        "Tx (Max Temp)": trend_df_cleaned["trend_max_temperature"].dropna(),
        "Tn (Min Temp)": trend_df_cleaned["trend_min_temperature"].dropna(),
        "Tavg (Avg Temp)": trend_df_cleaned["trend_avg_temperature"].dropna()
    }

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.boxplot(data_to_plot.values(), labels=data_to_plot.keys(), showfliers=True)
    ax.set_ylabel("Trend (°C/year)")
    ax.set_title("Distribution of Station Trends (Theil–Sen)")
    ax.grid(True)

    plt.tight_layout()
    save_path = os.path.join(output_folder_path, "trend_boxplot_considering_all_stations.jpg")
    plt.savefig(save_path, dpi=300)
    plt.close()
    # print(f"[INFO] Boxplot saved to {save_path}")

    
    ## Second plot with TXx and TNn
    trend_df_cleaned = trend_df.dropna(subset=[
        "trend_TXx", "trend_TNn"
    ], how="all")

    data_to_plot = {
        "TXx (Annual Max)": trend_df_cleaned["trend_TXx"].dropna(),
        "TNn (Annual Min)": trend_df_cleaned["trend_TNn"].dropna(),
    }

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.boxplot(data_to_plot.values(), labels=data_to_plot.keys(), showfliers=True)
    ax.set_ylabel("Trend (°C/year)")
    ax.set_title("Distribution of Station Trends (Theil–Sen)")
    ax.grid(True)

    plt.tight_layout()
    save_path = os.path.join(output_folder_path, "trend_boxplot_for_TXx_and_TNn_considering_all_stations.jpg")
    plt.savefig(save_path, dpi=300)
    plt.close()