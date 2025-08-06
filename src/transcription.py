#!/usr/bin/env python
import os, argparse, glob, tempfile, shutil, warnings
import cv2
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
# from paddleocr import PaddleOCR,draw_ocr
from PIL import Image, ImageDraw, ImageFont
import pytesseract
import easyocr
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
import math
import random
from scipy.stats import trim_mean
from scipy.cluster.hierarchy import fcluster, linkage
from sklearn.cluster import DBSCAN
from datetime import datetime
from calendar import monthrange


def get_max_rows_from_filename(filename):
    """ Extract year and month from filename and determine max rows. """
    try:
        # Extract YYYYMMDD part from filename (assuming format: STN_YYYYMMDD_SF)
        parts = filename.split('_')  # Splitting by '_'
        if len(parts) < 3:
            raise ValueError("Unexpected filename format")

        date_part = parts[1]  # Extract YYYYMMDD
        year = int(date_part[:4])  # First 4 characters = Year
        month = int(date_part[4:6])  # Next 2 characters = Month
        
        # Get the number of days in the month (handles leap years)
        num_days = monthrange(year, month)[1]

        # Calculate max_rows dynamically
        max_rows = num_days + 12  # Adding 5-day totals & averages
        
        return max_rows
    except Exception as e:
        print(f"Error processing filename {filename}: {e}")
        return 43  # Default to the max value (31 days)


def organize_contours_midpoint(contours, filename):
    """
    Organizes contours into `max_rows` using **midpoint clustering**, ensuring:
    - The row median is calculated using only the **50% closest** boxes.
    - Misaligned boxes are reassigned to the best row.
    - No row has more than `max_columns` boxes.

    Parameters:
    --------------
    contours: list
        List of bounding boxes (contours) detected in the image.

    max_rows: int
        Expected number of rows.

    max_columns: int
        Expected number of columns.

    Returns:
    --------------
    sorted_rows: list of lists
        Bounding boxes grouped into ordered rows.
    """

    if not contours:
        return []

    # Step 1: Extract **midpoint y-coordinates** for clustering
    midpoints = np.array([cv2.boundingRect(c)[1] + cv2.boundingRect(c)[3] // 2 for c in contours]).reshape(-1, 1)

    # Step 2: Perform K-Means clustering on midpoints
    max_rows = get_max_rows_from_filename(filename)
    kmeans = KMeans(n_clusters=min(max_rows, len(midpoints)), random_state=0, n_init=50, tol=1e-2)
    kmeans.fit(midpoints)
    labels = kmeans.labels_

    # Step 3: Assign contours to rows based on clustering
    row_dict = {i: [] for i in range(max_rows)}
    for label, contour in zip(labels, contours):
        row_dict[label].append(contour)

    # Step 4: Compute **super strict row medians** using only the closest 50%
    row_medians = {}

    for i, row in row_dict.items():
        if len(row) > 2:
            y_mids = np.array([cv2.boundingRect(c)[1] + cv2.boundingRect(c)[3] // 2 for c in row])
            y_mids.sort()

            # Keep **only the middle 50% closest values**
            trimmed_y_mids = y_mids[len(y_mids) // 4 : 3 * len(y_mids) // 4]

            # Compute median of the **trimmed** values
            row_medians[i] = np.median(trimmed_y_mids)
        else:
            row_medians[i] = np.median([cv2.boundingRect(c)[1] + cv2.boundingRect(c)[3] // 2 for c in row])

    # Step 5: Move misplaced boxes to the best row
    adjusted_rows = {i: [] for i in range(max_rows)}

    for i, row in row_dict.items():
        for box in row:
            y_mid = cv2.boundingRect(box)[1] + cv2.boundingRect(box)[3] // 2  # **Midpoint y-coordinate**
            
            # Find closest **trusted** row median (ignoring outliers)
            closest_row = i
            min_distance = abs(y_mid - row_medians[i])

            if i > 0:  # Check row above
                distance_up = abs(y_mid - row_medians[i-1])
                if distance_up < min_distance:
                    min_distance = distance_up
                    closest_row = i-1

            if i < max_rows - 1:  # Check row below
                distance_down = abs(y_mid - row_medians[i+1])
                if distance_down < min_distance:
                    closest_row = i+1

            adjusted_rows[closest_row].append(box)

    # Step 6: Sort each adjusted row again (left to right)
    for i in range(len(adjusted_rows)):
        adjusted_rows[i] = sorted(adjusted_rows[i], key=lambda c: cv2.boundingRect(c)[0])

    # Convert to final sorted list
    sorted_rows = [adjusted_rows[i] for i in range(max_rows)]

    return sorted_rows


def organize_contours_top(contours, filename):
    """
    Organizes contours into `max_rows` using **top edge clustering**, ensuring:
    - The row median is calculated using only the **50% closest** boxes.
    - Misaligned boxes are reassigned to the best row.
    - No row has more than `max_columns` boxes.

    Parameters:
    --------------
    contours: list
        List of bounding boxes (contours) detected in the image.

    max_rows: int
        Expected number of rows.

    max_columns: int
        Expected number of columns.

    Returns:
    --------------
    sorted_rows: list of lists
        Bounding boxes grouped into ordered rows.
    """

    if not contours:
        return []

    # Step 1: Extract **top edge y-coordinates** for clustering
    top_edges = np.array([cv2.boundingRect(c)[1] for c in contours]).reshape(-1, 1)

    # Step 2: Perform K-Means clustering on top edges
    max_rows = get_max_rows_from_filename(filename)
    kmeans = KMeans(n_clusters=min(max_rows, len(top_edges)), random_state=0, n_init=50, tol=1e-2)
    kmeans.fit(top_edges)
    labels = kmeans.labels_

    # Step 3: Assign contours to rows based on clustering
    row_dict = {i: [] for i in range(max_rows)}
    for label, contour in zip(labels, contours):
        row_dict[label].append(contour)

    # Step 4: Compute **super strict row medians** using only the closest 50%
    row_medians = {}

    for i, row in row_dict.items():
        if len(row) > 2:
            y_tops = np.array([cv2.boundingRect(c)[1] for c in row])
            y_tops.sort()

            # Keep **only the middle 50% closest values**
            trimmed_y_tops = y_tops[len(y_tops) // 4 : 3 * len(y_tops) // 4]

            # Compute median of the **trimmed** values
            row_medians[i] = np.median(trimmed_y_tops)
        else:
            row_medians[i] = np.median([cv2.boundingRect(c)[1] for c in row])

    # Step 5: Move misplaced boxes to the best row
    adjusted_rows = {i: [] for i in range(max_rows)}

    for i, row in row_dict.items():
        for box in row:
            y_top = cv2.boundingRect(box)[1]  # **Top edge y-coordinate**
            
            # Find closest **trusted** row median (ignoring outliers)
            closest_row = i
            min_distance = abs(y_top - row_medians[i])

            if i > 0:  # Check row above
                distance_up = abs(y_top - row_medians[i-1])
                if distance_up < min_distance:
                    min_distance = distance_up
                    closest_row = i-1

            if i < max_rows - 1:  # Check row below
                distance_down = abs(y_top - row_medians[i+1])
                if distance_down < min_distance:
                    closest_row = i+1

            adjusted_rows[closest_row].append(box)

    # Step 6: Sort each adjusted row again (left to right)
    for i in range(len(adjusted_rows)):
        adjusted_rows[i] = sorted(adjusted_rows[i], key=lambda c: cv2.boundingRect(c)[0])

    # Convert to final sorted list
    sorted_rows = [adjusted_rows[i] for i in range(max_rows)]

    return sorted_rows










def organize_contours_fraction(contours, max_rows, fraction=0.33):
    '''
    Organizes the bounding boxes (contours) in rows by a fractional point between the top and midpoint using KMeans clustering.

    Parameters
    --------------
    contours: list of bounding boxes (with their x, y, w, h coordinates)
        List of contours for the detected text in the table cells with coordinates.

    max_rows: int
        Maximum rows, adjusted based on your table's expected structure.

    fraction: float, optional (default=0.33)
        The fractional point of the height from the top for vertical clustering. A value between 0 (top) and 1 (bottom).
    
    Returns
    --------------
    rows: Bounding boxes organized in rows using KMeans clustering.
    '''

    # Calculate the point at the specified fraction between the top and bottom
    fraction_points = [(cv2.boundingRect(contour)[1] + int(cv2.boundingRect(contour)[3] * fraction)) for contour in contours]
    
    if len(fraction_points) == 0:
        return []

    # Apply KMeans clustering based on the fractional points
    kmeans = KMeans(n_clusters=min(max_rows, len(fraction_points)), random_state=0)
    kmeans.fit(np.array(fraction_points).reshape(-1, 1))
    labels = kmeans.labels_

    rows = [[] for _ in range(max_rows)]
    for label, contour in zip(labels, contours):
        rows[label].append(contour)

    # Sort each row by the x-coordinate for left-to-right ordering within rows
    for i in range(len(rows)):
        rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])

    return rows


def calculate_cell_reference(x, w, row_index, assigned_columns_per_row, max_columns, table_width):
    '''
    Ensures that there is always only one box per column by tracking assigned columns.

    If two cells are assigned to the same column, the second one is moved to the next available column.

    Parameters
    --------------
    x : float or int
        The x-coordinate of the cell's **top-left** corner.
    
    w : float or int
        The width of the bounding box.
    
    row_index : int
        The row number in the table to which the cell belongs.
    
    max_columns : int
        The maximum number of columns in the table. Ensures the column index does not exceed limits.
    
    table_width : int or float
        The total width of the table (in pixels). Used to determine the relative position of `x_center` 
        within the table and calculate the corresponding column index.

    assigned_columns_per_row : dict
        A dictionary tracking assigned columns for each row to ensure uniqueness.

    Returns
    --------------
    cell_reference : str
        The Excel-style cell reference (e.g., 'B3', 'C7') corresponding to the detected cell's position within the table.
    '''

    # Use the **center top x-coordinate** instead of just `x`
    x_center = x + (w / 2)

    # Calculate column index based on x_center instead of x
    column = math.floor(x_center / table_width * max_columns) + 1

    # Ensure the column index is within valid ranges
    column = max(1, min(column, max_columns))

    # Ensure the row exists in the tracking dictionary
    if row_index not in assigned_columns_per_row:
        assigned_columns_per_row[row_index] = set()

    # If the column is already assigned, shift to the next available column
    while column in assigned_columns_per_row[row_index] and column <= max_columns:
        column += 1  # Move to the next column

    # Ensure the new column is within range
    column = min(column, max_columns)

    # Register the assigned column
    assigned_columns_per_row[row_index].add(column)

    return f'{openpyxl.utils.get_column_letter(column)}{row_index}'



def generate_random_colors(n):
    '''
    Generates a list of `n` random distinct colors in BGR format.

    This function creates a list of distinct colors by generating random values for hue, saturation, and value (HSV) and then converting them to the BGR color space, which is commonly used in OpenCV. The generated colors are vivid and bright, ensuring they are visually distinct when used in applications such as object detection, image segmentation, or visual annotations.

    Parameters
    --------------
    n : int
        The number of distinct colors to generate.

    Returns
    --------------
    colors : list of tuples
        A list containing `n` tuples, each representing a color in BGR format. The colors are designed to be vivid and distinct for clear visualization.
    '''


    colors = []
    for i in range(n):
        # Generate random values for hue, saturation, and value
        hue = random.randint(0, 179)  # Hue range in OpenCV is [0, 179]
        saturation = random.randint(100, 255)  # To ensure the colors are vivid
        value = random.randint(100, 255)  # To ensure the colors are bright

        # Convert the random HSV color to BGR
        color = cv2.cvtColor(np.uint8([[[hue, saturation, value]]]), cv2.COLOR_HSV2BGR)[0][0].tolist()
        colors.append((int(color[0]), int(color[1]), int(color[2])))
    return colors


def draw_row_markers_and_boxes(image, rows, colors):
    '''
    Draws bounding boxes around contours in each row and adds numbered markers to each row with distinct colors.
    Ensures numbering is continuous (1 to 43), even for None rows.
    '''

    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.5
    thickness = 1

    for idx in range(len(rows)):  # Iterate over full 43-row structure
        row = rows[idx]  # Get row (can be None)
        color = colors[idx % len(colors)]  # Cycle through colors if needed

        y_position = 50 + (idx * 62)  # Approximate vertical position based on row index
        x_position = 10  # Arbitrary x position for marker

        # Always number rows, even if None
        cv2.putText(image, str(idx + 1), (x_position, y_position), font, font_scale, color, thickness)

        if row is None:
            continue  # Skip drawing bounding boxes for None rows

        # Draw bounding boxes for each contour in the row
        for contour in row:
            x, y, w, h = cv2.boundingRect(contour)

            # Define increase factors for bounding box modification
            increase_factor_width = 0.07  # 
            increase_factor_height = 0.35  # 

            # Expand width while keeping it centered
            new_w = int(w * (1 + increase_factor_width))  # Increase width
            x = max(0, x - (new_w - w) // 2)  # Adjust x to keep center fixed

            # Expand height symmetrically
            new_h = int(h * (1 + increase_factor_height * 2))  # Increase height
            y = max(0, y - (new_h - h) // 2)  # Adjust y to keep center fixed

            # Ensure bounding box remains within valid image bounds
            x = max(0, x)
            y = max(0, y)
            w = max(1, new_w)  # Avoid zero or negative width
            h = max(1, new_h)  # Avoid zero or negative height

            # Draw bounding box on the visualization image
            cv2.rectangle(image, (x, y), (x + w, y + h), color, 2)



def calculate_trimmed_mean(values, proportion_to_cut=0.2):
    '''
    Calculates the trimmed mean of a list of values, excluding a specified proportion of the smallest and largest values.

    This function computes the trimmed mean of a given list of numerical values. The trimmed mean is a measure of central tendency that removes a certain proportion of the lowest and highest values before calculating the mean. This helps to reduce the effect of outliers on the mean.

    Parameters
    --------------
    values : list
        A list of numerical values for which the trimmed mean is to be calculated.
    
    proportion_to_cut : float, optional
        The proportion of values to remove from each end of the sorted list before calculating the mean. 
        For example, a proportion of 0.2 means removing the lowest 20% and the highest 20% of the values. 
        The default value is 0.2.

    Returns
    --------------
    float
        The trimmed mean of the input values, after excluding the specified proportion of extreme values.
    '''

    return trim_mean(values, proportion_to_cut)


def merge_excel_files(file1, file2, output_file, start_row, end_row):
    '''
    Merges two Excel files for verification purposes: one organized by the mid-point coordinates of bounding boxes and the other by the top coordinates.

    This function merges two preprocessed Excel files that contain transcribed data organized differently (one by mid-point and the other by top coordinates of bounding boxes). 
    The merged output file allows for cross-checking to ensure that cells are correctly placed in their respective rows.

    Parameters
    --------------
    file1: str
        The path to the Excel file containing transcribed data organized in rows using the top coordinates of the bounding boxes (contours).
    file2: str
        The path to the Excel file containing transcribed data organized in rows using the mid-point coordinates of the bounding boxes (contours).
    output_file: str
        The path where the merged Excel file will be saved.
    start_row: int
        The starting row number from which to begin the merge.
    end_row: int
        The ending row number up to which the merge should be conducted.

    Returns
    --------------
    None
        The function creates and saves a merged Excel file at the specified `output_file` location. This file combines the data from `file1` and `file2` for further verification.
    '''


    # Load the Excel files into DataFrames without headers
    df1 = pd.read_excel(file1, header=None)
    df2 = pd.read_excel(file2, header=None)


    # If the indices are not simple integers or do not align with Excel rows as expected,
    # you might need to reset them or adjust how the Excel file is being read (e.g., `index_col=None`)
    df1 = df1.reset_index(drop=True)
    df2 = df2.reset_index(drop=True)

    # Convert start_row and end_row to zero-based index for Python
    start_idx = start_row - 1  # Convert 1-based index to 0-based
    end_idx = end_row - 1  # Convert 1-based index to 0-based

    # Slice to only include the range from start_idx to end_idx
    df1 = df1.iloc[start_idx:end_idx + 1]
    df2 = df2.iloc[start_idx:end_idx + 1]

    # Initialize a new DataFrame to hold merged results
    merged_df = pd.DataFrame(index=df1.index, columns=df1.columns)

    # Iterate over rows by index (assuming the indices are aligned)
    for idx in df1.index:
        for col in df1.columns:
            val1 = df1.at[idx, col]
            val2 = df2.at[idx, col]
            # Simple merge logic: prefer non-empty values from df1, then df2
            if pd.notna(val1):
                merged_df.at[idx, col] = val1
            else:
                merged_df.at[idx, col] = val2


    # Create a new workbook and select the active worksheet
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active

    # Append the merged DataFrame to the new worksheet
    for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            new_worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Merge cells for multi-column headers
    new_worksheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1) #No de la pentade
    new_worksheet.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2) #Date
    new_worksheet.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3) #Bellani
    new_worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8) #Températures extrêmes
    new_worksheet.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10) #Evaportation
    new_worksheet.merge_cells(start_row=1, start_column=11, end_row=3, end_column=11) #Pluies
    new_worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=16) #Température et Humidité de l'air à 6 heures
    new_worksheet.merge_cells(start_row=1, start_column=17, end_row=1, end_column=21) #Température et Humidité de l'air à 15 heures
    new_worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=26) #Température et Humidité de l'air à 18 heures
    new_worksheet.merge_cells(start_row=1, start_column=27, end_row=3, end_column=27) #Date
    # subheaders
    new_worksheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7) #Abri
    new_worksheet.merge_cells(start_row=2, start_column=9, end_row=2, end_column=10) #Piche
    new_worksheet.merge_cells(start_row=2, start_column=12, end_row=2, end_column=16) #(Psychromètre a aspiration)
    new_worksheet.merge_cells(start_row=2, start_column=17, end_row=2, end_column=21) #(Psychromètre a aspiration)
    new_worksheet.merge_cells(start_row=2, start_column=22, end_row=2, end_column=26) #(Psychromètre a aspiration)

    # Set up border styles for excel output
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    # Loop through cells to apply borders
    for row in new_worksheet.iter_rows(min_row=1, max_row=new_worksheet.max_row, min_col=1, max_col=new_worksheet.max_column):
        for cell in row:
            cell.border = thin_border
    new_workbook.save(output_file)
    
    # Iterate through all cells and set the alignment
    for row in new_worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Read headers from the first row of one of the files
    workbook = openpyxl.load_workbook(file1)
    copy_file1 = workbook.active
    headers = [cell.value for cell in copy_file1[1]]  
    for row in new_worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=new_worksheet.max_column):
        for col_num, header in enumerate(headers, start=1):
            cell = new_worksheet.cell(row=1, column=col_num, value=header)
            if header == "No de la pentade" or header == "Date" or header == "Bellani (gr. Cal/cm2) 6-6h" or header == "Pluies en mm. 6-6h":
                cell.alignment = Alignment(textRotation=90)

    # Save the workbook
    new_workbook.save(output_file)



def add_missing_boxes(sorted_rows, max_cell_width_threshold=130, max_cell_height_threshold=50, max_columns=24):
    updated_rows = []

    for row in sorted_rows:
        if any(cell is None for cell in row):
            continue

        bounding_boxes = [cv2.boundingRect(c) for c in row]
        bounding_boxes.sort(key=lambda b: b[0])

        new_boxes = bounding_boxes.copy()
        gaps = []

        # Handle missing boxes at the start of the row
        if new_boxes and new_boxes[0][0] > max_cell_width_threshold:
            first_x, first_y, first_w, first_h = new_boxes[0]
            num_missing_boxes = min(int(first_x // max_cell_width_threshold), max_columns - len(new_boxes))

            new_boxes_at_start = []
            for i in range(num_missing_boxes):
                new_x = max(0, first_x - (num_missing_boxes - i) * max_cell_width_threshold)
                new_box = (new_x, first_y, max_cell_width_threshold, max_cell_height_threshold)
                new_boxes_at_start.append(new_box)

            new_boxes = new_boxes_at_start + new_boxes

        for i in range(len(new_boxes) - 1):
            x1, y1, w1, h1 = new_boxes[i]
            x2, _, _, _ = new_boxes[i + 1]
            gap = x2 - (x1 + w1)

            if gap > max_cell_width_threshold:
                gaps.append((gap, i, x1 + w1, y1))

        gaps.sort(reverse=True, key=lambda g: g[0])

        for gap, i, gap_start_x, y1 in gaps:
            if len(new_boxes) >= max_columns:
                break

            num_missing_boxes = min(int(gap // max_cell_width_threshold), max_columns - len(new_boxes))

            total_box_width = num_missing_boxes * max_cell_width_threshold
            # start_x = gap_start_x + (gap - total_box_width) / 2

            # new_boxes_in_gap = []
            # for j in range(num_missing_boxes):
            #     new_x = start_x + j * max_cell_width_threshold
            #     new_box = (int(new_x), y1, max_cell_width_threshold, max_cell_height_threshold)
            #     new_boxes_in_gap.append(new_box)
            
            # Dynamically compute width so that all boxes fit perfectly into the gap
            dynamic_cell_width = gap / (num_missing_boxes + 1)

            new_boxes_in_gap = []
            for j in range(num_missing_boxes):
                if len(new_boxes) + len(new_boxes_in_gap) >= max_columns:
                    break

                center_x = gap_start_x + (j + 1) * dynamic_cell_width
                new_x = int(center_x - dynamic_cell_width / 2)

                new_box = (new_x, y1, int(dynamic_cell_width), max_cell_height_threshold)
                new_boxes_in_gap.append(new_box)
            
                # if len(new_boxes) + len(new_boxes_in_gap) >= max_columns:
                #     break

            new_boxes[i + 1:i + 1] = new_boxes_in_gap

        new_boxes = new_boxes[:max_columns]

        updated_contours = [
            np.array([[x, y], [x + w, y], [x + w, y + h], [x, y + h]], dtype=np.int32)
            for x, y, w, h in new_boxes
        ]
        updated_rows.append(updated_contours)

    return updated_rows




def transcription(detected_table_cells, ocr_model, tesseract_path, transient_transcription_output_dir, pre_QA_QC_transcribed_hydroclimate_data_dir_station, station, month_filename, no_of_rows, no_of_columns, no_of_rows_including_headers):
    '''
    Performs OCR (Optical Character Recognition) on detected table cells from a pre-processed image 
    to extract and organize textual data into an Excel workbook.

    This function processes detected tables within an image by isolating table cells using contours, 
    clipping the cell regions, and applying OCR/HTR (Handwritten Text Recognition) to transcribe text. 
    The transcribed data is organized into a structured Excel workbook, complete with headers and formatting. 
    Multiple methods are employed to double-check bounding box placements and ensure robust data extraction.

    Parameters
    --------------
    detected_table_cells : list
        A list containing:
        - detected_table_cells[0]: contours. Contours representing detected table cells.
        - detected_table_cells[1]: image_with_all_bounding_boxes. Image with bounding boxes drawn around detected cells.
        - detected_table_cells[2]: table_copy. Processed table image for further operations.
        - detected_table_cells[3]: table_original_image. Original table image before processing.
    
    ocr_model : str
        The OCR/HTR model used for text recognition. Options include 'Tesseract-OCR', 'EasyOCR', or PaddleOCR
    
    tesseract_path : str
        Path to the Tesseract executable, required if 'Tesseract-OCR' is the selected model.
    
    transient_transcription_output_dir : str
        Directory for saving intermediate results, such as detected regions of interest (ROIs).
    
    pre_QA_QC_transcribed_hydroclimate_data_dir_station : str
        Directory for saving the final transcribed Excel file before QA/QC steps.
    
    station : str
        Identifier of the station (station no.), used for organizing output files.
    
    month_filename : str
        Filename of the processed image, representing a specific month and year.
    
    no_of_rows : int
        Expected number of rows in the detected table, excluding headers.
    
    no_of_columns : int
        Expected number of columns in the detected table.
    
    no_of_rows_including_headers : int
        Total number of rows in the table, including headers.

    Returns
    --------------
    path_to_save_merged_excel_file : str
        Path to the final transcribed Excel workbook containing extracted and organized text.

    '''
    
    
    if ocr_model == 'Tesseract-OCR':
        ## Lauching Tesseract-OCR
        pytesseract.pytesseract.tesseract_cmd = tesseract_path ## Here input the PATH to the Tesseract executable on your computer. See more information here: https://pypi.org/project/pytesseract/
    # if ocr_model == 'PaddleOCR':
    #     ## Lauching PaddleOCR, which would be used by downloading necessary files as shown below
    #     paddle_ocr = PaddleOCR(use_angle_cls=True, lang = 'en', use_gpu=False) ## Run only once to download all required files
    if ocr_model == 'EasyOCR':
        ## Lauching EasyOCR
        easyocr_reader = easyocr.Reader(['en']) # this needs to run only once to load the model into memory

    easyocr_reader = easyocr.Reader(['en'])

    # Contours of bounding boxes detected in cell recognition
    new_contours = detected_table_cells[0]

    image_with_all_bounding_boxes = detected_table_cells[1]
    table_copy = detected_table_cells[2] # Binarized table image using Adaptive Thresholding
    # table_copy = detected_table_cells[3] # Table image but as a clip of the original image (No Binarization)

    # Get the dimensions of the loaded image. Here, particulary the image/table width is very important for the column placement of cells/bounding boxes
    image_height, image_width, image_channels = image_with_all_bounding_boxes.shape

    # Image onto which the regions of interest (ROIs) i.e the cells, will be drawn for illustration purposes
    ROIs_image = table_copy.copy()
    
    results = []

    # Here we use two methods to arrange the boundung boxes (as a double check): (1) Using the middle coordinates of the bounding boxes , and (2) Using the top coordinated of the boudning boxes.
    organize_methods = {  
        'Midpoint': organize_contours_midpoint,
        'Top': organize_contours_top
    }

    # organize_methods = {  
    #         'Midpoint': organize_contours_by_column,
    #         'Top': organize_contours_by_column
    #     }


    for method_name, organize_method in organize_methods.items():

        ## Create an Excel workbook and add a worksheet where the transcribed text will be saved
        wb = Workbook()
        ws = wb.active
        ws.title = 'OCR_Results'

        # Organize the contours
        # organized_rows = organize_method(new_contours, no_of_rows)

        organized_rows = organize_method(new_contours, month_filename)
        # organized_rows = organize_method(new_contours, max_rows=43, row_threshold=50)

        # Sorting the cell (bounding box) rows from first to last using trimmed mean of coordinates
        sorted_rows = sorted(organized_rows, key=lambda row: calculate_trimmed_mean([cv2.boundingRect(c)[1] + cv2.boundingRect(c)[3] // 2 for c in row]))

        # Get the expected max rows (dynamically)
        max_rows = get_max_rows_from_filename(month_filename)

        # Calculate how many rows need to be padded
        missing_rows = 43 - max_rows  # Always ensuring 43 rows

        # If there are missing rows, pad `None` before the last two rows (totals & averages)
        if missing_rows > 0:
            # Ensure we don't disturb the last two rows
            sorted_rows = sorted_rows[:-2] + [[None]] * missing_rows + sorted_rows[-2:]

        sorted_rows = add_missing_boxes(sorted_rows)  # Add missing boxes per row where there are gaps

        ## FOR ILLUSTRATION PURPOSES: Uncomment the following lines if you want to see an example of how the sorting function works. This is for illustration purposes only and not necessary for the main functionality of the script. 
        
        # Create a copy of the image for visualization
        image_before_sorting = table_copy.copy()
        image_after_sorting = table_copy.copy()

        # Generate colors for 43 rows. For sorted row visualitation purposes
        colors = generate_random_colors(no_of_rows)  # Random colours for the maximum number of rows, such that each row has its own color for easy identification

        # Draw markers on the image before sorting
        draw_row_markers_and_boxes(image_before_sorting, organized_rows, colors)  # Green color for original order

        # Draw markers on the image after sorting
        draw_row_markers_and_boxes(image_after_sorting, sorted_rows, colors)  # Red color for sorted order

        # Ensure save directory exists
        save_dir = os.path.join(transient_transcription_output_dir, station)
        os.makedirs(save_dir, exist_ok=True)

        # Define save paths
        save_path_before = os.path.join(save_dir, f'before_sorting_{method_name}_{month_filename}.png')
        save_path_after = os.path.join(save_dir, f'after_sorting_{method_name}_{month_filename}.png')

        # Save images
        cv2.imwrite(save_path_before, image_before_sorting)
        cv2.imwrite(save_path_after, image_after_sorting)

        # Load images for display
        img_before = cv2.imread(save_path_before)
        img_before = cv2.cvtColor(img_before, cv2.COLOR_BGR2RGB)
        
        img_after = cv2.imread(save_path_after)
        img_after = cv2.cvtColor(img_after, cv2.COLOR_BGR2RGB)

        # # # Save or display the images for inspection
        # plt.imshow(img_before)
        # plt.show()

        # plt.imshow(img_after)
        # plt.show()
    

        # Sort boxes within each column of each row by y-coordinate
        for row in sorted_rows:
            if row is not None:  # Skip None values to avoid errors
                row.sort(key=lambda c: cv2.boundingRect(c)[1])

        # Dictionary to track assigned columns per row
        assigned_columns_per_row = {}
        for row_index, row in enumerate(sorted_rows, start=1):
            if row is None:
                continue  # Skip processing for empty rows
            
            for contour in row:

                x, y, w, h = cv2.boundingRect(contour)

                # Define increase factors for bounding box modification
                increase_factor_width = 0.07  # Increase width by 7%
                increase_factor_height = 0.20  # Increase height by 20%

                # Expand width while keeping it centered
                new_w = int(w * (1 + increase_factor_width))  # Increase width
                x = max(0, x - (new_w - w) // 2)  # Adjust x to keep center fixed

                # Expand height symmetrically
                new_h = int(h * (1 + increase_factor_height * 2))  # Increase height
                y = max(0, y - (new_h - h) // 2)  # Adjust y to keep center fixed

                # Ensure bounding box remains within valid image bounds
                x = max(0, x)
                y = max(0, y)
                w = max(1, new_w)  # Avoid zero or negative width
                h = max(1, new_h)  # Avoid zero or negative height

                 # Draw bounding box on the visualization image
                cv2.rectangle(ROIs_image, (x, y), (x + w, y + h), (0, 0, 255), 3)

                # ********
                # OCR
                # Crop each cell using the bounding rectangle coordinates
                ROI = table_copy[y:y+h, x:x+w]  # Ensure consistency with visualization bounding boxes
                # ********

                # # Crop raw cell from original (non-binarized) image
                # cell_crop_color = table_copy[y:y+h, x:x+w]  # Now table_copy is the original image

                # # Convert to grayscale
                # gray_cell = cv2.cvtColor(cell_crop_color, cv2.COLOR_BGR2GRAY)

                # # Binarize using INVERTED thresholding (to thicken digits)
                # # binarized_inv = cv2.adaptiveThreshold(gray_cell, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 35, 4)  # (21,4) was good.
                # binarized_inv = cv2.adaptiveThreshold(gray_cell, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 51,6)   # ahd it at (51,6) before   (35,6) was good but make some light text disappear. 

                # # Invert back to get black text on white background
                # binarized_cell = cv2.bitwise_not(binarized_inv)

                # # Assign this as your final ROI for OCR
                # ROI = binarized_cell


                # ### FOR ILLUSTRATION PURPOSES: This line below is about drawing a rectangle on the image with the shape of the bounding box. Its not needed for the OCR. This is only for debugging purposes.
                # # image_with_all_bounding_boxes = cv2.rectangle(image_with_all_bounding_boxes, (x, y), (x + w, y + h), (0, 255, 0), 5)

                # # # Draw the adjusted ROI on the output image
                # # cv2.rectangle(ROIs_image, (x, y), (x + w, y + h), (0, 255, 0), 5)  # (0, 255, 0) represent a green color for ROI, and 5 is the thicnkess of the ROI bounbdary boxes
                
                # # Draw the updated bounding box
                # cv2.rectangle(ROIs_image, (x, y), (x + new_w, y + new_h), (0, 255, 0), 4)

                
                # OCR
                
                if ROI.size != 0:  # Check if the height and width are greater than zero. This is to prevent invalid ROIs
                    
                    # Save the detected text image/ROI
                    save_dir =  os.path.join(transient_transcription_output_dir, station)
                    os.makedirs(save_dir, exist_ok=True)  # Ensure the directory exists
                    save_path_detected_text = os.path.join(save_dir, 'detected.png')
                    cv2.imwrite(save_path_detected_text, ROI)
                    if ocr_model == 'Tesseract-OCR':
                    # Using Tesseract-OCR
                        ocr_result = pytesseract.image_to_string(save_path_detected_text, lang='cobecore-V6', config='--oem 1 --psm 7 -c tessedit_char_whitelist=0123456789') # Just added -c tessedit_char_whitelist=0123456789 to really limit the text type/values detected

                        # Here's a brief explanation of some Page Segmentation Modes (PSMs) available in Tesseract:
                        # 0: Orientation and script detection (OSD) only.
                        # 1: Automatic page segmentation with OSD.
                        # 2: Automatic page segmentation, but no OSD, or OCR.
                        # 3: Fully automatic page segmentation, but no OSD. (Default)
                        # 4: Assume a single column of text of variable sizes.
                        # 5: Assume a single uniform block of vertically aligned text.
                        # 6: Assume a single uniform block of text.
                        # 7: Treat the image as a single text line.
                        # 8: Treat the image as a single word.
                        # 9: Treat the image as a single word in a circle.
                        # 10: Treat the image as a single character.
                        # 11: Sparse text. Find as much text as possible in no particular order.
                        # 12: Sparse text with OSD.
                        # 13: Raw line. Treat the image as a single text line, bypassing hacks that are Tesseract-specific.

                    # Uncomment the following lines if you'd like to make use of PaddleOCR
                    # if ocr_model == 'PaddleOCR':
                    #     ## Using PaddleOCR
                    #     ocr_result = paddle_ocr.ocr('detected.png', cls = True)

                    if ocr_model == 'EasyOCR':
                    # Using EasyOCR
                        ocr_result = easyocr_reader.readtext(save_path_detected_text, detail = 0, allowlist='0123456789')
                        # In EasyOCR, the detail parameter specifies the level of detail in the output. 
                                # When using the readtext method, the detail parameter can be set to different values to control what kind of output you get. 
                                # Specifically:
                                # detail=1: The output will be a list of tuples, where each tuple contains detailed information about the detected text, 
                                # including the bounding box coordinates, the text string, and the confidence score. Example: [(bbox, text, confidence), ...].

                                # detail=0: The output will be a list of strings, where each string is the detected text without any additional details. 
                                # This is a simpler output format that only provides the recognized text. Example: ["text1", "text2", ...].
                        if isinstance(ocr_result, list): # This is because EasyOCR's results are returned as a list
                                ocr_result = ''.join(ocr_result)  # Convert list to a string
                    
                    
                    # Using OCR for handwritten text recognition
                    if ocr_result is not None:
                        
                        if not ocr_result.strip(): # Check if the result is empty or only whitespace. This could be due to the selected OCR (in this case: Tesseract-OCR) not being able to recognize the text in the ROI.
                            # For this reason, we can try another OCR, say for example Easy OCR, to try to recognize the text in this ROI
                            ocr_result = easyocr_reader.readtext(save_path_detected_text, detail = 0, allowlist='0123456789')
                            if isinstance(ocr_result, list): # This is because EasyOCR's results are returned as a list
                                ocr_result = ''.join(ocr_result)  # Convert list to a string


                                               
                        # Attain the Ms Excel Template cell coordinates
                        # # Determine the cell reference using the x coodrinate of the bounding box, row index, maximum column number, and image/table width
                        # cell_ref = calculate_cell_reference(x, row_index, max_columns=24, table_width=image_width) # e.g., A1, B5, etc.

                        # # Dictionary to track assigned columns per row
                        # assigned_columns_per_row = {}
                        # cell_ref = calculate_cell_reference(x, new_w, row_index, max_columns=24, table_width=image_width) # e.g., A1, B5, etc.
                        cell_ref = calculate_cell_reference(x, w, row_index, assigned_columns_per_row, max_columns=24, table_width=image_width)
                        # print(f"saving to {cell_ref}")
                        # # Additional check: Incase the cell (cell_ref) is already occupied with transcribed text. We then opt for the cell below within the same column
                        # column_letter = openpyxl.utils.get_column_letter(math.floor(x / image_width * no_of_columns) + 1)
                        # initial_row_index = row_index  # Store the initial row index
                        # # Check if the cell is already occupied
                        # if ws[cell_ref].value is not None:
                        #     row_index += 1
                        #     cell_ref = f'{column_letter}{row_index}'
                        
                        # Place the OCR/HTR recognized text in its respective Ms Excel cell 
                        ws[cell_ref].value = ocr_result   

                        # # Restore the row index to the initial value
                        # row_index = initial_row_index

                        # Set up border styles for excel output
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

                        # Loop through cells to apply borders
                        for row in ws.iter_rows(min_row=1, max_row=no_of_rows, min_col=1, max_col=no_of_columns):
                            for cell in row:
                                cell.border = thin_border

                        
                        ## CONTINOUS LEARNING (MACHINE-LEARNING) OF THE OCR
                        
                        # Adjust cell_ref to match final Excel layout: 2 columns inserted + 3 header rows added
                        adjusted_column_index = openpyxl.utils.column_index_from_string(cell_ref[0]) + 2  # +2 for two inserted columns on the left side of the excel sheet. In our case these were the 'Pentad No.' and the Date. Change this depending on your sheets or make it +0 (Zero) if there were no extra columns to the left
                        adjusted_row_index = int(cell_ref[1:]) + int(no_of_rows_including_headers - no_of_rows)  # In our case: = +3 for three inserted header rows. See # SPECIAL ADDITION TO CODE / CUSTOMIZATION below

                        # Convert adjusted column index back to letter
                        adjusted_column_letter = openpyxl.utils.get_column_letter(adjusted_column_index)

                        # Construct adjusted Excel-style reference
                        adjusted_cell_ref = f"{adjusted_column_letter}{adjusted_row_index}"
                        
                        # Construct a filename for the clipped cell imgae (ROI) using month_filename and Excel cell reference
                        roi_filename = f"{month_filename}_{adjusted_cell_ref}.png"

                        # Save the clipped cell image (ROI) for further training of the OCR model; For continous learning of different handwritting styles usign correct/confirmed values in the Quality Assessment and Quality Control module
                        # Set path to save
                        roi_save_dir = os.path.join(transient_transcription_output_dir, station, 'clipped_cells')
                        os.makedirs(roi_save_dir, exist_ok=True)

                        full_roi_save_path = os.path.join(roi_save_dir, roi_filename)

                        # Save the ROI image
                        cv2.imwrite(full_roi_save_path, ROI)

                    else:
                        print('No values detected in clip')
                else:
                    print('ROI is empty or invalid')


        ### SPECIAL ADDITION TO CODE / CUSTOMIZATION: The lines below are particular to our tables. Here we replace the headers and row labels as in the original table format. This will vary from your setup and this should be customized to your particular sheets.
        # Insert two columns on the left side of the excel sheet
        ws.insert_cols(1, 2)
        
        # Insert a new row at the top for headers
        ws.insert_rows(1, amount = 3)

        # Define your headers (adjust as needed)
        headers = ["No de la pentade", "Date", "Bellani (gr. Cal/cm2) 6-6h", "Températures extrêmes", "", "", "", "", "Evaportation en cm3 6 - 6h", "", "Pluies en mm. 6-6h", "Température et Humidité de l'air à 6 heures", "", "", "", "", "Température et Humidité de l'air à 15 heures",  "", "", "", "", "Température et Humidité de l'air à 18 heures",  "", "", "", "", "Date"]
        sub_headers_1 = ["", "", "", "Abri", "", "", "", "", "Piche", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", ""]
        sub_headers_2 =["", "", "", "Max.", "Min.", "(M+m)/2", "Ampl.", "Min. gazon", "Abri.", "Ext.", "", "T", "T'a", "e.", "U", "∆e", "T", "T'a", "e.", "U", "∆e","T", "T'a", "e.", "U", "∆e", ""]

        # Add the headers to the first row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            if header == "No de la pentade" or header == "Date" or header == "Bellani (gr. Cal/cm2) 6-6h" or header == "Pluies en mm. 6-6h":
                cell.alignment = Alignment(textRotation=90)

        # Add the first row of sub-headers to the second row
        for col_num, sub_header in enumerate(sub_headers_1, start=1):
            ws.cell(row=2, column=col_num, value=sub_header)

        # Add the second row of sub-headers to the third row
        for col_num, sub_header in enumerate(sub_headers_2, start=1):
            ws.cell(row=3, column=col_num, value=sub_header)
        
        # Merge cells for multi-column headers
        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1) #No de la pentade
        ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2) #Date
        ws.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3) #Bellani
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8) #Températures extrêmes
        ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10) #Evaportation
        ws.merge_cells(start_row=1, start_column=11, end_row=3, end_column=11) #Pluies
        ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=16) #Température et Humidité de l'air à 6 heures
        ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=21) #Température et Humidité de l'air à 15 heures
        ws.merge_cells(start_row=1, start_column=22, end_row=1, end_column=26) #Température et Humidité de l'air à 18 heures
        ws.merge_cells(start_row=1, start_column=27, end_row=3, end_column=27) #Date
        # subheaders
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7) #Abri
        ws.merge_cells(start_row=2, start_column=9, end_row=2, end_column=10) #Piche
        ws.merge_cells(start_row=2, start_column=12, end_row=2, end_column=16) #(Psychromètre a aspiration)
        ws.merge_cells(start_row=2, start_column=17, end_row=2, end_column=21) #(Psychromètre a aspiration)
        ws.merge_cells(start_row=2, start_column=22, end_row=2, end_column=26) #(Psychromètre a aspiration)


        # Label Date, Total and Average rows
        row_labels = ["1","2", "3", "4", "5", "Tot.", "Moy.", "6", "7", "8", "9", "10", "Tot.", "Moy.", "11", "12", "13", "14", "15", "Tot.", "Moy.", "16", "17", "18", "19", "20", "Tot.", "Moy.", "21", "22", "23", "24", "25", "Tot.", "Moy.", "26", "27", "28", "29", "30", "31", "Tot.", "Moy.", "Tot.", "Moy."]
        # Update the cells in the second and last column with the date values
        columns = [2, 27]
        for col in columns:
            for i, value in enumerate(row_labels, start=4):
                cell = ws.cell(row=i, column=col)
                cell.value = value
                # wb.save(new_version_of_file) # Save the modified workbook
        
        # Save Excel file
        file_path = os.path.join(save_dir, f'{method_name}_Excel_with_OCR_Results.xlsx')
        wb.save(file_path)
        results.append([file_path])

        
        ### FOR ILLUSTRATION PURPOSES: Uncomment the following lines if you want to see the ROIs per table
        # plt.imshow(ROIs_image)
        # plt.show()
    
    # Update paths to include the new Excel files (from the methods above) for merging. This is just a double check for correct cell placement
    top_excel_path = os.path.join(save_dir, 'Top_Excel_with_OCR_Results.xlsx')
    midpoint_excel_path = os.path.join(save_dir, 'Midpoint_Excel_with_OCR_Results.xlsx')
    path_to_save_merged_excel_file = os.path.join(pre_QA_QC_transcribed_hydroclimate_data_dir_station, f'{month_filename}_pre_QA_QC.xlsx')
    merge_excel_files(top_excel_path, midpoint_excel_path, path_to_save_merged_excel_file, 1, no_of_rows_including_headers) # this prioritizes the top coordinates of the bounding box to the mid point coordinates when placing the transcribed data into an excel sheet. But considers the best placement for both as double check. Here the 1 represent the first row and the 46 represents the last possible row to perform the merging of the excel files. In our case we have 43 rows with data + 3 header rows = 46.
    # merge_excel_files(midpoint_excel_path, top_excel_path, path_to_save_merged_excel_file, 1, no_of_rows_including_headers) # this prioritizes the midpoint coordinates of the bounding box to the top coordinates when placing the transcribed data into an excel sheet. But considers the best placement for both as double check. Here the 1 represent the first row and the 46 represents the last possible row to perform the merging of the excel files. In our case we have 43 rows with data + 3 header rows = 46.


    return path_to_save_merged_excel_file