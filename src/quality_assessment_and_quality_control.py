import os
import openpyxl
from openpyxl.styles import PatternFill
import shutil
import openpyxl.utils
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from statistics import mean
import re
import numpy as np

# Setting up the current working directory; for both the input and output folders
cwd = os.getcwd()



def is_string_convertible_to_float(value):
    '''
    Determines if a given value can be converted to a float.

    This function is used to verify if a value in a cell, such as one from an MS Excel sheet, can be converted to a float. 
    It ensures that the value can be used in subsequent numerical calculations without causing errors.

    Parameters
    --------------
    value: Any
        The value from a cell in an MS Excel sheet. This can be of any data type, including string, integer, float, or None.

    Returns
    --------------
    bool
        Returns True if the value can be converted to a float, indicating that it is safe for numerical calculations.
        Returns False if the value cannot be converted to a float or if the value is None (e.g., for empty cells).
    '''

    if value is None: # Check to handle None cases (empty cells)
        return False
    try:
        float(value)
        return True
    except ValueError:
        return False


def clean_numeric_string(value):
    """
    Cleans a transcribed numeric value by removing spaces, tabs, and newlines.

    This function ensures that the output is a properly formatted number string.

    Parameters
    ----------
    value : Any
        The value to clean.
    
    Returns
    -------
    str
        A cleaned string of transcribed values without spaces, tabs or new lines.
        Returns an empty string if the input is not convertible.
    """
    if value is None:
        return ""

    # Convert to string, remove spaces, tabs, and newlines
    cleaned_value = str(value).strip().replace("\n", "").replace("\t", "").replace(" ", "")

    return cleaned_value



def highlight_change(color, worksheet_and_cell_coordinate, filename):
    '''
    Highlights a cell in an Excel worksheet to indicate the status of data processing or transcription.

    This function applies a color highlight to a specific cell in an Excel worksheet to visually represent one of the following scenarios:
    (1) A correction has been made during data post processing (QA/QC checks) to fix an error.
    (2) A value has been confirmed as correctly transcribed.
    (3) A value has been identified as wrongly transcribed.

    Parameters
    --------------
    color: str
        The color to be used for highlighting, which corresponds to a specific status as detailed in the "Key_for_post_processed_data_sheets" document in docs (in the repository).
    worksheet_and_cell_coordinate: openpyxl.cell.Cell
        The specific cell in the worksheet that needs to be highlighted.
    filename: str
        The name or path of the Excel file containing the worksheet.

    Returns
    --------------
    None
        The function applies the highlight directly to the Excel file and does not return any value. The specified cell in the Excel file will be highlighted according to the provided color.
    '''
    
    # Highlight cells with strings instead of floats
    highlighting_color = color # Highlighting color of choice
    highlighting_strings = PatternFill(start_color = highlighting_color, end_color = highlighting_color, fill_type = 'solid')
    cell_to_highlight = worksheet_and_cell_coordinate
    cell_to_highlight.fill = highlighting_strings



def is_highlighted(cell, color):
    '''
    Checks if a cell is highlighted with a specific color during earlier post-processing steps.

    This function is used to determine whether a particular cell in an Excel worksheet has been highlighted with a specified color. 
    This check is essential for identifying cells that may need corrections based on the errors identified during earlier post-processing.

    Parameters
    --------------   
    cell: openpyxl.cell.Cell
        The cell in the worksheet to check for highlighting.
    color: str
        The color (in RGB or hex format) to compare against the cell's current highlight color.

    Returns
    -------------- 
    bool
        Returns True if the cell is highlighted with the specified color, indicating that it was marked during post-processing.
        Returns False if the cell is not highlighted with the specified color.
    '''


    fill = cell.fill.start_color
    if isinstance(fill, openpyxl.styles.colors.Color):
        return fill.rgb == color
    return False




def has_two_digits_in_order(value, calculated_value):
    """
    Determines if at least two consecutive digits from the calculated value appear in the correct order within the transcribed value, ignoring any decimal points.

    This function checks whether any pair of consecutive digits from the `calculated_value` are present in the transcribed `value`, in the same sequence. It is useful for validating or comparing numerical strings where the order of digits is important, such as in certain data verification tasks.

    Parameters:
    --------------
    value: str
        The string to be checked, typically a value extracted from a dataset.
    calculated_value: str
        The reference string containing the calculated value, against which the `value` is compared.

    Returns:
    --------------
    bool
        Returns True if at least two consecutive digits from `calculated_value` appear in the same order within `value`.
        Returns False otherwise.
    """

    # Remove decimal points by replacing them with an empty string
    value_str = str(value).replace('.', '')
    calculated_str = str(calculated_value).replace('.', '')

    # Sliding window approach to check for two consecutive matching digits
    for i in range(len(calculated_str) - 1):
        # Take two consecutive digits from the calculated value
        consecutive_digits = calculated_str[i:i+2]
        # Check if these two consecutive digits appear in the same order in the value
        if consecutive_digits in value_str:
            return True
    return False

# Save intermediate versions
def save_intermediate_version(workbook, stage, transient_transcription_output_dir_station, month_filename):
    '''
    Saves an intermediate version of the transcribed Excel workbook at a specified QA/QC stage.

    This function helps track changes and progress throughout the transcription and QA/QC process by saving 
    the Excel workbook at each defined stage. The intermediate file is named based on the stage identifier 
    and stored in the specified directory.

    Parameters
    --------------
    workbook : openpyxl.Workbook
        The Excel workbook object containing transcribed data to be saved.
    
    stage : str
        A label or identifier for the current QA/QC stage (e.g., 'decimal_place_correction', 'digit_manipulation').
    
    transient_transcription_output_dir_station : str
        Directory path for saving the intermediate file, typically associated with a specific station.
    
    month_filename : str
        Filename indicating the processed month and year, used for naming the intermediate file.

    Returns
    --------------
    intermediate_file : str
        The full path to the saved intermediate file.
    '''

    # To save intermediate versions of the trnascribed file throughout the different QA/QC steps 
    intermediate_file = os.path.join(transient_transcription_output_dir_station, f'{month_filename}_stage_{stage}.xlsx')
    workbook.save(intermediate_file)



def correct_80s_misreads(value):
    """
    If a value is in the 80s (80-89 range), replace the leading '8' with '3'.
    Works for both integers and decimals.
    """
    if isinstance(value, (int, float)):  # If it's a number, convert to string for checking
        value_str = str(value)
        if re.match(r"^8\d(\.\d+)?$", value_str):  # Match 80-89, including decimals
            corrected_value = "3" + value_str[1:]  # Replace the '8' with '3'
            return float(corrected_value)  # Convert back to float
    elif isinstance(value, str) and re.match(r"^8\d(\.\d+)?$", value.strip()):  
        return float("3" + value.strip()[1:])  # Ensure it converts to float after fixing
    return value  # Return original if no correction needed

def calculate_U(ea, delta_e):
    # Calculate relative humidity using the provided formula
    return (ea / (delta_e + ea)) * 100 if (delta_e + ea) != 0 else 0
    
def qa_qc(transcribed_table, station, transient_transcription_output_dir, post_QA_QC_transcribed_hydroclimate_data_dir_station, month_filename, max_temperature_threshold, min_temperature_threshold, decimal_places, uncertainty_margin, header_rows, multi_day_totals, multi_day_averages, max_days_for_multi_day_total, multi_day_totals_rows, final_totals_rows, excluded_rows, excluded_columns, daily_temperature_columns, daily_temperature_columns_and_diurnal_temperature_range, daily_precipitation_column, dry_and_wet_bulb_temperature_columns):
    '''
    Performs Quality Assessment and Quality Control (QA/QC) checks on transcribed hydroclimatic data.

    This function systematically verifies and corrects transcribed data by applying a series of checks, 
    recalculations, and adjustments based on predefined thresholds and logical rules. It handles validation 
    of individual cell values, multi-day totals and averages, as well as relationships between temperature 
    variables (e.g., max, min, average, and duirnal temperature range). It also highlights discrepancies and uncertainties, ensuring a clean, 
    reliable dataset for further use.

    Parameters
    --------------
    transcribed_table : str
        Path to the Excel file containing the initial transcribed data.
    station : str
        Identifier for the station (station no.) being processed, used for organizing outputs.
    transient_transcription_output_dir : str
        Directory for saving intermediate results during QA/QC steps.
    post_QA_QC_transcribed_hydroclimate_data_dir_station : str
        Directory for saving the final post-QA/QC dataset.
    month_filename : str
        Base name for the output files, typically including station metadata and the month/year of the data.
    max_temperature_threshold : float
        Upper limit for valid temperature values, used to flag potential errors.
    min_temperature_threshold : float
        Lower limit for valid temperature values, used to flag potential errors.
    decimal_places : int
        Number of decimal places to adjust the transcribed values. Values are divided by `10**decimal_places`.
    uncertainty_margin : float
        Allowed margin of error for numerical comparisons (e.g., in totals, averages).
    header_rows : int
        Number of header rows in the dataset to exclude from numerical checks.
    multi_day_totals : bool
        Whether the dataset includes multi-day total rows for validation.
    multi_day_averages : bool
        Whether the dataset includes multi-day average rows for validation.
    max_days_for_multi_day_total : int
        Maximum number of days included in a multi-day total calculation.
    multi_day_totals_rows : list of int
        Row indices of the multi-day total rows in the dataset.
    final_totals_rows : list of int
        Row indices of the final total rows in the dataset.
    excluded_rows : list of int
        Row indices to exclude from QA/QC checks.
    excluded_columns : list of int
        Column indices to exclude from QA/QC checks.
    columns_to_check : list of str
        List of column letters to check for temperature variables (e.g., max, min, average).
    columns_to_check_with_extra_variable : list of str
        List of column letters to check for temperature variables, including additional variables like amplitude (duirnal temperature range).

    Returns
    --------------
    new_workbook : openpyxl.Workbook
        The post-QA/QC Excel workbook object containing the cleaned and verified data.

    '''

    # Open directory with pre_QA_QC_transcribed data excel file
    workbook = openpyxl.load_workbook(transcribed_table)
    worksheet = workbook.active

    # Path to save a new copy of the workbook for post-processing
    new_version_of_file = os.path.join(post_QA_QC_transcribed_hydroclimate_data_dir_station, f'{month_filename}_post_QA_QC.xlsx')
    os.makedirs(post_QA_QC_transcribed_hydroclimate_data_dir_station, exist_ok=True)

    # Save the original workbook to ensure it's on disk
    transient_transcription_output_dir_station = os.path.join(transient_transcription_output_dir, station)
    os.makedirs(transient_transcription_output_dir_station, exist_ok=True)

    original_path = os.path.join(transient_transcription_output_dir_station, 'original_transcribed_data.xlsx')
    workbook.save(original_path)

    # Copy the original file to a new version for post-processing
    shutil.copy2(original_path, new_version_of_file)

    new_workbook = openpyxl.load_workbook(new_version_of_file)
    new_worksheet = new_workbook.active # To select the first worksheet of the workbook without requiring its name

    workbook.close() # Close original transcribed workbook

    # Process cells while excluding specified rows and columns
    # Inorder to already avoid very large incorrectly transcribed values right from the start, here we edit values that have more than 4 digits (thousands) in certain rows where we know it is impossible. This is strictly for our case study and may not apply to your case study. If not, comment the lines below until the '#*'
    # Iterate over all rows in the worksheet, with exception of the excluded rows (Headers and 5 day totals) and columns (e.g., those with U (relative humidity))
    for row in new_worksheet.iter_rows(min_col=0, max_col=new_worksheet.max_column):
        # Get the row index from the first cell of the row
        if row[0].row not in excluded_rows:
            for cell in row:
                if cell.column not in excluded_columns:
                    if is_string_convertible_to_float(cell.value): # Checking if the value transcribed in the cell is convertible to a float to avoid strings 
                        # Convert to string to check length and manipulate
                        str_value = str(cell.value)
                        if len(str_value) > 4: # Thousands. Here i chose 4 because it seems the strings are made up up numbers and a newline \n character. Check this when you use 3 instead
                            # Remove the first digit and convert back to the original type
                            new_value = str_value[1:]  # Remove the first character
                            cell.value = type(cell.value)(new_value)  # Convert back to int or float
                            highlight_change('FF9933', cell, new_version_of_file) #FF9933 is 2 ## To highlight manipulation of transcribed data
                            # Save Excel file
                            new_workbook.save(new_version_of_file)
    
    # Save Excel file
    new_workbook.save(new_version_of_file)
    # Save intermediate version after the first loop
    save_intermediate_version(new_workbook, "digit_manipulation", transient_transcription_output_dir_station, month_filename)
    #*



    # Since values transcribed ignore decimal points for all cells yet all the cell values initially are in decimals, here we Iterate over all rows, starting from row 4 to skip header rows
    # Note: we ignored the decimal points in the transcription module to reduce the noise in the transcripbed values since the sheets have dotted lines that would be recoognized by the OCR/HTR model as multiple decimal points.
    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row, min_col=3, max_col=new_worksheet.max_column-1): # For columns: avoid the first two columns and the last column (2nd with Date, and last also with the Date)
        for cell in row:
            cell.value = clean_numeric_string(cell.value)
            if is_string_convertible_to_float(cell.value): # Checking if the value transcribed in the cell is convertible to a float to avoid strings 
                cell.value = (float(cell.value))/(10**decimal_places) # This is set up in the configuration file.Here we had one decimal place through the entire sheet, so  we divide the cell value by 10
    new_workbook.save(new_version_of_file) # Save the modified workbook
    # Save intermediate version after the second loop
    save_intermediate_version(new_workbook, "decimal_place_correction", transient_transcription_output_dir_station, month_filename)

    
    # Cell coordinates to check the totals
    # Rows
    if multi_day_totals:
        rows_to_process = multi_day_totals_rows
    else:
        rows_to_process = final_totals_rows

    for current_index, row in enumerate(rows_to_process):

        for column in daily_temperature_columns + daily_precipitation_column + dry_and_wet_bulb_temperature_columns: # Where these columns to check represent [ Max Temperature, Min Temperature, Average Temperature] 
            # Get the value of the multi day total (in our case '5/6-day Total') that was transcribed. Note Months with 31 days ahve some 6 day todays considering 26th to 31st
            
            # Get the value of the multi-day total for the current row
            cell_for_total = new_worksheet[f"{column}{row}"]
            total_value = cell_for_total.value

            if total_value is None or total_value == '':
                highlight_change('FFC0CB', cell_for_total, new_version_of_file)  # Highlight empty cells in pink
                continue
            
            # Dynamically determine the number of offset cells (days considered in calculation of multi-day totals) based on row gaps, header rows, and presence of multi-day averages (placed immediatelly after the totals)
            if current_index == 0:
                # For the first row with multi-day totals, calculate offset_cells by subtracting the header rows. We also subtract by 1 here because the row index start at 0
                offset_cells = row - header_rows - 1
            else:
                # For other rows, calculate offset based on the gap between rows
                if multi_day_averages:
                    multi_day_average_row = 1
                else:
                    multi_day_average_row = 0

                offset_cells = multi_day_totals_rows[current_index] - multi_day_totals_rows[current_index - 1] - multi_day_average_row - 1
            
            # Ensure the offset cells do not exceed max_offset_days
            offset_cells = min(offset_cells, max_days_for_multi_day_total)

            # Calculate cell values for the days leading up to the multi-day totals
            cell_values_for_days = []
            blank_cells = []
            
            for day in range(offset_cells):
                cell_coordinate = new_worksheet[f"{column}{row - day - 1}"]
                cell_value = cell_coordinate.value
                if cell_value is None or cell_value == '':
                    cell_value = 0.0  # Replace empty values with 0.0 for summation. This is to avoid errors in Excel summation.
                    cell_values_for_days.append(cell_value)
                    highlight_change('FFFFFF', cell_coordinate, new_version_of_file) #No Fill 
                    blank_cells.append(1.0)
                else:
                    if is_string_convertible_to_float(cell_value):
                        cell_value = float(cell_value)
                        if column in daily_temperature_columns + dry_and_wet_bulb_temperature_columns:
                            if cell_value >= 100.0: 
                                manipulated_value = cell_value/10.0 # Special manipulation. If transcribed temperature value > 100, divide by 10.
                                cell_value = manipulated_value
                                highlight_change('CC3300', cell_coordinate, new_version_of_file) #CC3300 is Dark Red. Highlight to show that this special manupilation was done on this cell.
                            if cell_value < min_temperature_threshold or cell_value > max_temperature_threshold:
                                highlight_change('CC3300', cell_coordinate, new_version_of_file) #CC3300 is Dark Red. Highlight to show that temperature exceeds set thresholds for maximum or minimum 
                        cell_values_for_days.append(cell_value)
                        cell_coordinate.value = cell_value
                        new_workbook.save(new_version_of_file)
                    else:
                        cell_values_for_days.append(0.0) # Debugging: Incase the value transcribed in the cell is a string, convert it to 0.0 to avoid errors in the summation.


            # Sum the transcribed daily values and compare them with the transcribed multi-day totals
            total_sum = sum(cell_values_for_days)
            if abs(total_sum - float(total_value)) <= uncertainty_margin: 
                highlight_change('FF6DCD57', cell_for_total, new_version_of_file)  #FF6DCD57 is Green. Highlight to show that the transcribed multi-day total values confirmed that the transcribed daily values are correct, and vice versa
                
                # Additionally highlight the transcribed daily values that lead to correct transcribed multi-day totals
                for i in range(offset_cells):
                    cell_to_highlight = new_worksheet[f"{column}{row - i - 1}"]  # Access the previous offset cells. These are the cells above the multi-day total
                    if cell_to_highlight.value is not None or cell_to_highlight.value != '':
                        highlight_change('FF6DCD57', cell_to_highlight, new_version_of_file)  # FF6DCD57 is Green. Highlight to confirm these day values are part of the correct multi-day total calculation.
                new_workbook.save(new_version_of_file)
            else:
                highlight_change('75696F', cell_for_total, new_version_of_file)  #75696F is Grey. Highlight to show that transcribed multi-day total values are not equal to the transcribed daily values, and hence we cant confirm if the transcription was correct.
                new_workbook.save(new_version_of_file)
            
            if multi_day_averages and column in daily_temperature_columns + dry_and_wet_bulb_temperature_columns:
                cell_for_average = new_worksheet[f"{column}{row + 1}"]
                average_value = cell_for_average.value

                if average_value is None or average_value == '':
                    highlight_change('FFC0CB', cell_for_average, new_version_of_file)  #FFC0CB is pink.  Highlight empty cells in pink to show that the average value was not transcribed (or is missing)
                    continue
                # Calculate the average from the previous offset cells
                total_blank_cells = sum(blank_cells) # Count of blank cells. This is to avoid calculating the multi-day average whist considering missing days. i.e. for 28-30 day months, average in the last pentad isn't calculated using the same offset cells as that for 31 day months
                days = float(offset_cells - total_blank_cells) # Days considered for multi-day average calculation
                
                if average_value is not None and is_string_convertible_to_float(average_value) and total_sum is not None and total_sum != '' and days != 0:

                    calculated_average = total_sum / days
                    if abs(calculated_average - float(average_value)) <= uncertainty_margin:
                        highlight_change('FF6DCD57', cell_for_average, new_version_of_file)  # FF6DCD57 is Green. Highlight to show that the transcribed multi-day average values confirmed that the transcribed daily values are correct, and vice versa
                        
                        # Additionally highlight the transcribed daily values that lead to correct transcribed multi-day average
                        for i in range(offset_cells):
                            cell_to_highlight = new_worksheet[f"{column}{row - i - 1}"]  # Access the previous offset cells. These are the cells above the multi-day total
                            highlight_change('FF6DCD57', cell_to_highlight, new_version_of_file)  # FF6DCD57 is Green. Highlight to confirm these day values are part of the correct multi-day average calculation.

                        new_workbook.save(new_version_of_file)
                    else:
                        highlight_change('75696F', cell_for_average, new_version_of_file)  # #75696F is Grey. Highlight to show that transcribed multi-day total average value is not equal to the transcribed daily values, and hence we cant confirm if the transcription was correct.
                        new_workbook.save(new_version_of_file)                


    # Save intermediate version after the third loop
    save_intermediate_version(new_workbook, "multi_day_totals_and_avgs", transient_transcription_output_dir_station, month_filename)
        

    # Check row by row. Here we focus on the Daily Minimum, Average and Maximum Temperatures. Here Minimum temp. < Average temp. < Maximum temp.
    # Also: If two of the three temperature values (min, max or average) are correct (i.e. checked by the multi-day averages or total, and thus highlighted with green) and therefore correct the third value
    # We check the cells by rows, and manipulate where necessary
    
    # First, Convert column letters to numbers
    columns_to_check_indices = [openpyxl.utils.column_index_from_string(col) for col in daily_temperature_columns]

    # Get min_col and max_col from the columns_to_check
    min_col = min(columns_to_check_indices)
    max_col = max(columns_to_check_indices)

    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row, min_col=min_col, max_col=max_col):     
            
        # Create a dictionary to map column names (D, E, F, etc.) to the cell values
        row_cells = {daily_temperature_columns[i]: row[columns_to_check_indices[i] - min_col] for i in range(len(daily_temperature_columns))}

        # Now you can access the cells dynamically using the column names
        D = row_cells.get('D')  # Max Temp
        E = row_cells.get('E')  # Min Temp
        F = row_cells.get('F')  # Average Temp

        D_highlighted = is_highlighted(D, 'FF6DCD57') # Daily maximum temperature
        E_highlighted = is_highlighted(E, 'FF6DCD57') # Daily minimum temperature
        F_highlighted = is_highlighted(F, 'FF6DCD57') # Daily average temperature

        highlighted_count = D_highlighted + E_highlighted + F_highlighted  # Check how many of these daily temperature values (max, min or ave) are confirmed 'GREEN' from the prior checks)

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        exists_count = D_value_exists + E_value_exists + F_value_exists

        if highlighted_count == 2:
            if not D_highlighted:
                if is_string_convertible_to_float(E.value) and is_string_convertible_to_float(F.value):
                    if E.value > F.value:  # If Min Temp > Average Temp
                        # Highlight in red since Min Temp is greater than Average Temp (error condition)
                        highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                        highlight_change('FF0000', E, new_version_of_file)
                        highlight_change('FF0000', F, new_version_of_file)
                    else:
                        # Calculate D (Max Temp) based on valid F (Average Temp) and E (Min Temp) values
                        calculated_D = round(2 * float(F.value) - float(E.value), 1)
                        if D_value_exists and abs(float(D.value) - calculated_D) <= uncertainty_margin:
                            highlight_change('FF6DCD57', D, new_version_of_file)  # Highlight in green for correct value
                        else:
                            D.value = calculated_D
                            highlight_change('FF6DCD57', D, new_version_of_file)  # Highlight in green after correction
            elif not E_highlighted:
                if is_string_convertible_to_float(D.value) and is_string_convertible_to_float(F.value):
                    if D.value < F.value: # If Max Temp < Average Temp
                        # Highlight in red since Max Temp is less than Average Temp (error condition)
                        highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                        highlight_change('FF0000', E, new_version_of_file)
                        highlight_change('FF0000', F, new_version_of_file)
                    else:
                        # Calculate E (Min Temp) based on valid D (Max Temp) and F (Average) values
                        calculated_E = round(2 * float(F.value) - float(D.value), 1)
                        if E_value_exists and abs(float(E.value) - calculated_E) <= uncertainty_margin:
                            highlight_change('FF6DCD57', E, new_version_of_file)  # Highlight in green for correct value
                        else:
                            E.value = calculated_E
                            highlight_change('FF6DCD57', E, new_version_of_file)  # Highlight in green after correction
            elif not F_highlighted:
                if is_string_convertible_to_float(D.value) and is_string_convertible_to_float(E.value):
                    if D.value < E.value: # If Max Temp < Min Temp
                        # Highlight in red since Max Temp is less than Min Temp (error condition)
                        highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                        highlight_change('FF0000', E, new_version_of_file)
                        highlight_change('FF0000', F, new_version_of_file)
                    else:
                        # Calculate F (Average temp) based on valid D (Max Temp) and E (Min Temp) values
                        calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
                        if F_value_exists and abs(float(F.value) - calculated_F) <= uncertainty_margin:
                            highlight_change('FF6DCD57', F, new_version_of_file)  # Highlight in green for correct value
                        else:
                            F.value = calculated_F
                            highlight_change('FF6DCD57', F, new_version_of_file)  # Highlight in green after correction
            #new_workbook.save(new_version_of_file)
            
        # If one cell is highlighted but has no value, calculate its value
        elif highlighted_count == 3 and exists_count == 2:
            if D_highlighted and not D_value_exists and E_value_exists and F_value_exists:
                if E.value > F.value:  # If Min Temp > Average Temp
                    # Highlight in red since Min Temp is greater than Average Temp (error condition)
                    highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                    highlight_change('FF0000', E, new_version_of_file)
                    highlight_change('FF0000', F, new_version_of_file)
                else:
                    D.value = round(2 * float(F.value) - float(E.value), 1)
                    highlight_change('FF6DCD57', D, new_version_of_file) # Highlight in green for correct value
            elif E_highlighted and not E_value_exists and D_value_exists and F_value_exists:
                if D.value < F.value: # If Max Temp < Average Temp
                    # Highlight in red since Max Temp is less than Average Temp (error condition)
                    highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                    highlight_change('FF0000', E, new_version_of_file)
                    highlight_change('FF0000', F, new_version_of_file)
                else:
                    E.value = round(2 * float(F.value) - float(D.value), 1)
                    highlight_change('FF6DCD57', E, new_version_of_file) # Highlight in green for correct value
            elif F_highlighted and not F_value_exists and D_value_exists and E_value_exists:
                if D.value < E.value: # If Max Temp < Min Temp
                    # Highlight in red since Max Temp is less than Min Temp (error condition)
                    highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                    highlight_change('FF0000', E, new_version_of_file)
                    highlight_change('FF0000', F, new_version_of_file)
                else:
                    F.value = round((float(D.value) + float(E.value)) / 2, 1)
                    highlight_change('FF6DCD57', F, new_version_of_file) # Highlight in green for correct value
            #new_workbook.save(new_version_of_file)
        
        # If none of the cells are highlighted green but they all have values, or maybe only one of the cells is highlighted green
        elif highlighted_count < 2 and D_value_exists and E_value_exists and F_value_exists:

            calculated_D = round(2 * float(F.value) - float(E.value), 1)
            if abs(calculated_D - float(D.value)) <= uncertainty_margin:
                if E.value > F.value:  # If Min Temp > Average Temp
                    # Highlight in red since Min Temp is greater than Average Temp (error condition)
                    highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                    highlight_change('FF0000', E, new_version_of_file)
                    highlight_change('FF0000', F, new_version_of_file)
                else:
                    highlight_change('FF6DCD57', D, new_version_of_file) # Highlight in green for correct value
                #new_workbook.save(new_version_of_file)
            
            calculated_E = round(2 * float(F.value) - float(D.value), 1)
            if abs(calculated_E - float(E.value)) <= uncertainty_margin:
                if D.value < F.value: # If Max Temp < Average Temp
                    # Highlight in red since Max Temp is less than Average Temp (error condition)
                    highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                    highlight_change('FF0000', E, new_version_of_file)
                    highlight_change('FF0000', F, new_version_of_file)
                else:
                    highlight_change('FF6DCD57', E, new_version_of_file) # Highlight in green for correct value
                #new_workbook.save(new_version_of_file)
            

            calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
            if abs(calculated_F - float(F.value)) <= uncertainty_margin:
                if D.value < E.value: # If Max Temp < Min Temp
                    # Highlight in red since Max Temp is less than Min Temp (error condition)
                    highlight_change('FF0000', D, new_version_of_file)  # FF0000 is Red
                    highlight_change('FF0000', E, new_version_of_file)
                    highlight_change('FF0000', F, new_version_of_file)
                else:
                    highlight_change('FF6DCD57', F, new_version_of_file) # Highlight in green for correct value
                #new_workbook.save(new_version_of_file)
            


            # Check if at least two digits of the calculated value are present in the correct order in the given value. If so, replace the transcribed value with the corrected value
            if has_two_digits_in_order(D.value, calculated_D) and min_temperature_threshold <= calculated_D <= max_temperature_threshold and calculated_D > F.value > E.value:
                D.value = calculated_D
                highlight_change('FF6DCD57', D, new_version_of_file)
                # new_workbook.save(new_version_of_file)

            # Check if at least two digits of the calculated value are present in the correct order in the given value. If so, replace the transcribed value with the correcetd value
            if has_two_digits_in_order(E.value, calculated_E) and min_temperature_threshold <= calculated_E <= max_temperature_threshold and calculated_E < F.value < D.value :
                E.value = calculated_E
                highlight_change('FF6DCD57', E, new_version_of_file)
                # new_workbook.save(new_version_of_file)

            # Check if at least two digits of the calculated value are present in the correct order in the given value. If so, replace the transcribed value with the correcetd value
            if has_two_digits_in_order(F.value, calculated_F) and min_temperature_threshold <= calculated_F <= max_temperature_threshold and E.value < calculated_F < D.value:
                F.value = calculated_F
                highlight_change('FF6DCD57', F, new_version_of_file)
                # new_workbook.save(new_version_of_file)
                
        #new_workbook.save(new_version_of_file)    

        # RE-CHECKS
        D_highlighted = is_highlighted(D, 'FF6DCD57')
        E_highlighted = is_highlighted(E, 'FF6DCD57')
        F_highlighted = is_highlighted(F, 'FF6DCD57')

        highlighted_count = D_highlighted + E_highlighted + F_highlighted

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        if highlighted_count >=1 :
            if E_value_exists and F_value_exists and min_temperature_threshold <= E.value <= max_temperature_threshold and min_temperature_threshold <= F.value <= max_temperature_threshold and F.value > E.value:
                calculated_D = round(2 * float(F.value) - float(E.value), 1)
                if D.value is not None and D_value_exists and abs(float(D.value) - calculated_D) <= uncertainty_margin:
                    highlight_change('FF6DCD57', D, new_version_of_file) # Highlight in green for correct value
                else:
                    if not D_highlighted:
                        D.value = calculated_D
                        highlight_change('FF6DCD57', D, new_version_of_file) # Highlight in green for correct value
            if D_value_exists and F_value_exists and min_temperature_threshold <= D.value <= max_temperature_threshold and min_temperature_threshold <= F.value <= max_temperature_threshold and D.value > F.value:
                calculated_E = round(2 * float(F.value) - float(D.value), 1)
                if E.value is not None and E_value_exists and abs(float(E.value) - calculated_E) <= uncertainty_margin:
                    highlight_change('FF6DCD57', E, new_version_of_file) # Highlight in green for correct value
                else:
                    if not E_highlighted:
                        E.value = calculated_E
                        highlight_change('FF6DCD57', E, new_version_of_file) # Highlight in green for correct value
            if D_value_exists and E_value_exists and min_temperature_threshold <= D.value <= max_temperature_threshold and min_temperature_threshold <= E.value <= max_temperature_threshold and D.value > E.value:
                calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
                if F.value is not None and F_value_exists and abs(float(F.value) - calculated_F) <= uncertainty_margin:    
                    highlight_change('FF6DCD57', F, new_version_of_file) # Highlight in green for correct value
                else:
                    if not F_highlighted:
                        F.value = calculated_F
                        highlight_change('FF6DCD57', F, new_version_of_file) # Highlight in green for correct value
        
        if D_value_exists and E_value_exists and F_value_exists and min_temperature_threshold <= D.value <= max_temperature_threshold and min_temperature_threshold <= E.value <= max_temperature_threshold and min_temperature_threshold <= F.value <= max_temperature_threshold and E.value < F.value < D.value:
            # Calculate the average of D and E and check if it is within the uncertainty margin of F
            calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
            if abs(calculated_F - float(F.value)) <= uncertainty_margin:
                # Highlight all cells green if the condition is true
                highlight_change('FF6DCD57', D, new_version_of_file)
                highlight_change('FF6DCD57', E, new_version_of_file)
                highlight_change('FF6DCD57', F, new_version_of_file)


    new_workbook.save(new_version_of_file)

    # Save intermediate version after the checking the transcribed daily values of maximum, minimum amd average.
    save_intermediate_version(new_workbook, "daily_max_min_avg_temp", transient_transcription_output_dir_station, month_filename)
    

    # Check the cells by rows to ensure that Min temp threshold < Temp < Max Temp threshold
    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row, min_col=min_col, max_col=max_col):
        if row[0].row in excluded_rows:
            continue 
        
        # Create a dictionary to map column names (D, E, F, etc.) to the cell values
        row_cells = {daily_temperature_columns[i]: row[columns_to_check_indices[i] - min_col] for i in range(len(daily_temperature_columns))}

        # Now you can access the cells dynamically using the column names
        D = row_cells.get('D')  # Max Temp
        E = row_cells.get('E')  # Min Temp
        F = row_cells.get('F')  # Average Temp

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        if D_value_exists:
            corrected_D_value = correct_80s_misreads(D.value)
            D.value = corrected_D_value  # Update cell value if corrected
            D_value = float(D.value)
            if not (min_temperature_threshold <= D_value <= max_temperature_threshold): 
                highlight_change('CC3300', D, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct

        if E_value_exists:
            corrected_E_value = correct_80s_misreads(E.value)
            E.value = corrected_E_value  # Update cell value if corrected
            E_value = float(E.value)
            if not (min_temperature_threshold <= E_value <= max_temperature_threshold):
                highlight_change('CC3300', E, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct
        
        if F_value_exists:
            corrected_F_value = correct_80s_misreads(F.value)
            F.value = corrected_F_value  # Update cell value if corrected
            F_value = float(F.value)
            if not (min_temperature_threshold <= F_value <= max_temperature_threshold):
                highlight_change('CC3300', F, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct

    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)
    # Save intermediate version after rechecking the temperature thresholds for the transcribed daily values of maximum, minimum amd average.
    save_intermediate_version(new_workbook, "correct_values", transient_transcription_output_dir_station, month_filename)
    





    # Check of Max, Min and Avg Temperatures using the Amplitude (Ampl.) commonly known as the Diurnal Temperarure Range. This is column G of our worksheets
    # where: Ampl. = Max - Min ..................... (1)
    #        Ampl. = 2Avg - 2Min ................... (2)
    #        Ampl. = 2Max - 2Avg ................... (3)
    # The check the cells by rows, and manipulate where necessary

    #Adding the extra variable column i.e. Ampl. (Diurnal Temperature Range)
    # First, Convert column letters to numbers
    columns_to_check_indices_with_extra_variable = [openpyxl.utils.column_index_from_string(col) for col in daily_temperature_columns_and_diurnal_temperature_range]

    # Get min_col and max_col from the columns_to_check
    min_col_with_extra_variable = min(columns_to_check_indices_with_extra_variable)
    max_col_with_extra_variable = max(columns_to_check_indices_with_extra_variable)

    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row, min_col=min_col_with_extra_variable, max_col=max_col_with_extra_variable):    
            
        # Create a dictionary to map column names (D, E, F, etc.) to the cell values
        row_cells = {daily_temperature_columns_and_diurnal_temperature_range[i]: row[columns_to_check_indices_with_extra_variable[i] - min_col] for i in range(len(daily_temperature_columns_and_diurnal_temperature_range))}

        # Now you can access the cells dynamically using the column names
        D = row_cells.get('D')  # Max Temp
        E = row_cells.get('E')  # Min Temp
        F = row_cells.get('F')  # Average Temp
        G = row_cells.get('G')  # Diurnal Temp range

        D_highlighted = is_highlighted(D, 'FF6DCD57')
        E_highlighted = is_highlighted(E, 'FF6DCD57')
        F_highlighted = is_highlighted(F, 'FF6DCD57')

        # highlighted_count = D_highlighted + E_highlighted + F_highlighted

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)
        G_value_exists = is_string_convertible_to_float(G.value)

        if G_value_exists:
            G_value = float(G.value) # Here G_value is the Amplitude
            if G_value >+ 100:
                G_value /= 10 # Special manipulation. If transcribed duirnal temperature value > 100, divide by 10.
                highlight_change('CC3300', G, new_version_of_file) #CC3300 is Dark Red.  
            G.value = round(G_value, 1)
        
            # Calculate the amplitude based on the given formulas
            if D_value_exists and E_value_exists:
                D_value = float(D.value) # Max Temp
                E_value = float(E.value) # Min Temp

                if D_value > E_value: # Max temp > Min Temp
                    Ampl_1 = D_value - E_value  # Ampl. = Max - Min ..................... (1)
                    if abs(G_value - Ampl_1) <= uncertainty_margin:
                        # Highlight Ampl. cell green if the condition is true
                        highlight_change('FF6DCD57', G, new_version_of_file)
                        if not F_highlighted:
                            calculated_F = round((float(D.value) + float(E.value)) / 2, 1)
                            F.value = calculated_F
                            highlight_change('FF6DCD57', D, new_version_of_file) # Highlight in green for correct value
                            highlight_change('FF6DCD57', E, new_version_of_file) 
                            highlight_change('FF6DCD57', F, new_version_of_file)
                
            if E_value_exists and F_value_exists:
                E_value = float(E.value) # Min Temp
                F_value = float(F.value) # Avg Temp

                if E_value < F_value: # Min Temp < Avg temp
                    Ampl_2 = (2*F_value) - (2*E_value) # Ampl. = 2Avg - 2Min ................... (2)
                    if abs(G_value -Ampl_2) <= uncertainty_margin:
                        # Highlight Ampl. cell green if the condition is true
                        highlight_change('FF6DCD57', G, new_version_of_file)
                        if not D_highlighted:
                            calculated_D = round(2 * float(F.value) - float(E.value), 1)
                            D.value = calculated_D
                            highlight_change('FF6DCD57', D, new_version_of_file) # Highlight in green for correct value
                            highlight_change('FF6DCD57', E, new_version_of_file)
                            highlight_change('FF6DCD57', F, new_version_of_file)
                
            if D_value_exists and F_value_exists:
                D_value = float(D.value) # Max Temp
                F_value = float(F.value) # Avg Temp

                if D_value > F_value: # Max temp > Avg temp
                    Ampl_3 = (2*D_value) - (2*F_value) # Ampl. = 2Max - 2Avg ................... (3)
                    if abs(G_value -Ampl_3) <= uncertainty_margin:
                        # Highlight Ampl. cell green if the condition is true
                        highlight_change('FF6DCD57', G, new_version_of_file)
                        if not E_highlighted:
                            calculated_E = round(2 * float(F.value) - float(D.value), 1)
                            E.value = calculated_E
                            highlight_change('FF6DCD57', D, new_version_of_file) # Highlight in green for correct value
                            highlight_change('FF6DCD57', E, new_version_of_file)
                            highlight_change('FF6DCD57', F, new_version_of_file)

    new_workbook.save(new_version_of_file)
    # Save intermediate version after the checking the relation of duirnal temperature range with the transcribed daily values of maximum, minimum amd average.
    save_intermediate_version(new_workbook, "duirnal_temp_range", transient_transcription_output_dir_station, month_filename)
    

    ## RE-CHECKS
    # Check the cells by rows to ensure that Min temp threshold < Temp < Max Temp threshold
    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row, min_col=min_col, max_col=max_col):
        if row[0].row in excluded_rows:
            continue 
        
        # Create a dictionary to map column names (D, E, F, etc.) to the cell values
        row_cells = {daily_temperature_columns[i]: row[columns_to_check_indices[i] - min_col] for i in range(len(daily_temperature_columns))}

        # Now you can access the cells dynamically using the column names
        D = row_cells.get('D')  # Max Temp
        E = row_cells.get('E')  # Min Temp
        F = row_cells.get('F')  # Average Temp

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        if D_value_exists:
            corrected_D_value = correct_80s_misreads(D.value)
            D.value = corrected_D_value  # Update cell value if corrected
            D_value = float(D.value)
            if not (min_temperature_threshold <= D_value <= max_temperature_threshold): 
                highlight_change('CC3300', D, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct

        if E_value_exists:
            corrected_E_value = correct_80s_misreads(E.value)
            E.value = corrected_E_value  # Update cell value if corrected
            E_value = float(E.value)
            if not (min_temperature_threshold <= E_value <= max_temperature_threshold):
                highlight_change('CC3300', E, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct
        
        if F_value_exists:
            corrected_F_value = correct_80s_misreads(F.value)
            F.value = corrected_F_value  # Update cell value if corrected
            F_value = float(F.value)
            if not (min_temperature_threshold <= F_value <= max_temperature_threshold):
                highlight_change('CC3300', F, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct

    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)
    # Save intermediate version after rechecking the temperature thresholds for the transcribed daily values of maximum, minimum amd average.
    save_intermediate_version(new_workbook, "recheckthresh", transient_transcription_output_dir_station, month_filename)
    

    ## Additional checks for Dry bulb temperature (T), Wet bulb temperature (T'a), Actual Vapour pressure (e), Relative Humidity (U), and delta e 
    # Iterate over each row in the worksheet after header rows
    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row):
        if row[0].row in excluded_rows:
            continue 
        
        for i in range(0, len(dry_and_wet_bulb_temperature_columns), 2):
            dry_col = dry_and_wet_bulb_temperature_columns[i]
            wet_col = dry_and_wet_bulb_temperature_columns[i+1]

            ea_col = openpyxl.utils.column_index_from_string(wet_col) + 1
            U_col = ea_col + 1
            delta_e_col = U_col + 1

            dry_cell = row[openpyxl.utils.column_index_from_string(dry_col) - 1]
            wet_cell = row[openpyxl.utils.column_index_from_string(wet_col) - 1]
            ea_cell = row[ea_col - 1]
            U_cell = row[U_col - 1]
            delta_e_cell = row[delta_e_col - 1]

            if all(is_string_convertible_to_float(cell.value) for cell in [dry_cell, wet_cell, ea_cell, U_cell, delta_e_cell]):
                T = float(dry_cell.value)
                Ta = float(wet_cell.value)
                ea = float(ea_cell.value)
                delta_e = float(delta_e_cell.value)
                U = float(U_cell.value)
                calculated_U = calculate_U(ea, delta_e)
                
                # print("the U is: ")
                # print(U)
                # print("the calculated U is: ")
                # print(calculated_U)
                # print("Checking if the Actual Vapour pressure (e), Relative Humidity (U), and delta e calculation is accurate")
                # Update U_cell if different from calculated
                if np.isclose(U, calculated_U, atol=uncertainty_margin):
                    # print("The Actual Vapour pressure (e), Relative Humidity (U), and delta e calculation is accurate")
                    highlight_change('FF6DCD57', U_cell, new_version_of_file)  # Confirmed value. Highlight in green
                    highlight_change('FF6DCD57', ea_cell, new_version_of_file) # Confirmed value. Highlight in green
                    highlight_change('FF6DCD57', delta_e_cell, new_version_of_file)    # Confirmed value. Highlight in green

                # Temperature adjustments based on confirmed ea and delta_e
                if all(is_highlighted(cell, 'FF6DCD57') for cell in [ea_cell, U_cell, delta_e_cell]):
                    expected_es = ea + delta_e
                    # Reverse the es calculation to find T
                    calculated_T = (243.5 * np.log(expected_es / 6.112)) / (17.67 - np.log(expected_es / 6.112))

                    if np.abs(calculated_T - T) <= uncertainty_margin:
                        highlight_change('FF6DCD57', dry_cell, new_version_of_file) # Confirmed value. Highlight in green
                    else:
                        dry_cell.value = round(calculated_T, decimal_places)
                        highlight_change('FF6DCD57', dry_cell, new_version_of_file) # Confirmed value. Highlight in green


                    # Also check T'a and confirm or correct
                    # Reverse the ea calculation to find T
                    calculated_Ta = (243.5 * np.log(ea / 6.112)) / (17.67 - np.log(ea / 6.112))

                    if np.abs(calculated_Ta - Ta) <= uncertainty_margin:
                        highlight_change('FF6DCD57', wet_cell, new_version_of_file) # Confirmed value. Highlight in green
                    else:
                        wet_cell.value = round(calculated_Ta, decimal_places)
                        highlight_change('FF6DCD57', wet_cell, new_version_of_file) # Confirmed value. Highlight in green
                
            # If both T (dry) and T'a (wet) are confirmed, calculate ea, delta_e, and U
            if all(is_highlighted(cell, 'FF6DCD57') for cell in [dry_cell, wet_cell]) and all(is_string_convertible_to_float(cell.value) for cell in [dry_cell, wet_cell]):

                T = float(dry_cell.value)
                Ta = float(wet_cell.value)

                if not is_highlighted(ea_cell, 'FF6DCD57') or not is_highlighted(delta_e_cell, 'FF6DCD57') or not is_highlighted(U_cell, 'FF6DCD57'):
                    # Calculate actual vapour pressure (ea) from Ta (wet bulb)
                    calculated_ea = 6.112 * np.exp((17.67 * Ta) / (Ta + 243.5))
                    # Update and highlight
                    ea_cell.value = round(calculated_ea, decimal_places)
                    highlight_change('FF6DCD57', ea_cell, new_version_of_file) # Confirmed value. Highlight in green

                
                    # Calculate saturation vapour pressure from T (dry bulb)
                    es = 6.112 * np.exp((17.67 * T) / (T + 243.5))

                    # delta_e = es - ea
                    calculated_delta_e = es - calculated_ea

                    # U = ea / delta_e
                    calculated_U = calculate_U(calculated_ea, calculated_delta_e)

                    # Update and highlight
                    delta_e_cell.value = round(calculated_delta_e, decimal_places)
                    U_cell.value = round(calculated_U, decimal_places)

                    highlight_change('FF6DCD57', delta_e_cell, new_version_of_file) # Confirmed value. Highlight in green
                    highlight_change('FF6DCD57', U_cell, new_version_of_file) # Confirmed value. Highlight in green

            
            # If T is confirmed (green) while T'a is not confirmed and Δe is present and numeric, try calculating ea and validating T'a
            if is_highlighted(dry_cell, 'FF6DCD57') and not is_highlighted(wet_cell, 'FF6DCD57') and is_string_convertible_to_float(dry_cell.value) and is_string_convertible_to_float(delta_e_cell.value):
                T = float(dry_cell.value)
                delta_e = float(delta_e_cell.value)

                # Calculate saturation vapour pressure (es) from T
                es = 6.112 * np.exp((17.67 * T) / (T + 243.5))

                # Calculate ea = es - delta_e
                calculated_ea = es - delta_e

                # Now calculate T'a from ea
                #calculated_Ta = (243.5 * np.log(calculated_ea / 6.112)) / (17.67 - np.log(calculated_ea / 6.112))
                ratio = calculated_ea / 6.112
                if ratio > 0:
                    calculated_Ta = (243.5 * np.log(ratio)) / (17.67 - np.log(ratio))

                    if is_string_convertible_to_float(wet_cell.value):
                        Ta = float(wet_cell.value)
                        if np.abs(calculated_Ta - Ta) <= uncertainty_margin:
                            # All checks confirmed: update ea and U
                            ea_cell.value = round(calculated_ea, decimal_places)
                            highlight_change('FF6DCD57', ea_cell, new_version_of_file)
                            highlight_change('FF6DCD57', wet_cell, new_version_of_file)
                            highlight_change('FF6DCD57', dry_cell, new_version_of_file)
                            highlight_change('FF6DCD57', delta_e_cell, new_version_of_file)

                            # Calculate U from ea and delta_e
                            calculated_U = calculate_U(calculated_ea, delta_e)
                            U_cell.value = round(calculated_U, decimal_places)
                            highlight_change('FF6DCD57', U_cell, new_version_of_file)

            # If T'a is confirmed (green) but T is not, attempt to confirm ea, Δe, U, and T
            if is_highlighted(wet_cell, 'FF6DCD57') and not is_highlighted(dry_cell, 'FF6DCD57'):

                if is_string_convertible_to_float(wet_cell.value) and is_string_convertible_to_float(ea_cell.value) and is_string_convertible_to_float(dry_cell.value):
                    Ta = float(wet_cell.value)
                    T = float(dry_cell.value)
                    transcribed_ea = float(ea_cell.value)

                    # Step 1: Calculate ea from T'a
                    calculated_ea = 6.112 * np.exp((17.67 * Ta) / (Ta + 243.5))

                    if np.abs(calculated_ea - transcribed_ea) <= uncertainty_margin:
                        # Confirm ea
                        highlight_change('FF6DCD57', ea_cell, new_version_of_file)

                        # Step 2: Calculate es from T (dry bulb)
                        es = 6.112 * np.exp((17.67 * T) / (T + 243.5))

                        # Step 3: Compute Δe and U
                        calculated_delta_e = es - transcribed_ea
                        calculated_U = calculate_U(transcribed_ea, calculated_delta_e)

                        # Step 4: Compare with transcribed Δe and U
                        delta_e_ok = False
                        U_ok = False

                        if is_string_convertible_to_float(delta_e_cell.value):
                            transcribed_delta_e = float(delta_e_cell.value)
                            if np.abs(calculated_delta_e - transcribed_delta_e) <= uncertainty_margin:
                                highlight_change('FF6DCD57', delta_e_cell, new_version_of_file)
                                delta_e_ok = True

                        if is_string_convertible_to_float(U_cell.value):
                            transcribed_U = float(U_cell.value)
                            if np.abs(calculated_U - transcribed_U) <= uncertainty_margin:
                                highlight_change('FF6DCD57', U_cell, new_version_of_file)
                                U_ok = True

                        # Step 5: If either Δe or U matched, update/correct the other and confirm T
                        if delta_e_ok or U_ok:
                            if not delta_e_ok:
                                delta_e_cell.value = round(calculated_delta_e, decimal_places)
                                highlight_change('FF6DCD57', delta_e_cell, new_version_of_file)

                            if not U_ok:
                                U_cell.value = round(calculated_U, decimal_places)
                                highlight_change('FF6DCD57', U_cell, new_version_of_file)

                            # Now confirm T as correct
                            highlight_change('FF6DCD57', dry_cell, new_version_of_file)
    
    new_workbook.save(new_version_of_file)
    # Save intermediate version after rechecking the temperature thresholds for the transcribed daily values of maximum, minimum amd average.
    save_intermediate_version(new_workbook, "dryandwetbulb", transient_transcription_output_dir_station, month_filename)

    # RECALCULATIONS             
    # RECALCULATION (1)
    # New logic to recalculate multi-day totals and averages if values of all the days leading up to it are confirmed in previous QA/QC steps (highlighted green)
    for current_index, row in enumerate(rows_to_process):

        for column in daily_temperature_columns + daily_precipitation_column + dry_and_wet_bulb_temperature_columns: # Where these columns to check represent [ Max Temperature, Min Temperature, Average Temperature] 

            # Dynamically determine the number of offset cells (days considered in calculation of multi-day totals) based on row gaps, header rows, and presence of multi-day averages (placed immediatelly after the totals)
            if current_index == 0:
                # For the first row with multi-day totals, calculate offset_cells by subtracting the header rows. We also subtract by 1 here because the row index start at 0
                offset_cells = row - header_rows - 1
            else:
                # For other rows, calculate offset based on the gap between rows
                if multi_day_averages:
                    multi_day_average_row = 1
                else:
                    multi_day_average_row = 0

                offset_cells = multi_day_totals_rows[current_index] - multi_day_totals_rows[current_index - 1] - multi_day_average_row - 1
            
            # Ensure the offset cells do not exceed max_offset_days
            offset_cells = min(offset_cells, max_days_for_multi_day_total)
            
            # Days considered for total or average:
            days = []
            
            for day in range(offset_cells):
                cell_coordinate = new_worksheet[f"{column}{row - day - 1}"]
                days.append(cell_coordinate)
            
            # Get the value of the multi day total (in our case '5/6-day Total') that was transcribed. In months with 31 days, the period from the 26th to the 31st might form a 6-day total.
            cell_for_total = new_worksheet[f"{column}{row}"]
            confirmed_cell_for_total = is_highlighted(cell_for_total, 'FF6DCD57') # Check whether the multi-day total was confirmed (green)

            if multi_day_averages and column in daily_temperature_columns + dry_and_wet_bulb_temperature_columns:
                cell_for_average = new_worksheet[f"{column}{row + 1}"]
                confirmed_cell_for_average = is_highlighted(cell_for_average, 'FF6DCD57')  # Check if multi-day average was confirmed (green)

            else:
                confirmed_cell_for_average = False
            
            # Recalculate unconfirmed days if multi-day total and averages are confirmed
            if confirmed_cell_for_total or confirmed_cell_for_average:
                # If either the multi-day total or average is confirmed, use it to correct the missing values

                valid_day_values = [float(day.value) for day in days if day.value is not None and is_string_convertible_to_float(day.value) and is_highlighted(day, 'FF6DCD57')]  # Get valid (confirmed) day values
                
                if len(valid_day_values) == len(days) - 1:  # If only one day is unconfirmed
                    non_confirmed_day = next((day for day in days if not is_highlighted(day, 'FF6DCD57')), None)  # Identify the unconfirmed day
                    # Use confirmed total to calculate the missing day
                    if confirmed_cell_for_total and is_string_convertible_to_float(cell_for_total.value) and non_confirmed_day is not None:
                        total_value = float(cell_for_total.value)
                        non_confirmed_day_value = total_value - sum(valid_day_values)
                        if (min_temperature_threshold <= non_confirmed_day_value <= max_temperature_threshold): 
                            non_confirmed_day.value = round(non_confirmed_day_value, 1)
                            highlight_change('FF6DCD57', non_confirmed_day, new_version_of_file)  # Mark the corrected day in green (now confirmed)
                            new_workbook.save(new_version_of_file)
                    
                    # Use confirmed average to calculate the missing day
                    elif confirmed_cell_for_average and is_string_convertible_to_float(cell_for_average.value) and non_confirmed_day is not None:
                        average_value = float(cell_for_average.value)
                        total_value_from_average = average_value * len(days)
                        non_confirmed_day_value = total_value_from_average - sum(valid_day_values)
                        if (min_temperature_threshold <= non_confirmed_day_value <= max_temperature_threshold):
                            non_confirmed_day.value = round(non_confirmed_day_value, 1)
                            highlight_change('FF6DCD57', non_confirmed_day, new_version_of_file)  # Mark the corrected day in green
                            new_workbook.save(new_version_of_file)

            # Recalculate multi-day total if all days are confirmed
            if all(is_highlighted(day, 'FF6DCD57') for day in days) and not confirmed_cell_for_total:
                total_value = sum(float(day.value) for day in days if day.value is not None and is_string_convertible_to_float(day.value))
                cell_for_total.value = round(total_value, 1)
                highlight_change('FF6DCD57', cell_for_total, new_version_of_file)  # Highlight total in green
                new_workbook.save(new_version_of_file)

            # Recalculate multi-day average if all days are confirmed
            if multi_day_averages and column in daily_temperature_columns + dry_and_wet_bulb_temperature_columns and all(is_highlighted(day, 'FF6DCD57') for day in days) and not confirmed_cell_for_average:
                average_value = mean(float(day.value) for day in days if day.value is not None and is_string_convertible_to_float(day.value))
                cell_for_average.value = round(average_value, 1)
                highlight_change('FF6DCD57', cell_for_average, new_version_of_file)  # Highlight average in green
                new_workbook.save(new_version_of_file)

       
    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)
    # Save intermediate version after recalculating the multi-day totals and averages after confirmation of more daily values
    save_intermediate_version(new_workbook, "recalcmultiday", transient_transcription_output_dir_station, month_filename)
    

    # RECALCULATION (2)
    # Here we re-do all the multi-day total and average checks, after the previous QA/QC steps that corrected some previously wrongly transcribed values
    for current_index, row in enumerate(rows_to_process):

        for column in daily_temperature_columns + daily_precipitation_column + dry_and_wet_bulb_temperature_columns: # Where these columns to check represent [ Max Temperature, Min Temperature, Average Temperature] 
            # Get the value of the multi day total (in our case '5/6-day Total') that was transcribed. Note Months with 31 days ahve some 6 day todays considering 26th to 31st
            
            # Get the value of the multi-day total for the current row
            cell_for_total = new_worksheet[f"{column}{row}"]
            total_value = cell_for_total.value
            confirmed_cell_for_total = is_highlighted(cell_for_total, 'FF6DCD57') # Check whether the multi-day total was confirmed (green)


            if total_value is None or total_value == '':
                highlight_change('FFC0CB', cell_for_total, new_version_of_file)  # Highlight empty cells in pink
                continue
            
            # Dynamically determine the number of offset cells (days considered in calculation of multi-day totals) based on row gaps, header rows, and presence of multi-day averages (placed immediatelly after the totals)
            if current_index == 0:
                # For the first row with multi-day totals, calculate offset_cells by subtracting the header rows. We also subtract by 1 here because the row index start at 0
                offset_cells = row - header_rows - 1
            else:
                # For other rows, calculate offset based on the gap between rows
                if multi_day_averages:
                    multi_day_average_row = 1
                else:
                    multi_day_average_row = 0

                offset_cells = multi_day_totals_rows[current_index] - multi_day_totals_rows[current_index - 1] - multi_day_average_row - 1
            
            # Ensure the offset cells do not exceed max_offset_days
            offset_cells = min(offset_cells, max_days_for_multi_day_total)

            # Calculate cell values for the days leading up to the multi-day totals
            cell_values_for_days = []
            blank_cells = []

            # Days considered for total or average:
            days_considered = []
            
            for day in range(offset_cells):
                cell_coordinate = new_worksheet[f"{column}{row - day - 1}"]
                cell_value = cell_coordinate.value
                days_considered.append(cell_coordinate)
                if cell_value is None or cell_value == '':
                    cell_value = 0.0  # Replace empty values with 0.0 for summation. This is to avoid errors in Excel summation.
                    cell_values_for_days.append(cell_value)
                    highlight_change('FFFFFF', cell_coordinate, new_version_of_file) #No Fill 
                    blank_cells.append(1.0)
                else:
                    if is_string_convertible_to_float(cell_value):
                        cell_value = float(cell_value)
                        if column in daily_temperature_columns + dry_and_wet_bulb_temperature_columns:
                            if cell_value >= 100.0: 
                                manipulated_value = cell_value/10.0 # Special manipulation. If transcribed temperature value > 100, divide by 10.
                                cell_value = manipulated_value
                                highlight_change('CC3300', cell_coordinate, new_version_of_file) #CC3300 is Dark Red. Highlight to show that this special manupilation was done on this cell.
                            if cell_value < min_temperature_threshold or cell_value > max_temperature_threshold:
                                highlight_change('CC3300', cell_coordinate, new_version_of_file) #CC3300 is Dark Red. Highlight to show that temperature exceeds set thresholds for maximum or minimum 
                        cell_values_for_days.append(cell_value)
                        cell_coordinate.value = cell_value
                        new_workbook.save(new_version_of_file)
                    else:
                        cell_values_for_days.append(0.0) # Debugging: Incase the value transcribed in the cell is a string, convert it to 0.0 to avoid errors in the summation.


            # Sum the transcribed daily values and compare them with the transcribed multi-day totals
            total_sum = sum(cell_values_for_days)
            if abs(total_sum - float(total_value)) <= uncertainty_margin: 
                highlight_change('FF6DCD57', cell_for_total, new_version_of_file)  #FF6DCD57 is Green. Highlight to show that the transcribed multi-day total values confirmed that the transcribed daily values are correct, and vice versa
                
                # Additionally highlight the transcribed daily values that lead to correct transcribed multi-day totals
                for i in range(offset_cells):
                    cell_to_highlight = new_worksheet[f"{column}{row - i - 1}"]  # Access the previous offset cells. These are the cells above the multi-day total
                    highlight_change('FF6DCD57', cell_to_highlight, new_version_of_file)  # FF6DCD57 is Green. Highlight to confirm these day values are part of the correct multi-day total calculation.
                new_workbook.save(new_version_of_file)
            else:
                highlight_change('75696F', cell_for_total, new_version_of_file)  #75696F is Grey. Highlight to show that transcribed multi-day total values are not equal to the transcribed daily values, and hence we cant confirm if the transcription was correct.
                new_workbook.save(new_version_of_file)
            
            if multi_day_averages and column in daily_temperature_columns + dry_and_wet_bulb_temperature_columns:
                cell_for_average = new_worksheet[f"{column}{row + 1}"]
                average_value = cell_for_average.value
                confirmed_cell_for_average = is_highlighted(cell_for_average, 'FF6DCD57')  # Check if multi-day average was confirmed (green)

                if average_value is None or average_value == '':
                    highlight_change('FFC0CB', cell_for_average, new_version_of_file)  #FFC0CB is pink.  Highlight empty cells in pink to show that the average value was not transcribed (or is missing)
                    continue
                # Calculate the average from the previous offset cells
                total_blank_cells = sum(blank_cells) # Count of blank cells. This is to avoid calculating the multi-day average whist considering missing days. i.e. for 28-30 day months, average in the last pentad isn't calculated using the same offset cells as that for 31 day months
                days = float(offset_cells - total_blank_cells) # Days considered for multi-day average calculation
                
                if average_value is not None and is_string_convertible_to_float(average_value) and total_sum is not None and total_sum != '' and days != 0:

                    calculated_average = total_sum / days
                    if abs(calculated_average - float(average_value)) <= uncertainty_margin:
                        highlight_change('FF6DCD57', cell_for_average, new_version_of_file)  # FF6DCD57 is Green. Highlight to show that the transcribed multi-day average values confirmed that the transcribed daily values are correct, and vice versa
                        
                        # Additionally highlight the transcribed daily values that lead to correct transcribed multi-day average
                        for i in range(offset_cells):
                            cell_to_highlight = new_worksheet[f"{column}{row - i - 1}"]  # Access the previous offset cells. These are the cells above the multi-day total
                            highlight_change('FF6DCD57', cell_to_highlight, new_version_of_file)  # FF6DCD57 is Green. Highlight to confirm these day values are part of the correct multi-day average calculation.

                        new_workbook.save(new_version_of_file)
                    else:
                        highlight_change('75696F', cell_for_average, new_version_of_file)  # #75696F is Grey. Highlight to show that transcribed multi-day total average value is not equal to the transcribed daily values, and hence we cant confirm if the transcription was correct.
                        new_workbook.save(new_version_of_file)  


            # Recalculate multi-day total if all days are confirmed
            if all(is_highlighted(day, 'FF6DCD57') for day in days_considered) and not confirmed_cell_for_total:
                total_value = sum(float(day.value) for day in days_considered if day.value is not None and is_string_convertible_to_float(day.value))
                cell_for_total.value = round(total_value, 1)
                highlight_change('FF6DCD57', cell_for_total, new_version_of_file)  # Highlight total in green
                new_workbook.save(new_version_of_file)

            # Recalculate multi-day average if all days are confirmed
            if multi_day_averages and column in daily_temperature_columns + dry_and_wet_bulb_temperature_columns and all(is_highlighted(day, 'FF6DCD57') for day in days_considered) and not confirmed_cell_for_average:
                average_value = mean(float(day.value) for day in days_considered if day.value is not None and is_string_convertible_to_float(day.value))
                cell_for_average.value = round(average_value, 1)
                highlight_change('FF6DCD57', cell_for_average, new_version_of_file)  # Highlight average in green
                new_workbook.save(new_version_of_file)



    
    ## FINAL RE-CHECKS
    # Check the cells by rows to ensure that Min temp threshold < Temp < Max Temp threshold
    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row, min_col=min_col, max_col=max_col):
        if row[0].row in excluded_rows:
            continue 
        
        # Create a dictionary to map column names (D, E, F, etc.) to the cell values
        row_cells = {daily_temperature_columns[i]: row[columns_to_check_indices[i] - min_col] for i in range(len(daily_temperature_columns))}

        # Now you can access the cells dynamically using the column names
        D = row_cells.get('D')  # Max Temp
        E = row_cells.get('E')  # Min Temp
        F = row_cells.get('F')  # Average Temp

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        if D_value_exists:
            D_value = float(D.value)
            if not (min_temperature_threshold <= D_value <= max_temperature_threshold): 
                highlight_change('CC3300', D, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct

        if E_value_exists:
            E_value = float(E.value)
            if not (min_temperature_threshold <= E_value <= max_temperature_threshold):
                highlight_change('CC3300', E, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct
        
        if F_value_exists:
            F_value = float(F.value)
            if not (min_temperature_threshold <= F_value <= max_temperature_threshold):
                highlight_change('CC3300', F, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct

    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)
    # Save intermediate version after rechecking the temperature thresholds for the transcribed daily values of maximum, minimum amd average.
    save_intermediate_version(new_workbook, "2ndrechkthresh", transient_transcription_output_dir_station, month_filename)


    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)

    
    # Final recheck for totals and averages
    for current_index, row in enumerate(rows_to_process):

        for column in daily_temperature_columns + daily_precipitation_column + dry_and_wet_bulb_temperature_columns: # Where these columns to check represent [ Max Temperature, Min Temperature, Average Temperature] 
            # Get the value of the multi day total (in our case '5/6-day Total') that was transcribed. Note Months with 31 days ahve some 6 day todays considering 26th to 31st
            
            # Get the value of the multi-day total for the current row
            cell_for_total = new_worksheet[f"{column}{row}"]
            total_value = cell_for_total.value

            if total_value is None or total_value == '':
                highlight_change('FFC0CB', cell_for_total, new_version_of_file)  # Highlight empty cells in pink
                continue
            
            # Dynamically determine the number of offset cells (days considered in calculation of multi-day totals) based on row gaps, header rows, and presence of multi-day averages (placed immediatelly after the totals)
            if current_index == 0:
                # For the first row with multi-day totals, calculate offset_cells by subtracting the header rows. We also subtract by 1 here because the row index start at 0
                offset_cells = row - header_rows - 1
            else:
                # For other rows, calculate offset based on the gap between rows
                if multi_day_averages:
                    multi_day_average_row = 1
                else:
                    multi_day_average_row = 0

                offset_cells = multi_day_totals_rows[current_index] - multi_day_totals_rows[current_index - 1] - multi_day_average_row - 1
            
            # Ensure the offset cells do not exceed max_offset_days
            offset_cells = min(offset_cells, max_days_for_multi_day_total)

            # Calculate cell values for the days leading up to the multi-day totals
            confirmed_cell_values_for_days = []

            blank_cells = []

            for day in range(offset_cells):
                cell_coordinate = new_worksheet[f"{column}{row - day - 1}"]
                cell_value = cell_coordinate.value
                if cell_value is None or cell_value == '':
                    blank_cells.append(1.0)
                if is_string_convertible_to_float(cell_value) and is_highlighted(cell_coordinate, 'FF6DCD57'): # Check is day is confirmed
                    confirmed_cell_values_for_days.append(cell_value)
                    
                    
            # Sum the transcribed confirmed daily values and compare them with the transcribed multi-day totals
            total_sum = sum(confirmed_cell_values_for_days)
            if abs(total_sum - float(total_value)) <= uncertainty_margin: 
                highlight_change('FF6DCD57', cell_for_total, new_version_of_file)  #FF6DCD57 is Green. Highlight to show that the transcribed multi-day total values confirmed that the transcribed daily values are correct, and vice versa
                new_workbook.save(new_version_of_file)
            else:
                highlight_change('75696F', cell_for_total, new_version_of_file)  #75696F is Grey. Highlight to show that transcribed multi-day total values are not equal to the transcribed daily values, and hence we cant confirm if the transcription was correct.
                new_workbook.save(new_version_of_file)
            
            if multi_day_averages and column in daily_temperature_columns + dry_and_wet_bulb_temperature_columns:
                cell_for_average = new_worksheet[f"{column}{row + 1}"]
                average_value = cell_for_average.value

                if average_value is None or average_value == '':
                    highlight_change('FFC0CB', cell_for_average, new_version_of_file)  #FFC0CB is pink.  Highlight empty cells in pink to show that the average value was not transcribed (or is missing)
                    continue
                
                # Calculate the average from the previous offset cells
                total_blank_cells = sum(blank_cells) # Count of blank cells. This is to avoid calculating the multi-day average whist considering missing days. i.e. for 28-30 day months, average in the last pentad isn't calculated using the same offset cells as that for 31 day months
                days = float(offset_cells - total_blank_cells) # Days considered for multi-day average calculation
                
                if average_value is not None and is_string_convertible_to_float(average_value) and total_sum is not None and total_sum != '' and days !=0 and not is_highlighted(cell_for_average, 'CC3300'):

                    calculated_average = total_sum / days
                    if abs(calculated_average - float(average_value)) <= uncertainty_margin:
                        highlight_change('FF6DCD57', cell_for_average, new_version_of_file)  # FF6DCD57 is Green. Highlight to show that the transcribed multi-day average values confirmed that the transcribed daily values are correct, and vice versa
                        new_workbook.save(new_version_of_file)
                    if abs(calculated_average - float(total_value / days)) <= uncertainty_margin:
                        highlight_change('FF6DCD57', cell_for_average, new_version_of_file)  # FF6DCD57 is Green. Highlight to show that the transcribed multi-day average values confirmed that the transcribed daily values are correct, and vice versa
                        new_workbook.save(new_version_of_file)
                    else:
                        highlight_change('75696F', cell_for_average, new_version_of_file)  # #75696F is Grey. Highlight to show that transcribed multi-day total average value is not equal to the transcribed daily values, and hence we cant confirm if the transcription was correct.
                        new_workbook.save(new_version_of_file)                


    # Check the cells by rows to ensure that Min temp < Average Temp < Max Temp 
    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row, min_col=min_col, max_col=max_col):
        if row[0].row in excluded_rows:
            continue 
        
        # Create a dictionary to map column names (D, E, F, etc.) to the cell values
        row_cells = {daily_temperature_columns[i]: row[columns_to_check_indices[i] - min_col] for i in range(len(daily_temperature_columns))}

        # Now you can access the cells dynamically using the column names
        D = row_cells.get('D')  # Max Temp
        E = row_cells.get('E')  # Min Temp
        F = row_cells.get('F')  # Average Temp

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        if D_value_exists and E_value_exists and F_value_exists:
            D_value = float(D.value)
            E_value = float(E.value)
            F_value = float(F.value)
            if not (E_value <= F_value):
                highlight_change('CC3300', E, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct
                highlight_change('CC3300', F, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct
            if not (F_value <= D_value):
                highlight_change('CC3300', D, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct
                highlight_change('CC3300', F, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct


    # Check the cells by rows to ensure that Min temp threshold < Temp < Max Temp threshold
    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row, min_col=min_col, max_col=max_col):
        if row[0].row in excluded_rows:
            continue 
        
        # Create a dictionary to map column names (D, E, F, etc.) to the cell values
        row_cells = {daily_temperature_columns[i]: row[columns_to_check_indices[i] - min_col] for i in range(len(daily_temperature_columns))}

        # Now you can access the cells dynamically using the column names
        D = row_cells.get('D')  # Max Temp
        E = row_cells.get('E')  # Min Temp
        F = row_cells.get('F')  # Average Temp

        # Check if the highlighted cells have values
        D_value_exists = is_string_convertible_to_float(D.value)
        E_value_exists = is_string_convertible_to_float(E.value)
        F_value_exists = is_string_convertible_to_float(F.value)

        if D_value_exists:
            D_value = float(D.value)
            if not (min_temperature_threshold <= D_value <= max_temperature_threshold): 
                highlight_change('CC3300', D, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct

        if E_value_exists:
            E_value = float(E.value)
            if not (min_temperature_threshold <= E_value <= max_temperature_threshold):
                highlight_change('CC3300', E, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct
        
        if F_value_exists:
            F_value = float(F.value)
            if not (min_temperature_threshold <= F_value <= max_temperature_threshold):
                highlight_change('CC3300', F, new_version_of_file)  # Dark Red. Highlight to show that transcribed temp value may not be correct

    # Save the workbook after all changes
    new_workbook.save(new_version_of_file)

    # Save intermediate version after Last QA/QC check
    save_intermediate_version(new_workbook, "lastcheck", transient_transcription_output_dir_station, month_filename)


    # === FINAL STEP: Export confirmed green cells to text files === This will be important for further training (continous learning of handwritting styles) of the OCR
    roi_save_dir = os.path.join(transient_transcription_output_dir_station, 'clipped_cells')
    os.makedirs(roi_save_dir, exist_ok=True)

    for row in new_worksheet.iter_rows(min_row=header_rows+1, max_row=new_worksheet.max_row):
        for cell in row:
            # Check if the cell is green-highlighted (confirmed)
            if is_highlighted(cell, 'FF6DCD57') and is_string_convertible_to_float(cell.value):
                # Get Excel-style cell reference (e.g., D4)
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                cell_ref = f"{col_letter}{cell.row}"

                # Construct filename
                txt_filename = f"{month_filename}_{cell_ref}.txt"
                txt_save_path = os.path.join(roi_save_dir, txt_filename)

                # Remove decimal point (no rounding). This is because in the original transcription, decimal points were not transcribed (to reduce the noise on the sheets from dotted lines and poor maintenance conditions)
                clean_val = str(cell.value).replace('.', '')

                # Save to text file
                with open(txt_save_path, 'w', encoding='utf-8') as f:
                    f.write(clean_val)
    
    # Save final changes
    new_workbook.save(new_version_of_file)    
    new_workbook.close()

    return new_workbook

