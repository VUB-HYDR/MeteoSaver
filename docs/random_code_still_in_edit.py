#!/usr/bin/env python
import os, argparse, glob, tempfile, shutil, warnings
import cv2
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
#from paddleocr import PaddleOCR,draw_ocr
from PIL import Image, ImageDraw, ImageFont
import pytesseract
import easyocr
import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans


def organize_contours(contours, image_shape, max_rows):
    '''
    # Organizes the bounding boxes, here termed as contours, in rows by their center co-ordinates using the KMeans clustering

    Parameters
    --------------
    contours: List of contours for the detected text in the table cells with coordinates
    image_shape: Image of the table
    max_rows: Maximum rows, adjust based on your table's expected structure

    Returns
    -------------- 
    rows: Bounding boxes organised in rows using Kmeans clustering

    '''

    image_height, image_width, _ = image_shape
    midpoints = [(cv2.boundingRect(contour)[1] + cv2.boundingRect(contour)[3] // 2) for contour in contours]
    if len(midpoints) == 0:
        return []
    kmeans = KMeans(n_clusters=min(max_rows, len(midpoints)), random_state=0)
    kmeans.fit(np.array(midpoints).reshape(-1, 1))
    labels = kmeans.labels_
    rows = [[] for _ in range(max_rows)]
    for label, contour in zip(labels, contours):
        rows[label].append(contour)
    for i in range(len(rows)):
        rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])
    return rows

# def organize_contours_bottom(contours, image_shape, max_rows):
    # '''
    # # Organizes the bounding boxes, here termed as contours, in rows by their bottom co-ordinates using the KMeans clustering

    # Parameters
    # --------------
    # contours: List of contours for the detected text in the table cells with coordinates
    # image_shape: Image of the table
    # max_rows: Maximum rows, adjust based on your table's expected structure

    # Returns
    # -------------- 
    # rows: Bounding boxes organised in rows using Kmeans clustering

    # '''

#     image_height, image_width, _ = image_shape
#     bottom = [(cv2.boundingRect(contour)[1] + cv2.boundingRect(contour)[3]) for contour in contours]
#     if len(bottom) == 0:
#         return []
#     kmeans = KMeans(n_clusters=min(max_rows, len(bottom)), random_state=0)
#     kmeans.fit(np.array(bottom).reshape(-1, 1))
#     labels = kmeans.labels_

#     rows = [[] for _ in range(max_rows)]
#     for label, contour in zip(labels, contours):
#         rows[label].append(contour)
    
#     for i in range(len(rows)):
#         rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])
    
#     return rows


def organize_contours_top(contours, image_shape, max_rows):
    '''
    # Organizes the bounding boxes, here termed as contours, in rows by their top co-ordinates using the KMeans clustering

    Parameters
    --------------
    contours: List of contours for the detected text in the table cells with coordinates
    image_shape: Image of the table
    max_rows: Maximum rows, adjust based on your table's expected structure

    Returns
    -------------- 
    rows: Bounding boxes organised in rows using Kmeans clustering

    '''

    image_height, image_width, _ = image_shape
    # Top for vertical clustering
    top = [cv2.boundingRect(contour)[1] for contour in contours]
    if len(top) == 0:
        return []

    kmeans = KMeans(n_clusters=min(max_rows, len(top)), random_state=0)
    kmeans.fit(np.array(top).reshape(-1, 1))
    labels = kmeans.labels_

    rows = [[] for _ in range(max_rows)]
    for label, contour in zip(labels, contours):
        rows[label].append(contour)

    for i in range(len(rows)):
        rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])

    return rows


def get_excel_cell_coordinates(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    cell_coordinates = {}
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=43, max_col=24):
        for cell in row:
            if cell.value is None:
                cell_coordinates[cell.coordinate] = (cell.column, cell.row)
                
    return cell_coordinates

def normalize_coordinates(coord, dimension):
    return coord / dimension

def normalize_cell_coordinates(cell_coordinates, sheet_dimensions):
    normalized_cell_coords = {}
    for cell_ref, (col, row) in cell_coordinates.items():
        normalized_col = normalize_coordinates(col, sheet_dimensions['width'])
        normalized_row = normalize_coordinates(row, sheet_dimensions['height'])
        normalized_cell_coords[cell_ref] = (normalized_col, normalized_row)
    return normalized_cell_coords

def find_closest_cell(normalized_center_x, normalized_center_y, normalized_cell_coordinates):
    closest_cell = None
    min_distance = float('inf')
    
    for cell_ref, (normalized_col, normalized_row) in normalized_cell_coordinates.items():
        distance = np.sqrt((normalized_center_x - normalized_col) ** 2 + (normalized_center_y - normalized_row) ** 2)
        
        if distance < min_distance:
            min_distance = distance
            closest_cell = cell_ref
    
    return closest_cell

# def find_next_empty_cell(ws, start_cell_ref):
#     column = start_cell_ref[0]
#     start_row = int(start_cell_ref[1:])
    
#     current_row = start_row
#     while ws[f"{column}{current_row}"].value is not None:
#         current_row += 1
    
#     return f"{column}{current_row}"

# def find_previous_cell(ws, start_cell_ref):
#     column = start_cell_ref[0]
#     start_row = int(start_cell_ref[1:])

#     # Check the very previous cell only
#     if start_row > 1:
#         next_row = start_row - 1

#         return f"{column}{next_row}"
#     else:
#         return f"{column}{start_row}"

def find_next_cell(ws, start_cell_ref):
    column = start_cell_ref[0]
    start_row = int(start_cell_ref[1:])

    # Check the very next cell only
    next_row = start_row + 1

    return f"{column}{next_row}"

# def find_adjacent_empty_cell(ws, start_cell_ref):
#     column = start_cell_ref[0]
#     start_row = int(start_cell_ref[1:])

#     # Check the cell directly below
#     current_row = start_row
#     if ws[f"{column}{current_row}"].value is not None:
#         current_row += 1

#     # Check the cell directly above
#     if current_row > 0 and ws[f"{column}{current_row}"].value is not None:
#         current_row -= 1
    
#     return f"{column}{current_row}"

    # # If both adjacent cells are occupied or out of bounds, return None
    # elseif: 
    #     return f"{column}{start_row}

def transcription(detected_table_cells, ocr_model):
    '''
    # Makes use of a pre-processed image (in grayscale) to detect the table from the record sheets

    Parameters
    --------------
    detected_table_cells where: 
        detected_table_cells[0]: contours. Contours for the detected text in the table cells
        detected_table_cells[1]: image_with_all_bounding_boxes. Bounding boxex for each cell for which clips will be made later before optical character recognition 
        detected_table_cells[2]: table_copy
        detected_table_cells[3]: table_original_image
    

    ocr_model: Optical Character Recognition/Handwritten Text Recognition of choice
    
    Returns
    -------------- 
    wb : Ms Excel workbook with OCR Results
    ''' 
    if ocr_model == 'Tesseract-OCR':
        ## Lauching Tesseract-OCR
        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' ## Here input the PATH to the Tesseract executable on your computer. See more information here: https://pypi.org/project/pytesseract/
    # if ocr_model == 'PaddleOCR':
    #     ## Lauching PaddleOCR, which would be used by downloading necessary files as shown below
    #     paddle_ocr = PaddleOCR(use_angle_cls=True, lang = 'en', use_gpu=False) ## Run only once to download all required files
    if ocr_model == 'EasyOCR':
        ## Lauching EasyOCR
        easyocr_reader = easyocr.Reader(['en']) # this needs to run only once to load the model into memory

    contours = detected_table_cells[0]
    image_with_all_bounding_boxes = detected_table_cells[1]
    table_copy = detected_table_cells[2]
    table_original_image = detected_table_cells[3]

    # Get the dimensions of the loaded image
    image_height, image_width, image_channels = image_with_all_bounding_boxes.shape

    results = []

    organize_methods = {
        #'Bottom': organize_contours_bottom,
        'Midpoint': organize_contours,
        'Top': organize_contours_top
    }

    for method_name, organize_method in organize_methods.items():

        ## Create an Excel workbook and add a worksheet where the transcribed text will be saved
        wb = Workbook()
        ws = wb.active
        ws.title = 'OCR_Results'

        # Sort contours by y-coordinate
        contours_sorted = sorted(contours, key=lambda c: cv2.boundingRect(c)[1])

        max_rows = 43  # maximum rows, adjust based on your table's expected structure. Here, I had started with 45 rows and it was already giving good results . Previously had it at 43
        organized_rows = organize_method(contours_sorted, (image_height, image_width, image_channels), max_rows)
        # organized_rows = organize_and_merge_contours(contours_sorted, (image_height, image_width, image_channels), max_rows)


        # # Define the row index to start filling from
        # start_row_index = 1  # Change this to your desired starting row index

        # Sort contours by y-coordinate
        # contours_sorted = sorted(contours, key=lambda c: cv2.boundingRect(c)[1])

        # Calculate the median y-coordinate for each row and sort rows by this median
        sorted_rows = sorted(organized_rows, key=lambda row: np.median([cv2.boundingRect(c)[1] + cv2.boundingRect(c)[3] // 2 for c in row]))

        for row_index, row in enumerate(sorted_rows, start=1):
            for contour in row:
        ## Text detection using an OCR model; Here using TesseractOCR
        # for contour in contours_sorted:

        #for contour in contours:
                x, y, w, h = cv2.boundingRect(contour)
                
                # Adjust these threshold values according to your requirements
                min_width_threshold = 20 # 50 previously
                min_height_threshold = 28  # Had this at 13 previously. 20, 25, 30. trying 28, but it might be over fitting 

                max_width_threshold = 200 # 120 previously
                max_height_threshold = 90 # 60 previously

                # Filter out smaller bounding boxes
                if (min_width_threshold <= w <= max_width_threshold and min_height_threshold <= h <= max_height_threshold):

                    # Calculate a factor to increase the bounding box area (e.g., 80% larger)
                    factor_width = 0.05  # Modify this factor as needed
                    increase_factor_height = 0.25 # Modify this factor as needed
                
                    x += int(w * factor_width)  # Increase width
                    y -= int(h * increase_factor_height)  # Increase height
                    w -= int(1 * w * factor_width)  # Decrease width a little to avoid vertical lines that may be transcribed as the number 1 yet they aren't a number
                    h += int(2 * h * increase_factor_height)  # Increase height
                    
                    # This line below is about
                    # drawing a rectangle on the image with the shape of
                    # the bounding box. Its not needed for the OCR.
                    # Its just added for debugging purposes.
                    image_with_all_bounding_boxes = cv2.rectangle(image_with_all_bounding_boxes, (x, y), (x + w, y + h), (0, 255, 0), 5)
                    
                    # Calculate center coordinates of the bounding box
                    center_x = x + w // 2
                    center_y = y + h // 2
                    
                    # OCR
                    # Crop each cell using the bounding rectangle coordinates
                    ROI = table_copy[y:y+h, x:x+w]
                    
                    if ROI.size != 0:  # Check if the height and width are greater than zero
                        
                        cv2.imwrite('detected.png', ROI)
                        if ocr_model == 'Tesseract-OCR':
                        # Using Tesseract-OCR
                            ocr_result = pytesseract.image_to_string('detected.png', lang='cobedore-V6', config='--psm 7 -c tessedit_char_whitelist=0123456789.') # Just added -c tessedit_char_whitelist=0123456789. to really limit the text detected
                            #ocr_result = pytesseract.image_to_string('detected.png', lang='ccc-base+eng+cobedore-V6', config='--psm 7 -c tessedit_char_whitelist=0123456789.')

                            # Here's a brief explanation of some Page Segmentation Modes (PSMs) available in Tesseract:
                            # 0: Orientation and script detection (OSD) only.
                            # 1: Automatic page segmentation with OSD.
                            # 2: Automatic page segmentation, but no OSD, or OCR.
                            # 3: Fully automatic page segmentation, but no OSD. (Default)
                            # 4: Assume a single column of text of variable sizes.
                            # 5: Assume a single uniform block of vertically aligned text.
                            # 6: Assume a single uniform block of text.
                            # 7: Treat the image as a single text line.
                            # 8: Treat the image as a single word.
                            # 9: Treat the image as a single word in a circle.
                            # 10: Treat the image as a single character.
                            # 11: Sparse text. Find as much text as possible in no particular order.
                            # 12: Sparse text with OSD.
                            # 13: Raw line. Treat the image as a single text line, bypassing hacks that are Tesseract-specific.


                        # if ocr_model == 'PaddleOCR':
                        #     ## Using PaddleOCR
                        #     ocr_result = paddle_ocr.ocr('detected.png', cls = True)

                        if ocr_model == 'EasyOCR':
                        # Using EasyOCR
                            ocr_result = easyocr_reader.readtext('detected.png', detail = 0, allowlist='0123456789')
                        # Using OCR to recognize text/transcription
                        if ocr_result is not None:
                            
                            # Maximum number of columns and rows. These can be changed depending on the tables in the images
                            max_column_index = 24  # Number of columns in the table. Total number is original unclipped image are 27
                            max_row_index = 43  # Estimated number of rows in the table  .Previosly had it at 57 and results were good.       even 56 was good.     had this previously at 43
                            
                            cell_width = max(image_width // max_column_index, min_width_threshold)
                            cell_height = max(image_height//max_row_index, min_height_threshold)
                            
                            # Track filled cells using a set
                            filled_cells = []
                            
                            
                            # Convert the x-coordinate to a column letter # Assuming x-coordinate translates to columns cell_column
                            # Ensure x is within the valid range for Excel column indices
                            
                            if 1 <= center_x <= image_width:  # Excel's maximum column index
                                column_index = int(max((center_x) // cell_width, 0)) + 1 # Ensure column index is at least 1
                                #cell_column = openpyxl.utils.get_column_letter(min(column_index, max_column_index))
                            else:
                                column_index = 1
                                #cell_column = 'A'  # Set a default column if x is out of range
                            
                            # if 1 <= center_y <= image_height:
                            #     # Calculate the row index based on the row ratio
                            #     row_index = (center_y / cell_height) + 1  # Calculate row index as a floating-point number
                                
                            #     # Round the row index to the nearest integer
                            #     cell_row = int(round(row_index))
                                
                            #     # Ensure the row index is within the valid range
                            #     cell_row = min(max(cell_row, 1), max_row_index)
                            #     # row_ratio = (center_y) // cell_height
                            #     # cell_row = min(int(row_ratio) + 1, max_row_index)  # Convert row ratio to integer and ensure it's within valid range
                            #     # #cell_row = min(center_y // cell_height + 0.5, max_row_index)
                            # else:
                            #     cell_row = 1  # Set a default row if y is out of range  # Assuming y-coordinate translates to rows



                            # Ms Excel Template cell coordinates
                            file_path = f'docs\Table_structure.xlsx'
                            sheet_name = 'clipped_tilted'
                            cell_coordinates = get_excel_cell_coordinates(file_path, sheet_name)
                            sheet_dimensions = {'width': max_column_index, 'height': max_row_index}  # Assuming max column index and max row index

                            normalized_cell_coords = normalize_cell_coordinates(cell_coordinates, sheet_dimensions)


                            # Find closest cell in template
                            # Normalize bounding box center coordinates
                            normalized_center_x = normalize_coordinates(center_x, image_width)
                            normalized_center_y = normalize_coordinates(center_y, image_height)
                            closest_cell = find_closest_cell(normalized_center_x, normalized_center_y, normalized_cell_coords)
            

                            # if closest_cell:
                            #     cell = ws[closest_cell]
                            #     cell.value = ocr_result

                            if closest_cell:
                            # Check if the cell is occupied and find the next empty cell in the same column
                                cell_ref = closest_cell
                                while ws[cell_ref].value is not None:
                                    cell_ref = find_next_cell(ws, cell_ref)
                                
                                cell = ws[cell_ref]
                                cell.value = ocr_result








                            # # Write the OCR value to the cell in the Excel sheet
                            # cell = ws.cell(row=row_index, column=column_index)


                            # # # Set the cell value to the OCR result
                            # # cell.value = ocr_result

                            # current_cell_value = cell.value

                            # # If the cell is empty, add the OCR result directly
                            # if current_cell_value is None:
                            #     cell.value = ocr_result
                            # else:
                            #     # If the cell already has content, append the new result
                            #     cell.value = f"{current_cell_value}{ocr_result}"

                            # Set up border styles for excel output
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))

                            # Loop through cells to apply borders
                            for row in ws.iter_rows(min_row=1, max_row=max_row_index, min_col=1, max_col=max_column_index):
                                for cell in row:
                                    cell.border = thin_border
                            
                        else:
                            print('No values detected in clip')
                    else:
                        print('ROI is empty or invalid')


        # Insert two columns on the left side of the excel sheet
        ws.insert_cols(1, 2)
        
        # Insert a new row at the top for headers
        ws.insert_rows(1, amount = 3)

        # # Merging cells for headers
        # # Merge cells for headers that span multiple columns
        # ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8)  # For "Températures extrêmes"
        # #*** Do this for the other cells that need to be merged


        # Define your headers (adjust as needed)
        headers = ["No de la pentade", "Date", "Bellani (gr. Cal/cm2) 6-6h", "Températures extrêmes", "", "", "", "", "Evaportation en cm3 6 - 6h", "", "Pluies en mm. 6-6h", "Température et Humidité de l'air à 6 heures", "", "", "", "", "Température et Humidité de l'air à 15 heures",  "", "", "", "", "Température et Humidité de l'air à 18 heures",  "", "", "", "", "Date"]
        sub_headers_1 = ["", "", "", "Abri", "", "", "", "", "Piche", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", ""]
        sub_headers_2 =["", "", "", "Max.", "Min.", "(M+m)/2", "Ampl.", "Min. gazon", "Abri.", "Ext.", "", "T", "T'a", "e.", "U", "∆e", "T", "T'a", "e.", "U", "∆e","T", "T'a", "e.", "U", "∆e", ""]

        # Add the headers to the first row
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)
        # Add the first row of sub-headers to the second row
        for col_num, sub_header in enumerate(sub_headers_1, start=1):
            ws.cell(row=2, column=col_num, value=sub_header)

        # Add the second row of sub-headers to the third row
        for col_num, sub_header in enumerate(sub_headers_2, start=1):
            ws.cell(row=3, column=col_num, value=sub_header)

        file_path = f'src\output\{method_name}_Excel_with_OCR_Results.xlsx'
        wb.save(file_path)
        results.append([file_path])

        #wb.save(f'{method_name}_Excel_with_OCR_Results.xlsx') 

        # plt.imshow(image_with_all_bounding_boxes)
        # plt.show()

    return results



#### BELOW IS ANOTHER TRIAL OF THE SAME FUNCTIONS:




#!/usr/bin/env python
import os, argparse, glob, tempfile, shutil, warnings
import cv2
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
#from paddleocr import PaddleOCR,draw_ocr
from PIL import Image, ImageDraw, ImageFont
import pytesseract
import easyocr
import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
import pandas as pd
from openpyxl import load_workbook, Workbook


def organize_contours(contours, image_shape, max_rows):
    '''
    # Organizes the bounding boxes, here termed as contours, in rows by their center co-ordinates using the KMeans clustering

    Parameters
    --------------
    contours: List of contours for the detected text in the table cells with coordinates
    image_shape: Image of the table
    max_rows: Maximum rows, adjust based on your table's expected structure

    Returns
    -------------- 
    rows: Bounding boxes organised in rows using Kmeans clustering

    '''

    image_height, image_width, _ = image_shape
    midpoints = [(cv2.boundingRect(contour)[1] + cv2.boundingRect(contour)[3] // 2) for contour in contours]
    if len(midpoints) == 0:
        return []
    kmeans = KMeans(n_clusters=min(max_rows, len(midpoints)), random_state=0)
    kmeans.fit(np.array(midpoints).reshape(-1, 1))
    labels = kmeans.labels_
    rows = [[] for _ in range(max_rows)]
    for label, contour in zip(labels, contours):
        rows[label].append(contour)
    for i in range(len(rows)):
        rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])
    return rows

# def organize_contours_bottom(contours, image_shape, max_rows):
    # '''
    # # Organizes the bounding boxes, here termed as contours, in rows by their bottom co-ordinates using the KMeans clustering

    # Parameters
    # --------------
    # contours: List of contours for the detected text in the table cells with coordinates
    # image_shape: Image of the table
    # max_rows: Maximum rows, adjust based on your table's expected structure

    # Returns
    # -------------- 
    # rows: Bounding boxes organised in rows using Kmeans clustering

    # '''

#     image_height, image_width, _ = image_shape
#     bottom = [(cv2.boundingRect(contour)[1] + cv2.boundingRect(contour)[3]) for contour in contours]
#     if len(bottom) == 0:
#         return []
#     kmeans = KMeans(n_clusters=min(max_rows, len(bottom)), random_state=0)
#     kmeans.fit(np.array(bottom).reshape(-1, 1))
#     labels = kmeans.labels_

#     rows = [[] for _ in range(max_rows)]
#     for label, contour in zip(labels, contours):
#         rows[label].append(contour)
    
#     for i in range(len(rows)):
#         rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])
    
#     return rows


def organize_contours_top(contours, image_shape, max_rows):
    '''
    # Organizes the bounding boxes, here termed as contours, in rows by their top co-ordinates using the KMeans clustering

    Parameters
    --------------
    contours: List of contours for the detected text in the table cells with coordinates
    image_shape: Image of the table
    max_rows: Maximum rows, adjust based on your table's expected structure

    Returns
    -------------- 
    rows: Bounding boxes organised in rows using Kmeans clustering

    '''

    image_height, image_width, _ = image_shape
    # Top for vertical clustering
    top = [cv2.boundingRect(contour)[1] for contour in contours]
    if len(top) == 0:
        return []

    kmeans = KMeans(n_clusters=min(max_rows, len(top)), random_state=0)
    kmeans.fit(np.array(top).reshape(-1, 1))
    labels = kmeans.labels_

    rows = [[] for _ in range(max_rows)]
    for label, contour in zip(labels, contours):
        rows[label].append(contour)

    for i in range(len(rows)):
        rows[i] = sorted(rows[i], key=lambda c: cv2.boundingRect(c)[0])

    return rows


# def get_excel_cell_coordinates(file_path, sheet_name):
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb[sheet_name]
    
#     cell_coordinates = {}
#     for row in ws.iter_rows(min_row=1, min_col=1, max_row=43, max_col=24):
#         for cell in row:
#             if cell.value is None:
#                 cell_coordinates[cell.coordinate] = (cell.column, cell.row)
                
#     return cell_coordinates

def get_excel_cell_coordinates(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    cell_coordinates = {}
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=55, max_col=24):
        for cell in row:
            cell_coordinates[cell.coordinate] = (cell.column-1, cell.row-1) 
                
    return cell_coordinates

def normalize_coordinates(coord, dimension):
    return (coord) / (dimension)

# def normalize_coordinates(value, min_value, max_value):
#     return (value - min_value) / (max_value - min_value)

def get_table_boundaries(contours):
    x_coords = []
    y_coords = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        x_coords.extend([x, x + w])
        y_coords.extend([y, y + h])
    return min(x_coords), min(y_coords), max(x_coords), max(y_coords)

# def get_table_boundaries(contours):
#     min_x = min_y = float('inf')
#     max_x = max_y = float('-inf')
#     for contour in contours:
#         x, y, w, h = cv2.boundingRect(contour)
#         min_x = min(min_x, x)
#         min_y = min(min_y, y)
#         max_x = max(max_x, x + w)
#         max_y = max(max_y, y + h)
#     return min_x, min_y, max_x, max_y

def get_external_table_boundaries(contours):
    min_x = min_y = float('inf')
    max_x = max_y = float('-inf')
    
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        min_x = min(min_x, x)
        min_y = min(min_y, y)
        max_x = max(max_x, x + w)
        max_y = max(max_y, y + h)
    
    return min_x, min_y, max_x, max_y

# def normalize_cell_coordinates(cell_coordinates, sheet_dimensions):
#     width, height = sheet_dimensions['width'], sheet_dimensions['height']
#     normalized_coords = [(normalize_coordinates(x, 0, width), normalize_coordinates(y, 0, height)) for x, y in cell_coordinates]
#     return normalized_coords

def normalize_cell_coordinates(cell_coordinates, sheet_dimensions):
    normalized_cell_coords = {}
    for cell_ref, (col, row) in cell_coordinates.items():
        normalized_col = normalize_coordinates(col, sheet_dimensions['width'])
        normalized_row = normalize_coordinates(row, sheet_dimensions['height'])
        normalized_cell_coords[cell_ref] = (normalized_col, normalized_row)
    return normalized_cell_coords

def denormalize_coordinates(norm_value, max_value):
    return norm_value * (max_value )

# def denormalize_cell_coordinates(normalized_coords, sheet_dimensions):
#     denormalized_coords = {}
#     for cell_ref, (norm_col, norm_row) in normalized_coords.items():
#         col = denormalize_coordinates(norm_col, sheet_dimensions['width'])
#         row = denormalize_coordinates(norm_row, sheet_dimensions['height'])
#         denormalized_coords[cell_ref] = (int(col), int(row))
#     return denormalized_coords
def denormalize_cell_coordinates(normalized_coords, table_width, table_height):
    denormalized_coords = {}
    for cell_ref, (norm_col, norm_row) in normalized_coords.items():
        col = denormalize_coordinates(norm_col, table_width) 
        row = denormalize_coordinates(norm_row, table_height)
        denormalized_coords[cell_ref] = (int(col)-1, int(row)-1)
    return denormalized_coords



def find_closest_cell(normalized_center_x, normalized_center_y, normalized_cell_coordinates):
    closest_cell = None
    min_distance = float('inf')
    
    for cell_ref, (normalized_col, normalized_row) in normalized_cell_coordinates.items():
        distance = np.sqrt((normalized_center_x - normalized_col) ** 2 + (normalized_center_y - normalized_row) ** 2)
        
        if distance < min_distance:
            min_distance = distance
            closest_cell = cell_ref
    
    return closest_cell


# def find_closest_cell(normalized_center_x, normalized_center_y, normalized_cell_coordinates):
#     closest_cell = None
#     min_distance = float('inf')
    
#     for cell_ref, (normalized_col, normalized_row) in normalized_cell_coordinates.items():
#         distance = np.sqrt((normalized_center_x - normalized_col) ** 2 + (normalized_center_y - normalized_row) ** 2)
        
#         if distance < min_distance:
#             min_distance = distance
#             closest_cell = cell_ref
    
#     return closest_cell


# def find_closest_cell(x, y, cell_coords):
#     min_distance = float('inf')
#     closest_cell = None
#     for cell in cell_coords:
#         distance = np.sqrt((x - cell['center_x'])**2 + (y - cell['center_y'])**2)
#         if distance < min_distance:
#             min_distance = distance
#             closest_cell = cell
#     return closest_cell['cell_ref']


# def find_next_empty_cell(ws, start_cell_ref):
#     column = start_cell_ref[0]
#     start_row = int(start_cell_ref[1:])
    
#     current_row = start_row
#     while ws[f"{column}{current_row}"].value is not None:
#         current_row += 1
    
#     return f"{column}{current_row}"

# def find_previous_cell(ws, start_cell_ref):
#     column = start_cell_ref[0]
#     start_row = int(start_cell_ref[1:])

#     # Check the very previous cell only
#     if start_row > 1:
#         next_row = start_row - 1

#         return f"{column}{next_row}"
#     else:
#         return f"{column}{start_row}"



# def find_next_cell(ws, start_cell_ref):
#     column = start_cell_ref[0]
#     start_row = int(start_cell_ref[1:])

#     # Check the very next cell only
#     next_row = start_row + 1

#     return f"{column}{next_row}"


# def find_next_cell(cell_ref, denormalized_cell_coords):
#     col = ''.join(filter(str.isalpha, cell_ref))
#     row = int(''.join(filter(str.isdigit, cell_ref)))
    
#     next_col = col
#     next_row = row + 1
    
#     next_cell_ref = f"{next_col}{next_row}"
#     if next_cell_ref in denormalized_cell_coords:
#         return next_cell_ref
    
#     # If the next cell in the column does not exist, move to the next column
#     next_col_index = ord(col) - ord('A') + 1
#     next_col = chr(ord('A') + next_col_index)
#     next_row = 1
#     next_cell_ref = f"{next_col}{next_row}"
    
#     if next_cell_ref in denormalized_cell_coords:
#         return next_cell_ref
    
#     return None


def find_next_cell(ws, cell_ref):
    """
    Find the next empty cell in the same column.
    """
    col_letter = ''.join([char for char in cell_ref if char.isalpha()])
    row_number = int(''.join([char for char in cell_ref if char.isdigit()]))
    while ws[f'{col_letter}{row_number}'].value is not None:
        row_number += 1
    return f'{col_letter}{row_number}'


# def find_adjacent_empty_cell(ws, start_cell_ref):
#     column = start_cell_ref[0]
#     start_row = int(start_cell_ref[1:])

#     # Check the cell directly below
#     current_row = start_row
#     if ws[f"{column}{current_row}"].value is not None:
#         current_row += 1

#     # Check the cell directly above
#     if current_row > 0 and ws[f"{column}{current_row}"].value is not None:
#         current_row -= 1
    
#     return f"{column}{current_row}"

    # # If both adjacent cells are occupied or out of bounds, return None
    # elseif: 
    #     return f"{column}{start_row}


# def transcription(detected_table_cells, ocr_model):
#     '''
#     # Makes use of a pre-processed image (in grayscale) to detect the table from the record sheets

#     Parameters
#     --------------
#     detected_table_cells where: 
#         detected_table_cells[0]: contours. Contours for the detected text in the table cells
#         detected_table_cells[1]: image_with_all_bounding_boxes. Bounding boxex for each cell for which clips will be made later before optical character recognition 
#         detected_table_cells[2]: table_copy
#         detected_table_cells[3]: table_original_image
    

#     ocr_model: Optical Character Recognition/Handwritten Text Recognition of choice
    
#     Returns
#     -------------- 
#     wb : Ms Excel workbook with OCR Results
#     ''' 
#     if ocr_model == 'Tesseract-OCR':
#         ## Lauching Tesseract-OCR
#         pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' ## Here input the PATH to the Tesseract executable on your computer. See more information here: https://pypi.org/project/pytesseract/
#     # if ocr_model == 'PaddleOCR':
#     #     ## Lauching PaddleOCR, which would be used by downloading necessary files as shown below
#     #     paddle_ocr = PaddleOCR(use_angle_cls=True, lang = 'en', use_gpu=False) ## Run only once to download all required files
#     if ocr_model == 'EasyOCR':
#         ## Lauching EasyOCR
#         easyocr_reader = easyocr.Reader(['en']) # this needs to run only once to load the model into memory

#     contours = detected_table_cells[0]
#     image_with_all_bounding_boxes = detected_table_cells[1]
#     table_copy = detected_table_cells[2]
#     table_original_image = detected_table_cells[3]

#     # Get the dimensions of the loaded image
#     image_height, image_width, image_channels = image_with_all_bounding_boxes.shape

#     results = []

#     organize_methods = {
#         #'Bottom': organize_contours_bottom,
#         'Midpoint': organize_contours,
#         'Top': organize_contours_top
#     }

#     min_x, min_y, max_x, max_y = get_table_boundaries(contours)
#     table_width = max_x - min_x
#     table_height = max_y - min_y

#     for method_name, organize_method in organize_methods.items():

#         ## Create an Excel workbook and add a worksheet where the transcribed text will be saved
#         wb = Workbook()
#         ws = wb.active
#         ws.title = 'OCR_Results'

#         # Sort contours by y-coordinate
#         contours_sorted = sorted(contours, key=lambda c: cv2.boundingRect(c)[1])

#         max_rows = 43  # maximum rows, adjust based on your table's expected structure. Here, I had started with 45 rows and it was already giving good results . Previously had it at 43
#         organized_rows = organize_method(contours_sorted, (image_height, image_width, image_channels), max_rows)
#         # organized_rows = organize_and_merge_contours(contours_sorted, (image_height, image_width, image_channels), max_rows)


#         # # Define the row index to start filling from
#         # start_row_index = 1  # Change this to your desired starting row index

#         # Sort contours by y-coordinate
#         # contours_sorted = sorted(contours, key=lambda c: cv2.boundingRect(c)[1])

#         # Calculate the median y-coordinate for each row and sort rows by this median
#         sorted_rows = sorted(organized_rows, key=lambda row: np.median([cv2.boundingRect(c)[1] + cv2.boundingRect(c)[3] // 2 for c in row]))

#         for row_index, row in enumerate(sorted_rows, start=1):
#             for contour in row:
#         ## Text detection using an OCR model; Here using TesseractOCR
#         # for contour in contours_sorted:

#         #for contour in contours:
#                 x, y, w, h = cv2.boundingRect(contour)
                
#                 # Adjust these threshold values according to your requirements
#                 min_width_threshold = 20 # 50 previously
#                 min_height_threshold = 28  # Had this at 13 previously. 20, 25, 30. trying 28, but it might be over fitting 

#                 max_width_threshold = 200 # 120 previously
#                 max_height_threshold = 90 # 60 previously

#                 # Filter out smaller bounding boxes
#                 if (min_width_threshold <= w <= max_width_threshold and min_height_threshold <= h <= max_height_threshold):

#                     # Calculate a factor to increase the bounding box area (e.g., 80% larger)
#                     factor_width = 0.05  # Modify this factor as needed
#                     increase_factor_height = 0.25 # Modify this factor as needed
                
#                     x += int(w * factor_width)  # Increase width
#                     y -= int(h * increase_factor_height)  # Increase height
#                     w -= int(1 * w * factor_width)  # Decrease width a little to avoid vertical lines that may be transcribed as the number 1 yet they aren't a number
#                     h += int(2 * h * increase_factor_height)  # Increase height
                    
#                     # This line below is about
#                     # drawing a rectangle on the image with the shape of
#                     # the bounding box. Its not needed for the OCR.
#                     # Its just added for debugging purposes.
#                     image_with_all_bounding_boxes = cv2.rectangle(image_with_all_bounding_boxes, (x, y), (x + w, y + h), (0, 255, 0), 5)
                    
#                     # Calculate center coordinates of the bounding box
#                     center_x = x + w // 2
#                     center_y = y + h // 2
                    
#                     # OCR
#                     # Crop each cell using the bounding rectangle coordinates
#                     ROI = table_copy[y:y+h, x:x+w]
                    
#                     if ROI.size != 0:  # Check if the height and width are greater than zero
                        
#                         cv2.imwrite('detected.png', ROI)
#                         if ocr_model == 'Tesseract-OCR':
#                         # Using Tesseract-OCR
#                             ocr_result = pytesseract.image_to_string('detected.png', lang='cobedore-V6', config='--psm 7 -c tessedit_char_whitelist=0123456789.') # Just added -c tessedit_char_whitelist=0123456789. to really limit the text detected
#                             #ocr_result = pytesseract.image_to_string('detected.png', lang='ccc-base+eng+cobedore-V6', config='--psm 7 -c tessedit_char_whitelist=0123456789.')

#                             # Here's a brief explanation of some Page Segmentation Modes (PSMs) available in Tesseract:
#                             # 0: Orientation and script detection (OSD) only.
#                             # 1: Automatic page segmentation with OSD.
#                             # 2: Automatic page segmentation, but no OSD, or OCR.
#                             # 3: Fully automatic page segmentation, but no OSD. (Default)
#                             # 4: Assume a single column of text of variable sizes.
#                             # 5: Assume a single uniform block of vertically aligned text.
#                             # 6: Assume a single uniform block of text.
#                             # 7: Treat the image as a single text line.
#                             # 8: Treat the image as a single word.
#                             # 9: Treat the image as a single word in a circle.
#                             # 10: Treat the image as a single character.
#                             # 11: Sparse text. Find as much text as possible in no particular order.
#                             # 12: Sparse text with OSD.
#                             # 13: Raw line. Treat the image as a single text line, bypassing hacks that are Tesseract-specific.


#                         # if ocr_model == 'PaddleOCR':
#                         #     ## Using PaddleOCR
#                         #     ocr_result = paddle_ocr.ocr('detected.png', cls = True)

#                         if ocr_model == 'EasyOCR':
#                         # Using EasyOCR
#                             ocr_result = easyocr_reader.readtext('detected.png', detail = 0, allowlist='0123456789')
#                         # Using OCR to recognize text/transcription
#                         if ocr_result is not None:
                            
#                             # Maximum number of columns and rows. These can be changed depending on the tables in the images
#                             max_column_index = 24  # Number of columns in the table. Total number is original unclipped image are 27
#                             max_row_index = 43  # Estimated number of rows in the table  .Previosly had it at 57 and results were good.       even 56 was good.     had this previously at 43
                            
#                             cell_width = max(image_width // max_column_index, min_width_threshold)
#                             cell_height = max(image_height//max_row_index, min_height_threshold)
                            
#                             # Track filled cells using a set
#                             filled_cells = []
                            
                            
#                             # Convert the x-coordinate to a column letter # Assuming x-coordinate translates to columns cell_column
#                             # Ensure x is within the valid range for Excel column indices
                            
#                             if 1 <= center_x <= image_width:  # Excel's maximum column index
#                                 column_index = int(max((center_x) // cell_width, 0)) + 1 # Ensure column index is at least 1
#                                 #cell_column = openpyxl.utils.get_column_letter(min(column_index, max_column_index))
#                             else:
#                                 column_index = 1
#                                 #cell_column = 'A'  # Set a default column if x is out of range
                            
#                             # if 1 <= center_y <= image_height:
#                             #     # Calculate the row index based on the row ratio
#                             #     row_index = (center_y / cell_height) + 1  # Calculate row index as a floating-point number
                                
#                             #     # Round the row index to the nearest integer
#                             #     cell_row = int(round(row_index))
                                
#                             #     # Ensure the row index is within the valid range
#                             #     cell_row = min(max(cell_row, 1), max_row_index)
#                             #     # row_ratio = (center_y) // cell_height
#                             #     # cell_row = min(int(row_ratio) + 1, max_row_index)  # Convert row ratio to integer and ensure it's within valid range
#                             #     # #cell_row = min(center_y // cell_height + 0.5, max_row_index)
#                             # else:
#                             #     cell_row = 1  # Set a default row if y is out of range  # Assuming y-coordinate translates to rows



#                             # Ms Excel Template cell coordinates
#                             file_path = f'docs\Table_structure.xlsx'
#                             sheet_name = 'clipped_tilted'
#                             cell_coordinates = get_excel_cell_coordinates(file_path, sheet_name)
#                             sheet_dimensions = {'width': max_column_index, 'height': max_row_index}  # Assuming max column index and max row index

#                             normalized_cell_coords = normalize_cell_coordinates(cell_coordinates, sheet_dimensions)


#                             # Find closest cell in template
#                             # Normalize bounding box center coordinates
#                             # normalized_center_x = normalize_coordinates(center_x, image_width)
#                             # normalized_center_y = normalize_coordinates(center_y, image_height)
#                             normalized_center_x = normalize_coordinates(center_x, image_width)
#                             normalized_center_y = normalize_coordinates(center_y, table_height)

#                             closest_cell = find_closest_cell(normalized_center_x, normalized_center_y, normalized_cell_coords)
            

#                             # if closest_cell:
#                             #     cell = ws[closest_cell]
#                             #     cell.value = ocr_result

#                             if closest_cell:
#                             # Check if the cell is occupied and find the next empty cell in the same column
#                                 cell_ref = closest_cell
#                                 while ws[cell_ref].value is not None:
#                                     cell_ref = find_next_cell(ws, cell_ref)
                                
#                                 cell = ws[cell_ref]
#                                 cell.value = ocr_result








#                             # # Write the OCR value to the cell in the Excel sheet
#                             # cell = ws.cell(row=row_index, column=column_index)


#                             # # # Set the cell value to the OCR result
#                             # # cell.value = ocr_result

#                             # current_cell_value = cell.value

#                             # # If the cell is empty, add the OCR result directly
#                             # if current_cell_value is None:
#                             #     cell.value = ocr_result
#                             # else:
#                             #     # If the cell already has content, append the new result
#                             #     cell.value = f"{current_cell_value}{ocr_result}"

#                             # Set up border styles for excel output
#                             thin_border = Border(
#                                 left=Side(style='thin'),
#                                 right=Side(style='thin'),
#                                 top=Side(style='thin'),
#                                 bottom=Side(style='thin'))

#                             # Loop through cells to apply borders
#                             for row in ws.iter_rows(min_row=1, max_row=max_row_index, min_col=1, max_col=max_column_index):
#                                 for cell in row:
#                                     cell.border = thin_border
                            
#                         else:
#                             print('No values detected in clip')
#                     else:
#                         print('ROI is empty or invalid')


#         # Insert two columns on the left side of the excel sheet
#         ws.insert_cols(1, 2)
        
#         # Insert a new row at the top for headers
#         ws.insert_rows(1, amount = 3)

#         # # Merging cells for headers
#         # # Merge cells for headers that span multiple columns
#         # ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8)  # For "Températures extrêmes"
#         # #*** Do this for the other cells that need to be merged


#         # Define your headers (adjust as needed)
#         headers = ["No de la pentade", "Date", "Bellani (gr. Cal/cm2) 6-6h", "Températures extrêmes", "", "", "", "", "Evaportation en cm3 6 - 6h", "", "Pluies en mm. 6-6h", "Température et Humidité de l'air à 6 heures", "", "", "", "", "Température et Humidité de l'air à 15 heures",  "", "", "", "", "Température et Humidité de l'air à 18 heures",  "", "", "", "", "Date"]
#         sub_headers_1 = ["", "", "", "Abri", "", "", "", "", "Piche", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", ""]
#         sub_headers_2 =["", "", "", "Max.", "Min.", "(M+m)/2", "Ampl.", "Min. gazon", "Abri.", "Ext.", "", "T", "T'a", "e.", "U", "∆e", "T", "T'a", "e.", "U", "∆e","T", "T'a", "e.", "U", "∆e", ""]

#         # Add the headers to the first row
#         for col_num, header in enumerate(headers, start=1):
#             ws.cell(row=1, column=col_num, value=header)
#         # Add the first row of sub-headers to the second row
#         for col_num, sub_header in enumerate(sub_headers_1, start=1):
#             ws.cell(row=2, column=col_num, value=sub_header)

#         # Add the second row of sub-headers to the third row
#         for col_num, sub_header in enumerate(sub_headers_2, start=1):
#             ws.cell(row=3, column=col_num, value=sub_header)

#         file_path = f'src\output\{method_name}_Excel_with_OCR_Results.xlsx'
#         wb.save(file_path)
#         results.append([file_path])

#         #wb.save(f'{method_name}_Excel_with_OCR_Results.xlsx') 

#         # plt.imshow(image_with_all_bounding_boxes)
#         # plt.show()

#     return results




def detect_lines(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, 30, 100, apertureSize=3)
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold=100, minLineLength=100, maxLineGap=10)
    
    vertical_lines = set()
    horizontal_lines = set()
    if lines is not None:
        for line in lines:
            x1, y1, x2, y2 = line[0]
            if x1 == x2:  # Vertical line
                vertical_lines.add(x1)
            elif y1 == y2:  # Horizontal line
                horizontal_lines.add(y1)
    
    return sorted(vertical_lines), sorted(horizontal_lines)


def calculate_intersection_area(rect1, rect2):
    x1, y1, w1, h1 = rect1
    x2, y2, w2, h2 = rect2

    # Calculate the overlap coordinates
    x_overlap = max(0, min(x1 + w1, x2 + w2) - max(x1, x2))
    y_overlap = max(0, min(y1 + h1, y2 + h2) - max(y1, y2))

    # Calculate the area of overlap
    overlap_area = x_overlap * y_overlap

    return overlap_area

def get_closest_contour(contours, x, y, w, h):
    min_distance = float('inf')
    closest_contour = None
    contour_rect = (x, y, w, h)

    for contour in contours:
        contour_x, contour_y, contour_w, contour_h = cv2.boundingRect(contour)
        filtered_contour_rect = (contour_x, contour_y, contour_w, contour_h)
        overlap_area = calculate_intersection_area(contour_rect, filtered_contour_rect)
        
        if overlap_area >= (contour_w * contour_h) / 3: # If ateleast 1/3 of the are overlaps
            # Calculate the distance between the centroids of the two rectangles
            contour_center = (contour_x + contour_w / 2, contour_y + contour_h / 2)
            rect_center = (x + w / 2, y + h / 2)
            distance = np.sqrt((contour_center[0] - rect_center[0]) ** 2 + (contour_center[1] - rect_center[1]) ** 2)
            
            if distance < min_distance:
                min_distance = distance
                closest_contour = contour

    return closest_contour

def calculate_cell_reference(center_x, center_y, max_rows, max_columns, table_width, table_height):
    row = int(center_y / table_height * max_rows) + 1
    column = int(center_x / table_width * max_columns) + 1
    return f'{openpyxl.utils.get_column_letter(column)}{row}'

# # This compares the two contours with a point (x,y)
# def get_closest_contour(contours, x, y, w, h):
#     # Function to get the closest contour to the given ROI
#     min_distance = float('inf')
#     closest_contour = None
#     for contour in contours:
#         px, py, pw, ph = cv2.boundingRect(contour)
#         distance = np.sqrt((px - x) ** 2 + (py - y) ** 2)
#         if distance < min_distance:
#             min_distance = distance
#             closest_contour = contour
#     return closest_contour

def transcription(detected_table_cells, ocr_model):

    
    wb = Workbook()
    ws = wb.active
    ws.title = 'OCR_Results'

    if ocr_model == 'Tesseract-OCR':
        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    elif ocr_model == 'EasyOCR':
        easyocr_reader = easyocr.Reader(['en'])

    contours_for_detected_text = detected_table_cells[0]
    image_with_all_bounding_boxes = detected_table_cells[1]
    table_copy = detected_table_cells[2]
    table_original_image = detected_table_cells[3]

    image_height, image_width, _ = image_with_all_bounding_boxes.shape

    min_x, min_y, max_x, max_y = get_table_boundaries(contours_for_detected_text)
    table_width = max_x - min_x
    table_height = max_y - min_y

    # min_x, min_y, max_x, max_y = 0, 0, image_width, image_height
    # table_width = max_x - min_x
    # table_height = max_y - min_y

    # Load and normalize cell coordinates from the template
    file_path = 'docs/Table_structure.xlsx'
    sheet_name = 'clipped_tilted'
    cell_coordinates = get_excel_cell_coordinates(file_path, sheet_name)
    sheet_dimensions = {'width': 24, 'height': 55}  # Based on the template
    normalized_cell_coords = normalize_cell_coordinates(cell_coordinates, sheet_dimensions)

    # Denormalize cell coordinates to match the table image dimensions
    denormalized_cell_coords = denormalize_cell_coordinates(normalized_cell_coords, table_width, table_height)

    # Rows to skip
    rows_to_skip = [6, 9, 15, 18, 24, 27, 40, 43, 50, 53]

    # Create a new dictionary excluding the rows to skip
    filtered_cell_coords = {
        cell_ref: (x, y)
        for cell_ref, (x, y) in denormalized_cell_coords.items()
        if openpyxl.utils.coordinate_to_tuple(cell_ref)[0] not in rows_to_skip
    }

    contours = []
    for cell_ref, (x, y) in filtered_cell_coords.items():

        cell_width = table_width / sheet_dimensions['width']
        cell_height = table_height / sheet_dimensions['height']

        # Store the bounding box as a contour (list of points)
        contour = [(x, y + 25), (x + int(cell_width), y + 25), (x + int(cell_width), y + 25 + int(cell_height)), (x, y + 25 + int(cell_height))]
        contours.append(contour)

        # Example manipulation: Draw the contour as a rectangle
        cv2.drawContours(image_with_all_bounding_boxes, [np.array(contour)], 0, (0, 255, 0), 2)

    # Print the image with the created bounding boxes from the template. These are not necessary covering the entire cell but are pretty close. Therefore this is strictly for illustration and checking purposes
    plt.imshow(cv2.cvtColor(image_with_all_bounding_boxes, cv2.COLOR_BGR2RGB))
    plt.show()

    # Filtering contours_for_detected_text based on size thresholds
    min_width_threshold = 20
    min_height_threshold = 28
    max_width_threshold = 200
    max_height_threshold = 90

    filtered_contours_for_detected_text = [
        contour for contour in contours_for_detected_text
        if min_width_threshold <= cv2.boundingRect(contour)[2] <= max_width_threshold and
           min_height_threshold <= cv2.boundingRect(contour)[3] <= max_height_threshold
    ]

    # Create a copy of table_copy to draw the ROIs
    ROIs_image = table_copy.copy()
    for contour in contours:
        # Extract coordinates for Region of Interest (ROI)
        x, y = contour[0]
        x2, y2 = contour[2]
        w = x2 - x
        h = y2 - y

        # Adjust ROI boundaries based on the closest detected text contour
        closest_contour = get_closest_contour(filtered_contours_for_detected_text, x, y, w, h)
        if closest_contour is not None:
            cx, cy, cw, ch = cv2.boundingRect(closest_contour)

            # Adjust the ROI boundaries
            x = max(x, cx)
            y = max(y, cy)
            w = min(w, cw + x)
            h = min(h, ch + y)

        # Calculate the center of the adjusted ROI
        center_x = x + w / 2
        center_y = y + h / 2

        # Draw the adjusted ROI on the output image
        cv2.rectangle(ROIs_image, (x, y-10), (x + w, y-10 + h+10), (255, 0, 0), 2)  # Red color for ROI

        ROI = table_copy[y-10:y+h+5, x:x +w]  # Use calculated cell dimensions
        
        if ROI.size != 0:
            cv2.imwrite('detected.png', ROI)
            if ocr_model == 'Tesseract-OCR':
                ocr_result = pytesseract.image_to_string('detected.png', lang='cobedore-V6', config='--psm 7 -c tessedit_char_whitelist=0123456789.') # Just added -c tessedit_char_whitelist=0123456789. to really limit the text detected

                # Calculate the cell reference
                cell_ref = calculate_cell_reference(center_x, center_y, max_rows=43, max_columns=24, table_width=table_width, table_height=table_height)
                # Write the OCR result to the Excel cell
                ws[cell_ref].value = ocr_result

    
    # Display the output image with all the drawn ROIs
    plt.imshow(cv2.cvtColor(ROIs_image, cv2.COLOR_BGR2RGB))
    plt.show()




    # wb = Workbook()
    # ws = wb.active
    # ws.title = 'OCR_Results'


    # # gray = cv2.cvtColor(image_with_all_bounding_boxes, cv2.COLOR_BGR2GRAY)

    # # # Edge detection
    # # edges = cv2.Canny(gray, 50, 150, apertureSize=3)

    # # # Hough Line Transform
    # # lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=100, minLineLength=100, maxLineGap=10)

    # # # Initialize variables to store the coordinates of the first vertical and horizontal lines
    # # first_vertical_line_x = None
    # # first_horizontal_line_y = None

    # # # Identify the first vertical and horizontal lines
    # # for line in lines:
    # #     x1, y1, x2, y2 = line[0]
    # #     if x1 == x2:  # Vertical line
    # #         if first_vertical_line_x is None or x1 < first_vertical_line_x:
    # #             first_vertical_line_x = x1
    # #     elif y1 == y2:  # Horizontal line
    # #         if first_horizontal_line_y is None or y1 < first_horizontal_line_y:
    # #             first_horizontal_line_y = y1

    # # # If lines are not found, set them to zero
    # # if first_vertical_line_x is None:
    # #     first_vertical_line_x = 0
    # # if first_horizontal_line_y is None:
    # #     first_horizontal_line_y = 0
    
    # # # # Detect vertical and horizontal lines in the image
    # # # vertical_lines, horizontal_lines = detect_lines(image_with_all_bounding_boxes)

    # # # # Define the height of each cell in the image
    # # # cell_height = 80

    # # # Continue with your original code to draw bounding boxes
    # # for cell_ref, (x, y) in denormalized_cell_coords.items():
    # #     col, row = openpyxl.utils.coordinate_to_tuple(cell_ref)
    # #     # # Convert to zero-based indexing
    # #     # col -= 1
    # #     # row -= 1

    # #     # Adjust x and y by the first vertical and horizontal line offsets
    # #     adjusted_x = x - first_vertical_line_x
    # #     adjusted_y = y - first_horizontal_line_y

    # #     # Calculate the width and height of the cell correctly
    # #     if col < sheet_dimensions['width'] - 1:
    # #         next_col = col + 1
    # #         next_x = denormalized_cell_coords.get(openpyxl.utils.get_column_letter(next_col + 1) + str(row + 1), (table_width, y))[0]
    # #         cell_width = next_x - x
    # #     else:
    # #         cell_width = table_width / sheet_dimensions['width']

    # #     if row < sheet_dimensions['height'] - 1:
    # #         next_row = row + 1
    # #         next_y = denormalized_cell_coords.get(openpyxl.utils.get_column_letter(col + 1) + str(next_row + 1), (x, table_height))[1]
    # #         cell_height = next_y - y
    # #     else:
    # #         cell_height = table_height / sheet_dimensions['height']

    # #     # Draw bounding boxes on the table image based on the adjusted coordinates
    # #     cv2.rectangle(image_with_all_bounding_boxes, (adjusted_x, adjusted_y), 
    # #                 (adjusted_x + int(cell_width), adjusted_y + int(cell_height)), 
    # #                 (0, 255, 0), 2)




    # # WORKING CODE
    # # List to store contours
    # contours = []
    # # First loop to create the bounding boxes and store as contours
    # # Rows to skip
    # rows_to_skip = [6, 9, 15, 18, 24, 27, 40, 43, 50, 53]  
    # for cell_ref, (x, y) in denormalized_cell_coords.items():
    #     row, col = openpyxl.utils.coordinate_to_tuple(cell_ref)
    #     # Skip specified rows (this is because in the template it is always empty in these rows)
    #     if row in rows_to_skip:
    #         continue

    #     # Calculate the width and height of the cell correctly
    #     # Convert to zero-based indexing
    #     # col -= 1
    #     # row -= 1

    #     cell_width = table_width / sheet_dimensions['width']
    #     cell_height = table_height / sheet_dimensions['height']


    #     # # Calculate the width and height of the cell correctly
    #     # if col < sheet_dimensions['width'] - 1:
    #     #     next_col = col
    #     #     next_x = denormalized_cell_coords.get(openpyxl.utils.get_column_letter(next_col+1) + str(row), (table_width, y))[0]
    #     #     cell_width = next_x - x
    #     # else:
    #     #     cell_width = table_width / sheet_dimensions['width']

    #     # if row < sheet_dimensions['height'] - 1:
    #     #     next_row = row
    #     #     next_y = denormalized_cell_coords.get(openpyxl.utils.get_column_letter(col+1) + str(next_row+1), (x, table_height))[1]
    #     #     cell_height = next_y - y
    #     # else:
    #     #     cell_height = table_height / sheet_dimensions['height']

    #     # Store the bounding box as a contour (list of points)
    #     contour = [(x, y + 25), (x + int(cell_width), y + 25), (x + int(cell_width), y + 25 + int(cell_height)), (x, y + 25 + int(cell_height))]
    #     contours.append((contour, row))

    #     # Example manipulation: Draw the contour as a rectangle
    #     cv2.drawContours(image_with_all_bounding_boxes, [np.array(contour)], 0, (0, 255, 0), 2)


    #     # # Draw bounding boxes on the table image based on the denormalized coordinates
    #     # cv2.rectangle(image_with_all_bounding_boxes, (x, y+20), (x + int(cell_width), y+20 + int(cell_height)), (0, 255, 0), 2)

    # ### Not being used at the moment
    # # for coord, (col, row) in denormalized_cell_coords.items():
    # #     if col >= len(vertical_lines) - 1 or row >= len(horizontal_lines) - 1:
    # #         continue  # Skip if there are not enough lines to form a box
        
    # #     # Find the closest vertical and horizontal lines to align the bounding box
    # #     x1 = vertical_lines[col]
    # #     x2 = vertical_lines[col + 1] if col + 1 < len(vertical_lines) else image_with_all_bounding_boxes.shape[1]
    # #     y1 = horizontal_lines[row]
    # #     y2 = horizontal_lines[row + 1] if row + 1 < len(horizontal_lines) else image_with_all_bounding_boxes.shape[0]
        
    # #     # Draw the rectangle on the image
    # #     cv2.rectangle(image_with_all_bounding_boxes, (x1, y1), (x2, y2), (0, 255, 0), 2)

    # # # Second loop: Manipulate contours and draw them
    # # # Rows to skip
    # # rows_to_skip = [6, 9, 15, 18, 24, 27, 40, 43, 50, 53]  # Zero-based indices for 5th, 6th, and 9th rows
    # # for contour, row in contours:
    # #     # Skip specified rows (this is because in the template it is always empty in these rows)
    # #     if row in rows_to_skip:
    # #         continue

    # #     # Example manipulation: Draw the contour as a rectangle
    # #     cv2.drawContours(image_with_all_bounding_boxes, [np.array(contour)], 0, (0, 255, 0), 2)

    # #     # Extract coordinates for ROI
    # #     x, y = contour[0]
    # #     x2, y2 = contour[2]
    # #     cell_width = x2 - x
    # #     cell_height = y2 - y 

    # for contour in contours_for_detected_text:
    #     x, y, w, h = cv2.boundingRect(contour)
        
    #     # Adjust these threshold values according to your requirements
    #     min_width_threshold = 20 # 50 previously
    #     min_height_threshold = 28  # Had this at 13 previously. 20, 25, 30. trying 28, but it might be over fitting 

    #     max_width_threshold = 200 # 120 previously
    #     max_height_threshold = 90 # 60 previously

    #     # Filter out smaller bounding boxes
    #     if (min_width_threshold <= w <= max_width_threshold and min_height_threshold <= h <= max_height_threshold):

    #         # Calculate a factor to increase the bounding box area (e.g., 80% larger)
    #         factor_width = 0.05  # Modify this factor as needed
    #         increase_factor_height = 0.25 # Modify this factor as needed
        
    #         x += int(w * factor_width)  # Increase width
    #         y -= int(h * increase_factor_height)  # Increase height
    #         w -= int(1 * w * factor_width)  # Decrease width a little to avoid vertical lines that may be transcribed as the number 1 yet they aren't a number
    #         h += int(2 * h * increase_factor_height)  # Increase height
            
    #         # This line below is about
    #         # drawing a rectangle on the image with the shape of
    #         # the bounding box. Its not needed for the OCR.
    #         # Its just added for debugging purposes.
    #         image_with_all_bounding_boxes = cv2.rectangle(image_with_all_bounding_boxes, (x, y), (x + w, y + h), (0, 255, 0), 5)
            
    #         # Calculate center coordinates of the bounding box
    #         center_x = x + w // 2
    #         center_y = y + h // 2
            
    #         # OCR

    #         # Extract the ROI from the table image
    #         # ROI = table_copy[y:y+25 + int(cell_height), x+10:x-5 + int(cell_width)]
    #         ROI = image_with_all_bounding_boxes[y:y+h, x:x +w]  # Use calculated cell dimensions
        
    #         if ROI.size != 0:
    #             cv2.imwrite('detected.png', ROI)
    #             if ocr_model == 'Tesseract-OCR':
    #                 ocr_result = pytesseract.image_to_string('detected.png', lang='cobedore-V6', config='--psm 7 -c tessedit_char_whitelist=0123456789.') # Just added -c tessedit_char_whitelist=0123456789. to really limit the text detected

    #         #     # if ocr_result:
    #         #     #     ws[cell_ref] = ocr_result
                
    #         #     if ocr_result:
    #         #         col = openpyxl.utils.get_column_letter((x // cell_width) + 1)
    #         #         row = (y // cell_height) + 1
    #         #         cell_ref = f'{col}{row}'
    #         #         ws[cell_ref] = ocr_result

    #         #Ms Excel Template cell coordinates
    #         file_path = f'docs\Table_structure.xlsx'
    #         sheet_name = 'clipped_tilted'
    #         cell_coordinates = get_excel_cell_coordinates(file_path, sheet_name)
    #         sheet_dimensions = {'width': 24, 'height': 43}  # Assuming max column index and max row index

    #         normalized_cell_coords = normalize_cell_coordinates(cell_coordinates, sheet_dimensions)


    #         # Find closest cell in template
    #         # Normalize bounding box center coordinates
    #         # normalized_center_x = normalize_coordinates(center_x, image_width)
    #         # normalized_center_y = normalize_coordinates(center_y, image_height)
    #         normalized_center_x = normalize_coordinates(x, image_width)
    #         normalized_center_y = normalize_coordinates(y, image_height)

    #         closest_cell = find_closest_cell(normalized_center_x, normalized_center_y, normalized_cell_coords)


    #         # if closest_cell:
    #         #     cell = ws[closest_cell]
    #         #     cell.value = ocr_result

    #         if closest_cell:
    #         # Check if the cell is occupied and find the next empty cell in the same column
    #             cell_ref = closest_cell
    #             while ws[cell_ref].value is not None:
    #                 cell_ref = find_next_cell(ws, cell_ref)
                
    #             cell = ws[cell_ref]
    #             cell.value = ocr_result

        
    #         # Plotting bounding boxes on the image
    #         cv2.rectangle(table_copy, (x+10, y), (x2-5, y2+25), (255, 0, 0), 2)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=43, min_col=1, max_col=24):
        for cell in row:
            cell.border = thin_border

    # Insert two columns on the left side of the excel sheet
    ws.insert_cols(1, 2)
    
    # Insert a new row at the top for headers
    ws.insert_rows(1, amount=3)

    # Define your headers (adjust as needed)
    headers = ["No de la pentade", "Date", "Bellani (gr. Cal/cm2) 6-6h", "Températures extrêmes", "", "", "", "", "Evaporation en cm3 6 - 6h", "", "Pluies en mm. 6-6h", "Température et Humidité de l'air à 6 heures", "", "", "", "", "Température et Humidité de l'air à 15 heures", "", "", "", "", "Température et Humidité de l'air à 18 heures", "", "", "", "", "Date"]
    sub_headers_1 = ["", "", "", "Abri", "", "", "", "", "Piche", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", ""]
    sub_headers_2 = ["", "", "", "Max.", "Min.", "(M+m)/2", "Ampl.", "Min. gazon", "Abri.", "Ext.", "", "T", "T'a", "e.", "U", "∆e", "T", "T'a", "e.", "U", "∆e", "T", "T'a", "e.", "U", "∆e", ""]

    # Add the headers to the first row
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Add the first row of sub-headers to the second row
    for col_num, sub_header in enumerate(sub_headers_1, start=1):
        ws.cell(row=2, column=col_num, value=sub_header)

    # Add the second row of sub-headers to the third row
    for col_num, sub_header in enumerate(sub_headers_2, start=1):
        ws.cell(row=3, column=col_num, value=sub_header)

    # Ensure the output directory exists
    output_dir = 'src/output'
    os.makedirs(output_dir, exist_ok=True)
    
    file_path = os.path.join(output_dir, 'newtrial_Excel_with_OCR_Results.xlsx')
    wb.save(file_path)

    # # Print the image with the bounding boxes
    # plt.imshow(cv2.cvtColor(image_with_all_bounding_boxes, cv2.COLOR_BGR2RGB))
    # plt.show()

    # Save the Excel workbook
    wb.save('OCR_Results.xlsx')

    return contours






# def transcription(detected_table_cells, ocr_model):
#     '''
#     Makes use of a pre-processed image (in grayscale) to detect the table from the record sheets

#     Parameters
#     --------------
#     detected_table_cells where: 
#         detected_table_cells[0]: contours. Contours for the detected text in the table cells
#         detected_table_cells[1]: image_with_all_bounding_boxes. Bounding boxes for each cell for which clips will be made later before optical character recognition 
#         detected_table_cells[2]: table_copy
#         detected_table_cells[3]: table_original_image

#     ocr_model: Optical Character Recognition/Handwritten Text Recognition of choice
    
#     Returns
#     -------------- 
#     wb : Ms Excel workbook with OCR Results
#     ''' 
#     if ocr_model == 'Tesseract-OCR':
#         ## Lauching Tesseract-OCR
#         pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' ## Here input the PATH to the Tesseract executable on your computer. See more information here: https://pypi.org/project/pytesseract/
#     elif ocr_model == 'EasyOCR':
#         ## Lauching EasyOCR
#         easyocr_reader = easyocr.Reader(['en']) # this needs to run only once to load the model into memory

#     contours = detected_table_cells[0]
#     image_with_all_bounding_boxes = detected_table_cells[1]
#     table_copy = detected_table_cells[2]
#     table_original_image = detected_table_cells[3]

#     # Get the dimensions of the loaded image
#     image_height, image_width, image_channels = image_with_all_bounding_boxes.shape

#     results = []

#     min_x, min_y, max_x, max_y = get_table_boundaries(contours)
#     table_width = max_x - min_x
#     table_height = max_y - min_y

#     ## Create an Excel workbook and add a worksheet where the transcribed text will be saved
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'OCR_Results'

#     # Load the Excel template to get row heights
#     excel_path = f'docs\Table_structure.xlsx'  # Replace with the path to your Excel template
#     df = pd.read_excel(excel_path, header=None)
#     workbook = load_workbook(excel_path)
#     sheet = workbook['clipped_tilted']

#     # Slice the DataFrame to include only the specified range
#     df = df.iloc[0:43, 0:24]

#     # Get the dimensions of the Excel sheet
#     rows, cols = df.shape

#     # Calculate the cumulative heights of the rows
#     row_heights = []
#     for i in range(1, rows + 1):
#         row_height = sheet.row_dimensions[i].height
#         row_heights.append(row_height if row_height else sheet.default_row_height)

#     # Calculate the cumulative sum of row heights to determine cell centers
#     cumulative_heights = [0] * rows
#     cumulative_heights[0] = row_heights[0]
#     for i in range(1, rows):
#         cumulative_heights[i] = cumulative_heights[i - 1] + row_heights[i]

#     # Compute the expected center coordinates of each cell in the Excel template
#     excel_cells = []
#     cell_width = image_width // cols

#     for i in range(rows):
#         for j in range(cols):
#             center_x = j * cell_width + cell_width // 2
#             if i == 0:
#                 center_y = cumulative_heights[i] // 2
#             else:
#                 center_y = (cumulative_heights[i - 1] + cumulative_heights[i]) // 2
#             cell_ref = f'{openpyxl.utils.get_column_letter(1 + j)}{1 + i}'  # Create cell reference
#             excel_cells.append({'row': i, 'col': j, 'center_x': center_x, 'center_y': center_y, 'cell_ref': cell_ref})


#     for contour in contours:
#         x, y, w, h = cv2.boundingRect(contour)
        
#         # Adjust these threshold values according to your requirements
#         min_width_threshold = 20
#         min_height_threshold = 28
#         max_width_threshold = 200
#         max_height_threshold = 90

#         # Filter out smaller bounding boxes
#         if (min_width_threshold <= w <= max_width_threshold and min_height_threshold <= h <= max_height_threshold):

#             # Calculate a factor to increase the bounding box area (e.g., 80% larger)
#             factor_width = 0.05
#             increase_factor_height = 0.25
        
#             x += int(w * factor_width)
#             y -= int(h * increase_factor_height)
#             w -= int(1 * w * factor_width)
#             h += int(2 * h * increase_factor_height)
            
#             image_with_all_bounding_boxes = cv2.rectangle(image_with_all_bounding_boxes, (x, y), (x + w, y + h), (0, 255, 0), 5)
            
#             # Calculate center coordinates of the bounding box
#             center_x = x + w // 2
#             center_y = y + h // 2
            
#             # OCR
#             ROI = table_copy[y:y+h, x:x+w]
            
#             if ROI.size != 0:  # Check if the height and width are greater than zero
                
#                 cv2.imwrite('detected.png', ROI)
#                 if ocr_model == 'Tesseract-OCR':
#                     ocr_result = pytesseract.image_to_string('detected.png', lang='cobedore-V6', config='--psm 7 -c tessedit_char_whitelist=0123456789.')
#                 elif ocr_model == 'EasyOCR':
#                     ocr_result = easyocr_reader.readtext('detected.png', detail=0, allowlist='0123456789')

#                 if ocr_result is not None:
#                     normalized_center_x = normalize_coordinates(center_x, image_width)
#                     normalized_center_y = normalize_coordinates(center_y, table_height)

#                     closest_cell = find_closest_cell(normalized_center_x, normalized_center_y, excel_cells)

#                     if closest_cell:
#                         cell_ref = closest_cell
#                         while ws[cell_ref].value is not None:
#                             cell_ref = find_next_cell(ws, cell_ref)
                        
#                         cell = ws[cell_ref]
#                         cell.value = ocr_result

#                         thin_border = Border(
#                             left=Side(style='thin'),
#                             right=Side(style='thin'),
#                             top=Side(style='thin'),
#                             bottom=Side(style='thin'))

#                         for row in ws.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=cols):
#                             for cell in row:
#                                 cell.border = thin_border
#                 else:
#                     print('No values detected in clip')
#             else:
#                 print('ROI is empty or invalid')

#     # Insert two columns on the left side of the excel sheet
#     ws.insert_cols(1, 2)
    
#     # Insert a new row at the top for headers
#     ws.insert_rows(1, amount=3)

#     # Define your headers (adjust as needed)
#     headers = ["No de la pentade", "Date", "Bellani (gr. Cal/cm2) 6-6h", "Températures extrêmes", "", "", "", "", "Evaportation en cm3 6 - 6h", "", "Pluies en mm. 6-6h", "Température et Humidité de l'air à 6 heures", "", "", "", "", "Température et Humidité de l'air à 15 heures",  "", "", "", "", "Température et Humidité de l'air à 18 heures",  "", "", "", "", "Date"]
#     sub_headers_1 = ["", "", "", "Abri", "", "", "", "", "Piche", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", "", "(Psychromètre a aspiration)", "", "", "", ""]
#     sub_headers_2 = ["", "", "", "Max.", "Min.", "(M+m)/2", "Ampl.", "Min. gazon", "Abri.", "Ext.", "", "T", "T'a", "e.", "U", "∆e", "T", "T'a", "e.", "U", "∆e","T", "T'a", "e.", "U", "∆e", ""]

#     # Add the headers to the first row
#     for col_num, header in enumerate(headers, start=1):
#         ws.cell(row=1, column=col_num, value=header)
#     # Add the first row of sub-headers to the second row
#     for col_num, sub_header in enumerate(sub_headers_1, start=1):
#         ws.cell(row=2, column=col_num, value=sub_header)

#     # Add the second row of sub-headers to the third row
#     for col_num, sub_header in enumerate(sub_headers_2, start=1):
#         ws.cell(row=3, column=col_num, value=sub_header)

#     file_path = f'src/output/OCR_Results.xlsx'
#     wb.save(file_path)
#     results.append([file_path])

#     return results






















# def transcription(detected_table_cells, ocr_model):

#     if ocr_model == 'Tesseract-OCR':
#         pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
#     elif ocr_model == 'EasyOCR':
#         easyocr_reader = easyocr.Reader(['en'])

#     contours = detected_table_cells[0]
#     image_with_all_bounding_boxes = detected_table_cells[1]
#     table_copy = detected_table_cells[2]
#     table_original_image = detected_table_cells[3]

#     image_height, image_width, _ = image_with_all_bounding_boxes.shape

#     min_x, min_y, max_x, max_y = get_table_boundaries(contours)
#     table_width = max_x - min_x
#     table_height = max_y - min_y

#     # min_x, min_y, max_x, max_y = 0, 0, image_width, image_height
#     # table_width = max_x - min_x
#     # table_height = max_y - min_y

#     # Load and normalize cell coordinates from the template
#     file_path = 'docs/Table_structure.xlsx'
#     sheet_name = 'clipped_tilted'
#     cell_coordinates = get_excel_cell_coordinates(file_path, sheet_name)
#     sheet_dimensions = {'width': 24, 'height': 50}  # Based on the template
#     normalized_cell_coords = normalize_cell_coordinates(cell_coordinates, sheet_dimensions)

#     # Denormalize cell coordinates to match the table image dimensions
#     denormalized_cell_coords = denormalize_cell_coordinates(normalized_cell_coords, table_width, table_height)

#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'OCR_Results'

#     for cell_ref, (x, y) in denormalized_cell_coords.items():
#         col, row = openpyxl.utils.coordinate_to_tuple(cell_ref)
#         # Calculate the width and height of the cell correctly
#         # Convert to zero-based indexing
#         col -= 1
#         row -= 1

#         # Calculate the width and height of the cell correctly
#         if col < sheet_dimensions['width'] - 1:
#             next_col = col
#             next_x = denormalized_cell_coords.get(openpyxl.utils.get_column_letter(next_col+1) + str(row), (table_width, y))[0]
#             cell_width = next_x - x
#         else:
#             cell_width = table_width / sheet_dimensions['width']

#         if row < sheet_dimensions['height'] - 1:
#             next_row = row
#             next_y = denormalized_cell_coords.get(openpyxl.utils.get_column_letter(col+1) + str(next_row+1), (x, table_height))[1]
#             cell_height = next_y - y
#         else:
#             cell_height = table_height / sheet_dimensions['height']

#         # Draw bounding boxes on the table image based on the denormalized coordinates
#         cv2.rectangle(image_with_all_bounding_boxes, (x, y), (x + int(cell_width), y + int(cell_height)), (0, 255, 0), 2)
    

