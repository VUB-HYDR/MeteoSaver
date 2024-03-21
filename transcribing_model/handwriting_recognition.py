#!/usr/bin/env python
import os, argparse, glob, tempfile, shutil, warnings
import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from paddleocr import PaddleOCR,draw_ocr
from PIL import Image, ImageDraw, ImageFont
import pytesseract
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import easyocr
import tensorflow as tf

#%% OCR
## Lauching Tesseract-OCR
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' ## Here input the PATH to the Tesseract executable on your computer. See more information here: https://pypi.org/project/pytesseract/

## Another option for OCR is PaddleOCR, which would be used by downloading necessary files as shown below
#ocr = PaddleOCR(use_angle_cls=True, lang = 'en', use_gpu=False) ## Run only once to download all required files

## Another option for OCR: EasyOCR
#reader = easyocr.Reader(['en']) # this needs to run only once to load the model into memory

## Another option using Tensorflow and the MNIST dataset
# mnist = tf.keras.datasets.mnist
# (training_data, training_labels), (test_data, test_labels) = mnist.load_data()
# training_data, test_data = training_data / 255, test_data / 255

# model = tf.keras.Sequential([
#     tf.keras.layers.Flatten(input_shape=(28, 28)),
#     tf.keras.layers.Dense(128, activation=tf.nn.relu),
#     tf.keras.layers.Dense(10, activation=tf.nn.softmax)
# ])

# model.compile(optimizer=tf.keras.optimizers.Adam(),
#               loss='sparse_categorical_crossentropy',
#               metrics=['accuracy'])

# model.fit(training_data, training_labels, epochs=5)

# model.evaluate(test_data, test_labels)



# def preprocess_image(image_path):
#     image = Image.open(image_path).convert('L')  # Convert image to grayscale
#     image = image.resize((28, 28))  # Resize image to match the input shape of the model
#     image_array = np.array(image) / 255  # Normalize pixel values
#     return image_array

# # Function to predict handwritten text in the image
# def predict_text(image_path):
#     # Preprocess the input image
#     image_array = preprocess_image(image_path)
#     # Reshape the image array to match the input shape of the model
#     image_input = np.expand_dims(image_array, axis=0)
#     # Use the trained model to make predictions
#     predictions = model.predict(image_input)
#     # Get the index of the predicted class (i.e., the predicted digit)
#     predicted_digit = np.argmax(predictions)
#     return predicted_digit

# # Example usage
# image_path = 'notebooks/clipped_9.png'
# predicted_digit = predict_text(image_path)
# print("Predicted Digit:", predicted_digit)


# import paddle # load paddle module
# import paddle.fluid as fluid

# # As seen on https://github.com/PaddlePaddle/book/blob/develop/02.recognize_digits/README.md 
# def softmax_regression():
#     """
#     Define softmax classifier:
#     A fully connected layer with activation function  softmax
#     Return:
#     predict_image -- result of classification
#     """
#     # input original image data in size of 28*28*1
#     img = fluid.data(name='img', shape=[None, 1, 28, 28], dtype='float32')
#     # With softmax as the fully connected layer of the activation function, the size of the output layer must be 10
#     predict = fluid.layers.fc(
#     input=img, size=10, act='softmax')
#     return predict

# def multilayer_perceptron():
#     """
#     Define multilayer perceptron classifier:
#     Multilayer perceptron with two hidden layers (fully connected layers)
#     The activation function of the first two hidden layers uses ReLU, and the activation function of the output layer uses Softmax.

#     Return:
#     predict_image -- result of classification
#     """
#     # input raw image data in size of 28*28*1
#     img = fluid.data(name='img', shape=[None, 1, 28, 28], dtype='float32')
#     # the first fully connected layer, whose activation function is ReLU
#     hidden = fluid.layers.fc(input=img, size=200, act='relu')
#     # the second fully connected layer, whose activation function is ReLU
#     hidden = fluid.layers.fc(input=hidden, size=200, act='relu')
#     # With softmax as the fully connected output layer of the activation function, the size of the output layer must be 10
#     prediction = fluid.layers.fc(input=hidden, size=10, act='softmax')
#     return prediction

# def conv_pool(input, num_filters, filter_size, pool_size, pool_stride, act="relu"):
#     """
#     Define convolution-pooling layer：
#         Conv_pool layer has a convolutional layer and a pooling layer
#     Args：
#         input —— Input
#         num_filters —— The number of filter
#         filter_size —— The filter size
#         pool_size —— The pool kernel size
#         pool_stride —— The pool stride size
#         act —— Activation type

#     Return:
#         out -- output
#     """
#     conv_out = fluid.layers.conv2d(
#         input=input,
#         num_filters=num_filters,
#         filter_size=filter_size,
#         act=act)
#     out = fluid.layers.pool2d(
#         input=conv_out,
#         pool_size=pool_size,
#         pool_stride=pool_stride)
#     return out

# def convolutional_neural_network():
#     """
#     Define convolutional neural network classifier:
#     The input 2D image passes through two convolution-pooling layers, using the fully connected layer with softmax as the output layer

#     Return:
#     predict -- result of classification
#     """
#     # input raw image data in size of 28*28*1
#     img = fluid.data(name='img', shape=[None, 1, 28, 28], dtype='float32')
#     # the first convolution-pooling layer
#     # Use 20 5*5 filters, the pooling size is 2, the pooling step is 2, and the activation function is Relu.
#     conv_pool_1 = conv_pool(
#         input=img,
#         filter_size=5,
#         num_filters=20,
#         pool_size=2,
#         pool_stride=2,
#         act="relu")
#     conv_pool_1 = fluid.layers.batch_norm(conv_pool_1)
#     # the second convolution-pooling layer
#     # Use 20 5*5 filters, the pooling size is 2, the pooling step is 2, and the activation function is Relu.
#     conv_pool_2 = conv_pool(
#         input=conv_pool_1,
#         filter_size=5,
#         num_filters=50,
#         pool_size=2,
#         pool_stride=2,
#         act="relu")
#     # With softmax as the fully connected output layer of the activation function, the size of the output layer must be 10
#     prediction = fluid.layers.fc(input=conv_pool_2, size=10, act='softmax')
#     return prediction


# def train_program():
#     """
#     Configure train_program

#     Return:
#     predict -- result of classification
#     avg_cost -- mean loss
#     acc -- accuracy of classification

#     """
#     # label layer, called label, correspondent with label category of input picture
#     label = fluid.data(name='label', shape=[None, 1], dtype='int64')

#     # predict = softmax_regression() # cancel note and run Softmax regression
#     # predict = multilayer_perceptron() # cancel note and run multiple perceptron
#     predict = convolutional_neural_network() # cancel note and run LeNet5 convolutional neural network

#     # use class cross-entropy function to compute loss function between predict and label
#     cost = fluid.layers.cross_entropy(input=predict, label=label)
#     # compute mean loss
#     avg_cost = fluid.layers.mean(cost)
#     # compute accuracy of classification
#     acc = fluid.layers.accuracy(input=predict, label=label)
#     return predict, [avg_cost, acc]


# def optimizer_program():
#     return fluid.optimizer.Adam(learning_rate=0.001)

# # there are 64 data in a minibatch
# BATCH_SIZE = 64

# # read 500 data in train dataset, randomly disorder them and then transfer it into batched reader which yield 64 data each time.
# train_reader = paddle.batch(
#         paddle.reader.shuffle(
#             paddle.dataset.mnist.train(), buf_size=500),
#         batch_size=BATCH_SIZE)
# # read data in test dataset and yield 64 data every time
# test_reader = paddle.batch(
#             paddle.dataset.mnist.test(), batch_size=BATCH_SIZE)

# def event_handler(pass_id, batch_id, cost):
#     # print the intermediate results of training, like
#     # training iterations, number of batch, and loss function
#     print("Pass %d, Batch %d, Cost %f" % (pass_id,batch_id, cost))

# from paddle.utils.plot import Ploter

# train_prompt = "Train cost"
# test_prompt = "Test cost"
# cost_ploter = Ploter(train_prompt, test_prompt)

# # visualize training process
# def event_handler_plot(ploter_title, step, cost):
#     cost_ploter.append(ploter_title, step, cost)
#     cost_ploter.plot()

# # the model is run on single CPU
# use_cuda = False # If you want to use GPU, please set it True
# place = fluid.CUDAPlace(0) if use_cuda else fluid.CPUPlace()

# # call train_program to get prediction value and loss value,
# prediction, [avg_loss, acc] = train_program()

# # input original image data in size of 28*28*1
# # label layer, called label, correspondent with label category of input picture.

# # It is informed that data in network consists of two parts. One is img value, the other is label value.
# feeder = fluid.DataFeeder(feed_list=['img', 'label'], place=place)

# # choose Adam optimizer
# optimizer = tf.keras.optimizers.Adam()
# #optimizer.minimize(avg_loss)

# PASS_NUM = 5 #train 5 iterations
# epochs = [epoch_id for epoch_id in range(PASS_NUM)]

# # save parameters of model into save_dirname file
# save_dirname = "recognize_digits.inference.model"

# def train_test(train_test_program,
#                    train_test_feed, train_test_reader):

#     # save classification accuracy into acc_set
#     acc_set = []
#     # save mean loss in avg_loss_set
#     avg_loss_set = []
#     # transfer each data which is the output of testing reader_yield into network to train
#     for test_data in train_test_reader():
#         acc_np, avg_loss_np = exe.run(
#             program=train_test_program,
#             feed=train_test_feed.feed(test_data),
#             fetch_list=[acc, avg_loss])
#         acc_set.append(float(acc_np))
#         avg_loss_set.append(float(avg_loss_np))
#     # get accuracy and loss value on the test data
#     acc_val_mean = numpy.array(acc_set).mean()
#     avg_loss_val_mean = numpy.array(avg_loss_set).mean()
#     # return mean loss value and mean accuracy
#     return avg_loss_val_mean, acc_val_mean

# from __future__ import print_function #load print of python3 into current version

# exe = fluid.Executor(place)

# main_program = fluid.default_main_program()
# test_program = fluid.default_main_program().clone(for_test=True)

# lists = []
# step = 0
# for epoch_id in epochs:
#     for step_id, data in enumerate(train_reader()):
#         metrics = exe.run(main_program,
#                           feed=feeder.feed(data),
#                           fetch_list=[avg_loss, acc])
#         if step % 100 == 0: # print a log for every 100 times of training
#             print("Pass %d, Batch %d, Cost %f" % (step, epoch_id, metrics[0]))
#             event_handler_plot(train_prompt, step, metrics[0])
#         step += 1

#     # test classification result of each epoch
#     avg_loss_val, acc_val = train_test(train_test_program=test_program,
#                                        train_test_reader=test_reader,
#                                        train_test_feed=feeder)

#     print("Test with Epoch %d, avg_cost: %s, acc: %s" %(epoch_id, avg_loss_val, acc_val))
#     event_handler_plot(test_prompt, step, metrics[0])

#     lists.append((epoch_id, avg_loss_val, acc_val))

#     # save parameters of trained model for prediction
#     if save_dirname is not None:
#         fluid.io.save_inference_model(save_dirname,
#                                       ["img"], [prediction], exe,
#                                       model_filename=None,
#                                       params_filename=None)

# # Choose the best pass
# best = sorted(lists, key=lambda list: float(list[1]))[0]
# print('Best pass is %s, testing Avgcost is %s' % (best[0], best[1]))
# print('The classification accuracy is %.2f%%' % (float(best[2]) * 100))


#%% FUNCTIONS

## Function for detecting the table and the text within the table
def table_detection_model(image_path):

    ## Read Image from the given image path
    original_image  = cv2.imread(image_path)
    # Image in grayscale
    image = cv2.cvtColor(original_image, cv2.COLOR_BGR2GRAY)



    ## Pre-processing image to detect the table from the record sheets
    # Here, the threshold value for pixel intensities = 0, and the value 255 is assigned if the pixel value is above the threshold
    thresh = cv2.threshold(image,0,255,cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1] 
    # Perform morphological operations (like dilation and erosion) for better segmentation
    kernel = np.ones((5,5),np.uint8)
    thresh = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
    # Find contours
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    # Minimum dimensions of table
    threshold_area = 4  # Minimum contour area to consider as a cell
    threshold_height = 2   # Minimum height of the cell
    threshold_width = 2    # Minimum width of the cell
    # Initialize variables for the largest contour
    largest_contour_area = 0
    largest_contour = None
    # Filter and extract individual cells, focusing on the largest contour
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        # Filter out small contours or undesired regions based on area or aspect ratio
        if cv2.contourArea(contour) > threshold_area and h > threshold_height and w > threshold_width:
            # Find the largest contour by area
            contour_area = cv2.contourArea(contour)
            if contour_area > largest_contour_area:
                largest_contour_area = contour_area
                largest_contour = contour
    # Draw bounding box for the largest contour (if found), which here represents the table on the record sheets
    if largest_contour is not None:
        x, y, w, h = cv2.boundingRect(largest_contour)
        cv2.rectangle(image, (x, y), (x + w, y + h), (0, 255, 0), 2)
        table = image[y:y + h, x:x + w] # clip out the table (here, the largest contour) from the original image.
    else:
        table = image # Incase the main table is not detected as the largest contour, we just use the original image/ whole record sheet as the image with the table



    ## Detecting the vertical and horizontal (both dotted and bold) in the table
    # Thresholding to reduce the image to black or white pixels
    table_img_bin = cv2.adaptiveThreshold(table, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 91,6)
    # Save the binary image for use later in detecting text
    cv2.imwrite('table_binarized.jpg', table_img_bin)
    #thresh,img_bin = cv2.threshold(table,100,255,cv2.THRESH_BINARY)
    img_bin = 255-table_img_bin
    # Detect the vertical lines in the image
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, np.array(table).shape[1]//50)) # The '//50' divides the length of the array (table) by 50, likely to obtain a fraction of the length for the structuring element,
    eroded_image = cv2.erode(img_bin, vertical_kernel, iterations=1)
    vertical_lines = cv2.dilate(eroded_image, vertical_kernel, iterations=5)
    # Detect the horizontal lines in the image
    hor_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (np.array(table).shape[1]//20, 1)) # The '//20' divides the width of the array (table) by 20, likely to obtain a fraction of the width for the structuring element.
    eroded_image= cv2.erode(img_bin, hor_kernel, iterations=1)
    horizontal_lines = cv2.dilate(eroded_image, hor_kernel, iterations=5)
    # Blending the imaegs with the vertical lines and the horizontal lines 
    combined_vertical_and_horizontal_lines = cv2.addWeighted(vertical_lines, 1, horizontal_lines, 1, 1)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    combined_image_dilated = cv2.dilate(combined_vertical_and_horizontal_lines, kernel, iterations=5)
    # Remove the lines from the image (table)
    image_without_lines = cv2.subtract(img_bin, combined_image_dilated)
    # Remove smaller 'still-visible' lines through noise removal
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    image_without_lines_noise_removed = cv2.erode(image_without_lines, kernel, iterations=1)
    image_without_lines_noise_removed = cv2.dilate(image_without_lines_noise_removed, kernel, iterations=1)
    # Convert words into blobs using dilation
    kernel_to_remove_gaps_between_words = np.array([
            [1,1,1,1,1],
            [1,1,1,1,1]])
    image_with_word_blobs = cv2.dilate(image_without_lines_noise_removed, kernel_to_remove_gaps_between_words, iterations=5)
    simple_kernel = np.ones((3,3), np.uint8)
    image_with_word_blobs = cv2.dilate(image_with_word_blobs, simple_kernel, iterations=1)
    # Detecting the dotted lines using horizontal line detection and erosion. ### ADDITIONAL STEP: This is because the original images ahve dotted horizontal lines which cvan still be detected after the first removal of main (undotted) horizontal lines
    hor_kernel_2 = cv2.getStructuringElement(cv2.MORPH_RECT, (np.array(image_with_word_blobs).shape[1]//20, 1))
    image_3 = cv2.erode(image_with_word_blobs, hor_kernel_2, iterations=1)
    horizontal_lines_2 = cv2.dilate(image_3, hor_kernel_2, iterations=1)
    # Removing the dotted liens by substracting them from the original image 
    image_without_lines_2 = cv2.subtract(image_with_word_blobs, horizontal_lines_2)



    ## Using contours in order to detect text in the table after removing the vertical and horizontal lines
    # Assuming 'table' is your input image in BGR format
    result = cv2.findContours(image_without_lines_2, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    contours = result[0]
    # Original image of table in binarizesd format
    image_with_all_bounding_boxes = cv2.imread('table_binarized.jpg')
    table_copy = image_with_all_bounding_boxes.copy()
    # Get the dimensions of the loaded image
    image_height, image_width, image_channels = image_with_all_bounding_boxes.shape


    ## Create an Excel workbook and add a worksheet where the transcribed text will be saved
    wb = Workbook()
    ws = wb.active
    ws.title = 'OCR_Results'


    ## Text detection using an OCR model; Here using TesseractOCR
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        
        # Adjust these threshold values according to your requirements
        min_width_threshold = 20
        min_height_threshold = 10  # Had this at 10 previously

        # Filter out smaller bounding boxes
        if w >= min_width_threshold and h >= min_height_threshold:
            
            # Calculate a factor to increase the bounding box area (e.g., 80% larger)
            increase_factor_width = 0.1  # Modify this factor as needed
            increase_factor_height = 0.5 # Modify this factor as needed
        
            x -= int(w * increase_factor_width)  # Increase width
            y -= int(h * increase_factor_height)  # Increase height
            w += int(2 * w * increase_factor_width)  # Increase width
            h += int(2 * h * increase_factor_height)  # Increase height
            
            
            #bounding_boxes.append((x, y, w, h))
            # This line below is about
            # drawing a rectangle on the image with the shape of
            # the bounding box. Its not needed for the OCR.
            # Its just added for debugging purposes.
            image_with_all_bounding_boxes = cv2.rectangle(image_with_all_bounding_boxes, (x, y), (x + w, y + h), (0, 255, 0), 5)
            
            # Calculate center coordinates of the bounding box
            center_x = x + w // 2
            center_y = y + h // 2
            
            # OCR
            # Crop each cell using the bounding rectangle coordinates
            ROI = table_copy[y:y+h, x:x+w]
            
            if ROI.size != 0:  # Check if the height and width are greater than zero
                
                cv2.imwrite('detected.png', ROI)

                # Using Tesseract-OCR
                ocr_result = pytesseract.image_to_string('detected.png', lang='cobedore-V6', config='--psm 7')
                #ocr_result = pytesseract.image_to_string('detected.png', lang='cobedore-V6', config='--psm 6 -c tessedit_char_whitelist=0123456789 ')
                if ocr_result is not None:
                    
                    # Maximum number of columns and rows. These can be changed depending on the tables in the images
                    max_column_index = 26  # Number of columns in the table. 
                    max_row_index = 57  # Estimated number of rows in the table  . Initital runs had 56             
                    
                    cell_width = max(image_width // max_column_index, min_width_threshold)
                    cell_height = max(image_height//max_row_index, min_height_threshold)
                    
                    # Track filled cells using a set
                    filled_cells = []
                    
                    
                    # Convert the x-coordinate to a column letter # Assuming x-coordinate translates to columns cell_column
                    # Ensure x is within the valid range for Excel column indices
                    
                    if 1 <= center_x <= image_width:  # Excel's maximum column index
                        column_index = max((center_x) // cell_width, 0) + 1 # Ensure column index is at least 1
                        #cell_column = openpyxl.utils.get_column_letter(min(column_index, max_column_index))
                    else:
                        column_index = 0
                        #cell_column = 'A'  # Set a default column if x is out of range
                        
                    if 1 <= center_y <= image_height:
                        row_ratio = (center_y) // cell_height
                        cell_row = min(int(row_ratio) + 1, max_row_index)  # Convert row ratio to integer and ensure it's within valid range
                        #cell_row = min(center_y // cell_height + 0.5, max_row_index)
                    else:
                        cell_row = 1  # Set a default row if y is out of range  # Assuming y-coordinate translates to rows
                
                    
                    cell_key = (cell_row, column_index)      
                    # Check if the cell is already filled
                    if cell_key in filled_cells:
                        cell_row = cell_row + 1  # Move to the next row if the cell is filled
                        cell_key = (cell_row, column_index)  # Update the cell key
                        filled_cells.append(cell_key)  # Add the filled cell coordinates to the set
                    else:
                        filled_cells.append(cell_key)  # Still add the filled cell coordinates to the set

                    # Write the OCR value to the cell in the Excel sheet
                    cell = ws.cell(row=cell_row, column=column_index)
                    cell.value = ocr_result
                
                    # Set up border styles for excel output
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

                    # Loop through cells to apply borders
                    for row in ws.iter_rows(min_row=1, max_row=max_row_index, min_col=1, max_col=max_column_index):
                        for cell in row:
                            cell.border = thin_border

                    ## Using PaddleOCR
                    # ocr_result = ocr.ocr('detected.png', cls = True)
                    # if ocr_result is not None and ocr_result[0] is not None:
                    #     cell_values = ocr_result[0]
                    #     text_in_cells = [line[1][0] for line in cell_values]
                        
                    #     boxes = [line[0] for line in result]
                        
                    #     # Maximum number of columns and rows. These can be changed depending on the tables in the images
                    #     max_column_index = 26  # Number of columns in the table. 
                    #     max_row_index = 57  # Estimated number of rows in the table  . Initital runs had 56             
                        
                    #     cell_width = max(image_width // max_column_index, min_width_threshold)
                    #     cell_height = max(image_height//max_row_index, min_height_threshold)
                        
                    #     # Track filled cells using a set
                    #     filled_cells = []
                        
                    #     for text in range(len(text_in_cells)):
                    #         value =  text_in_cells[text]

                    #         # Convert the x-coordinate to a column letter # Assuming x-coordinate translates to columns cell_column
                    #         # Ensure x is within the valid range for Excel column indices
                            
                    #         if 1 <= center_x <= image_width:  # Excel's maximum column index
                    #             column_index = max((center_x) // cell_width, 0) + 1 # Ensure column index is at least 1
                    #             #cell_column = openpyxl.utils.get_column_letter(min(column_index, max_column_index))
                    #         else:
                    #             column_index = 0
                    #             #cell_column = 'A'  # Set a default column if x is out of range
                                
                    #         if 1 <= center_y <= image_height:
                    #             row_ratio = (center_y) // cell_height
                    #             cell_row = min(int(row_ratio) + 1, max_row_index)  # Convert row ratio to integer and ensure it's within valid range
                    #             #cell_row = min(center_y // cell_height + 0.5, max_row_index)
                    #         else:
                    #             cell_row = 1  # Set a default row if y is out of range  # Assuming y-coordinate translates to rows
                        
                            
                    #         cell_key = (cell_row, column_index)      
                    #         # Check if the cell is already filled
                    #         if cell_key in filled_cells:
                    #             cell_row = cell_row + 1  # Move to the next row if the cell is filled
                    #             cell_key = (cell_row, column_index)  # Update the cell key
                    #             filled_cells.append(cell_key)  # Add the filled cell coordinates to the set
                    #         else:
                    #             filled_cells.append(cell_key)  # Still add the filled cell coordinates to the set

                    #         # Write the OCR value to the cell in the Excel sheet
                    #         cell = ws.cell(row=cell_row, column=column_index)
                    #         cell.value = value
                        
                    #         # Set up border styles for excel output
                    #         thin_border = Border(
                    #             left=Side(style='thin'),
                    #             right=Side(style='thin'),
                    #             top=Side(style='thin'),
                    #             bottom=Side(style='thin'))

                    #         # Loop through cells to apply borders
                    #         for row in ws.iter_rows(min_row=1, max_row=max_row_index, min_col=1, max_col=max_column_index):
                    #             for cell in row:
                    #                 cell.border = thin_border
                
                

                # Using EasyOCR
                # ocr_result = reader.readtext('detected.png', detail = 0, allowlist='0123456789')
                # if len(ocr_result) != 0:
                #     #cell_values = ocr_result[0]
                #     text_in_cells = ocr_result
                    
                #     #boxes = [line[0] for line in result]
                    
                #     # Maximum number of columns and rows. These can be changed depending on the tables in the images
                #     max_column_index = 26  # Number of columns in the table. 
                #     max_row_index = 57  # Estimated number of rows in the table  . Initital runs had 56             
                    
                #     cell_width = max(image_width // max_column_index, min_width_threshold)
                #     cell_height = max(image_height//max_row_index, min_height_threshold)
                    
                #     # Track filled cells using a set
                #     filled_cells = []
                    
                #     for text in range(len(text_in_cells)):
                #         value =  text_in_cells[text]

                #         # Convert the x-coordinate to a column letter # Assuming x-coordinate translates to columns cell_column
                #         # Ensure x is within the valid range for Excel column indices
                        
                #         if 1 <= center_x <= image_width:  # Excel's maximum column index
                #             column_index = max((center_x) // cell_width, 0) + 1 # Ensure column index is at least 1
                #             #cell_column = openpyxl.utils.get_column_letter(min(column_index, max_column_index))
                #         else:
                #             column_index = 0
                #             #cell_column = 'A'  # Set a default column if x is out of range
                            
                #         if 1 <= center_y <= image_height:
                #             row_ratio = (center_y) // cell_height
                #             cell_row = min(int(row_ratio) + 1, max_row_index)  # Convert row ratio to integer and ensure it's within valid range
                #             #cell_row = min(center_y // cell_height + 0.5, max_row_index)
                #         else:
                #             cell_row = 1  # Set a default row if y is out of range  # Assuming y-coordinate translates to rows
                    
                        
                #         cell_key = (cell_row, column_index)      
                #         # Check if the cell is already filled
                #         if cell_key in filled_cells:
                #             cell_row = cell_row + 1  # Move to the next row if the cell is filled
                #             cell_key = (cell_row, column_index)  # Update the cell key
                #             filled_cells.append(cell_key)  # Add the filled cell coordinates to the set
                #         else:
                #             filled_cells.append(cell_key)  # Still add the filled cell coordinates to the set

                #         # Write the OCR value to the cell in the Excel sheet
                #         cell = ws.cell(row=cell_row, column=column_index)
                #         cell.value = value
                    
                #         # Set up border styles for excel output
                #         thin_border = Border(
                #             left=Side(style='thin'),
                #             right=Side(style='thin'),
                #             top=Side(style='thin'),
                #             bottom=Side(style='thin'))

                #         # Loop through cells to apply borders
                #         for row in ws.iter_rows(min_row=1, max_row=max_row_index, min_col=1, max_col=max_column_index):
                #             for cell in row:
                #                 cell.border = thin_border

                else:
                    print('No values detected in clip')
            else:
                print('ROI is empty or invalid')
            
    wb.save('Excel_with_OCR_Results.xlsx') 

    return wb



#%% SETTING UP THE CURRENT WORKING DIRECTORY; for both the input and output folders
cwd = os.getcwd()

images_folder = os.path.join(cwd, 'data') #folder containing all images
sample_images = os.path.join(images_folder, '10_sample_different_images') # sample images

# TRIAL ON ONE TEST IMAGE FROM THE FOLDER
## This will be replace with a 'for' loop after testing all the functions
one_test_image =  os.path.join(sample_images, '203_196503_SF_YAN.JPG')
image_read = cv2.imread(one_test_image)
plt.imshow(image_read)
plt.show()

table_detection_model(one_test_image)

  
ocr_result = reader.readtext('notebooks/clipped100.png')













